# -------------------------------------------------------------------------
# FEMMA 12.0 - BiH Edition (Modern UI)
# Full Application Code - Imports, Globals, Pricing
# -------------------------------------------------------------------------

import os
import sys
import re
import math
import json
import copy
import unicodedata
import numpy as np
from datetime import datetime, date
import xml.etree.ElementTree as ET
import urllib.request
import urllib.error
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
import pandas as pd
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import fills as fills_mod
from openpyxl.chart import BarChart, Reference

# -------------------------------------------------------------------------
# GUI THEME
# -------------------------------------------------------------------------

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

ROSEGOLD = "#D4AF37"
ROSEGOLD_DARK = "#B9962F"
TEXT_MUTED = "#C8C8C8"

# -------------------------------------------------------------------------
# GLOBAL STATE
# -------------------------------------------------------------------------

file_path = ""  # aktivni .xlsx fajl (prodaja/nabavka)
output_file_path = ""  # opcioni output fajl za izvjestaje
pantheon_xls_path = ""  # Pantheon .xls
promet_sales_path = ""  # ukupna prodaja (Pantheon .xlsx)
promet_stanje_path = ""  # trenutno stanje (Pantheon .xlsx)
promet_output_path = ""  # output fajl za proracun prometa i zaliha


def app_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


KALK_NABAVNE_PATH = os.path.join(app_base_dir(), "nabavnecijene_kalkulacije.json")
KALK_FILE_CACHE_PATH = os.path.join(app_base_dir(), "kalk_file_cache.json")
DEFAULT_KALK_FOLDER = os.path.join(app_base_dir(), "Kalkulacije")
KALK_NABAVNE_BY_SKU = {}
KALK_MPC_WO_BY_SKU = {}
KALK_MPC_W_BY_SKU = {}
KALK_VAT_BY_SKU = {}


def _set_output_file_path(p):
    global output_file_path
    output_file_path = p or ""
    label_text = (
        "Output fajl: (isti kao input)"
        if not output_file_path
        else f"Output fajl: {os.path.basename(output_file_path)}"
    )
    if "output_file_label_sales" in globals():
        output_file_label_sales.configure(text=label_text)
    if "output_file_label_proc" in globals():
        output_file_label_proc.configure(text=label_text)


def _set_promet_output_path(p):
    global promet_output_path
    promet_output_path = p or ""
    label_text = (
        "Output fajl: (automatski)"
        if not promet_output_path
        else f"Output fajl: {os.path.basename(promet_output_path)}"
    )
    if "output_file_label_promet" in globals():
        output_file_label_promet.configure(text=label_text)


def backup_kalk_db():
    if not os.path.exists(KALK_NABAVNE_PATH):
        return
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = os.path.join(app_base_dir(), "db_backup")
    try:
        os.makedirs(backup_dir, exist_ok=True)
    except Exception:
        return
    backup_path = os.path.join(
        backup_dir, f"nabavnecijene_kalkulacije_backup_{stamp}.json"
    )
    try:
        with open(KALK_NABAVNE_PATH, "r", encoding="utf-8") as f_src:
            content = f_src.read()
        with open(backup_path, "w", encoding="utf-8") as f_dst:
            f_dst.write(content)
    except Exception:
        pass


def load_kalk_file_cache():
    if not os.path.exists(KALK_FILE_CACHE_PATH):
        return {}
    try:
        with open(KALK_FILE_CACHE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    return data


def save_kalk_file_cache(cache):
    try:
        with open(KALK_FILE_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2, ensure_ascii=True)
    except Exception:
        pass


def _default_output_name():
    if file_path:
        base = os.path.splitext(os.path.basename(file_path))[0]
        return f"{base}_output.xlsx"
    return "output.xlsx"


def _default_promet_output_name(base_path: str | None):
    if base_path:
        base = os.path.splitext(os.path.basename(base_path))[0]
        return f"{base}_promet_zaliha.xlsx"
    return "promet_zaliha.xlsx"

# -------------------------------------------------------------------------
# PODESAVANJA (DEFAULTI + UČITAVANJE)
# -------------------------------------------------------------------------

SETTINGS_PATH = os.path.join(app_base_dir(), "femma_settings.json")
MP_CJENOVNIK_PATH = os.path.join(app_base_dir(), "mp_cjenovnik.json")
META_TMP_PATH = os.path.join(app_base_dir(), "meta_period_tmp.json")

DEFAULT_PREFIX_MAP = {
    "AF-": "Afro rep",
    "RR-": "Ravni rep",
    "OPK-": "Repovi OPK",
    "AR-": "Ariana repovi",
    "KRR-": "Kratki repovi",
    "KRO-": "Kratki repovi",
    "KRA-": "Kratki repovi",
    "TRK-": "Repovi trakica",
    "DR-": "Dugi repovi",
    "U-": "U klipse",
    "BD-": "Blowdry klipse",
    "BDR-": "Blowdry repovi",
    "EKS-": "Ekstenzije",
    "EKSOPK": "Ekstenzije OPK",
    "SIS-": "Siske",
    "P0": "Klasične perike",
    "PR": "Premium perike",
}
DEFAULT_USD_KM_RATE = 0.0
DEFAULT_NET_MARGIN_TRANSPORT_USD = 0.0
DEFAULT_NET_MARGIN_CUSTOMS_PCT = 0.0
DEFAULT_NET_MARGIN_MARKETING_KM = 0.0
DEFAULT_NET_MARGIN_SPACE_KM = 1300.0
DEFAULT_NET_MARGIN_LABOR_KM = 15000.0
DEFAULT_NET_MARGIN_ACCOUNTING_KM = 350.0
DEFAULT_NET_MARGIN_UTILITIES_KM = 300.0

DEFAULT_CUSTOM_SKU_LIST = [
    "U-10",
    "U-1B",
    "U-2",
    "U-350",
    "U-4",
    "U-6",
    "U-613",
    "U-K16",
    "U-PLAT",
    "U-SMEDPR",
    "U-SMO",
    "U-SRP",
    "U-VIOLET",
    "AF-SMEDPR",
    "AF-PLAT",
    "AR-SMEDPR",
    "AR-MHG",
    "AR-10",
    "AR-6",
    "AR-SRP",
    "BD-10",
    "BD-1B",
    "BD-2",
    "BD-4",
    "BD-6",
    "BD-613",
    "BD-K16",
    "BD-PLAT",
    "BD-SMEDPR",
    "BD-SRP",
    "EKS-MGH",
    "EKS-PLAT",
    "EKS-SMEDPR",
    "KRO-SMEDPR",
    "KRO-PLAT",
    "KRR-10",
    "KRR-PLAT",
    "KRR-SRP",
    "OPK-8",
    "OPK-PLAT",
    "OPK-SMEDPR",
    "OPK-SRP",
    "RR-8",
    "RR-PLAT",
    "RR-SMEDPR",
    "BDR-1B",
    "BDR-2",
    "BDR-4",
    "BDR-6",
    "BDR-SRP",
    "BDR-k16",
    "BDR-613",
    "BDR-SMEDPR",
]
DEFAULT_SKU_CATEGORY_OVERRIDES = {}
DEFAULT_CIJENE_PREFIXI = {
    "RR-": 6.50,
    "AF-": 7.50,
    "OPK-": 7.00,
    "KRR-": 6.00,
    "KRO-": 7.00,
    "KRA-": 6.00,
    "AR-": 6.00,
    "TRK-": 6.00,
    "DR-": 8.00,
    "EKS-": 7.00,
    "EKSOPK": 8.00,
    "U-": 9.50,
    "BD-": 9.00,
    "P0": 8.50,
    "PR": 31.00,
    "BDR-": 7.50,
}


def _normalize_sku_list(values):
    result = []
    for v in values or []:
        if isinstance(v, str):
            v = v.strip().upper()
            if v:
                result.append(v)
    return result


def _normalize_prefix_map(values):
    result = {}
    for k, v in (values or {}).items():
        if not isinstance(k, str) or not isinstance(v, str):
            continue
        key = k.strip().upper()
        val = v.strip()
        if key and val:
            result[key] = val
    return result


def _normalize_overrides(values):
    result = {}
    for k, v in (values or {}).items():
        if not isinstance(k, str) or not isinstance(v, str):
            continue
        key = k.strip().upper()
        val = v.strip()
        if key and val:
            result[key] = val
    return result


def _normalize_price_map(values):
    result = {}
    for k, v in (values or {}).items():
        if not isinstance(k, str):
            continue
        key = k.strip().upper()
        if not key:
            continue
        try:
            price = float(v)
        except Exception:
            continue
        result[key] = float(price)
    return result


def _normalize_kalk_offsets(values):
    defaults = {
        "sku": 0,
        "nabavna": 0,
        "mpc_wo": 0,
        "mpc_w": 0,
        "pdv_rate": 0,
        "kolicina": 0,
    }
    if not isinstance(values, dict):
        return defaults
    result = defaults.copy()
    for key in defaults.keys():
        if key not in values:
            continue
        try:
            val = int(values.get(key, 0))
        except Exception:
            continue
        val = max(-2, min(2, val))
        result[key] = val
    return result


def load_settings():
    prefix = DEFAULT_PREFIX_MAP.copy()
    custom_list = list(DEFAULT_CUSTOM_SKU_LIST)
    overrides = DEFAULT_SKU_CATEGORY_OVERRIDES.copy()
    price_prefix = DEFAULT_CIJENE_PREFIXI.copy()
    price_sku = {}
    mp_category_prices = {}
    usd_bam_rate = DEFAULT_USD_KM_RATE
    net_transport_usd = DEFAULT_NET_MARGIN_TRANSPORT_USD
    net_customs_pct = DEFAULT_NET_MARGIN_CUSTOMS_PCT
    net_marketing_bam = DEFAULT_NET_MARGIN_MARKETING_KM
    net_space_bam = DEFAULT_NET_MARGIN_SPACE_KM
    net_labor_bam = DEFAULT_NET_MARGIN_LABOR_KM
    net_accounting_bam = DEFAULT_NET_MARGIN_ACCOUNTING_KM
    net_utilities_bam = DEFAULT_NET_MARGIN_UTILITIES_KM
    kalk_folder = DEFAULT_KALK_FOLDER
    kalk_processed_files = {}
    kalk_last_offsets = {
        "sku": 0,
        "nabavna": 0,
        "mpc_wo": 0,
        "mpc_w": 0,
        "pdv_rate": 0,
        "kolicina": 0,
    }

    if os.path.exists(SETTINGS_PATH):
        try:
            with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            prefix.update(_normalize_prefix_map(data.get("prefix_map")))
            custom_list = _normalize_sku_list(data.get("custom_skus", custom_list))
            overrides = _normalize_overrides(data.get("sku_category_overrides", overrides))
            price_prefix.update(
                _normalize_price_map(data.get("purchase_prices_by_prefix"))
            )
            price_sku = _normalize_price_map(data.get("purchase_price_overrides"))
            mp_category_prices = _normalize_price_map(
                data.get("mp_prices_by_category")
            )
            try:
                usd_bam_rate = float(data.get("usd_bam_rate", usd_bam_rate))
            except Exception:
                usd_bam_rate = usd_bam_rate
            try:
                net_transport_usd = float(
                    data.get("net_margin_transport_usd", net_transport_usd)
                )
            except Exception:
                net_transport_usd = net_transport_usd
            try:
                net_customs_pct = float(
                    data.get("net_margin_customs_pct", net_customs_pct)
                )
            except Exception:
                net_customs_pct = net_customs_pct
            try:
                net_marketing_bam = float(
                    data.get("net_margin_marketing_bam", net_marketing_bam)
                )
            except Exception:
                net_marketing_bam = net_marketing_bam
            try:
                net_space_bam = float(
                    data.get("net_margin_space_bam", net_space_bam)
                )
            except Exception:
                net_space_bam = net_space_bam
            try:
                net_labor_bam = float(
                    data.get("net_margin_labor_bam", net_labor_bam)
                )
            except Exception:
                net_labor_bam = net_labor_bam
            try:
                net_accounting_bam = float(
                    data.get("net_margin_accounting_bam", net_accounting_bam)
                )
            except Exception:
                net_accounting_bam = net_accounting_bam
            try:
                net_utilities_bam = float(
                    data.get("net_margin_utilities_bam", net_utilities_bam)
                )
            except Exception:
                net_utilities_bam = net_utilities_bam
            try:
                kalk_folder = str(data.get("kalkulacije_folder", kalk_folder))
            except Exception:
                kalk_folder = kalk_folder
            try:
                kalk_processed_files = data.get("kalk_processed_files", {})
                if not isinstance(kalk_processed_files, dict):
                    kalk_processed_files = {}
            except Exception:
                kalk_processed_files = {}
            try:
                kalk_last_offsets = _normalize_kalk_offsets(
                    data.get("kalk_last_offsets", kalk_last_offsets)
                )
            except Exception:
                kalk_last_offsets = kalk_last_offsets
        except Exception as e:
            print(f"Upozorenje: ne mogu ucitati {SETTINGS_PATH}: {e}")

    return (
        prefix,
        custom_list,
        overrides,
        price_prefix,
        price_sku,
        mp_category_prices,
        usd_bam_rate,
        net_transport_usd,
        net_customs_pct,
        net_marketing_bam,
        net_space_bam,
        net_labor_bam,
        net_accounting_bam,
        net_utilities_bam,
        kalk_folder,
        kalk_processed_files,
        kalk_last_offsets,
    )


def save_settings():
    data = {
        "prefix_map": prefix_map,
        "custom_skus": sorted(CUSTOM_SKU_SET),
        "sku_category_overrides": SKU_CATEGORY_OVERRIDES,
        "purchase_prices_by_prefix": PRICE_BY_PREFIX,
        "purchase_price_overrides": PRICE_BY_SKU,
        "mp_prices_by_category": MP_PRICE_BY_CATEGORY,
        "usd_bam_rate": usd_bam_rate,
        "net_margin_transport_usd": net_margin_transport_usd,
        "net_margin_customs_pct": net_margin_customs_pct,
        "net_margin_marketing_bam": net_margin_marketing_bam,
        "net_margin_space_bam": net_margin_space_bam,
        "net_margin_labor_bam": net_margin_labor_bam,
        "net_margin_accounting_bam": net_margin_accounting_bam,
        "net_margin_utilities_bam": net_margin_utilities_bam,
        "kalkulacije_folder": kalkulacije_folder,
        "kalk_processed_files": KALK_PROCESSED_FILES,
        "kalk_last_offsets": KALK_LAST_OFFSETS,
    }
    try:
        with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=True)
    except Exception as e:
        messagebox.showerror("Greska", f"Ne mogu sacuvati podesavanja: {e}")


def write_meta_tmp(start_date, end_date, period_days):
    data = {
        "SalesPeriodStart": start_date.strftime("%Y-%m-%d"),
        "SalesPeriodEnd": end_date.strftime("%Y-%m-%d"),
        "SalesPeriodDays": int(period_days),
    }
    try:
        with open(META_TMP_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=True)
    except Exception:
        pass


def read_meta_tmp():
    if not os.path.exists(META_TMP_PATH):
        return None
    try:
        with open(META_TMP_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


(
    prefix_map,
    CUSTOM_SKU_LIST,
    SKU_CATEGORY_OVERRIDES,
    PRICE_BY_PREFIX,
    PRICE_BY_SKU,
    MP_PRICE_BY_CATEGORY,
    usd_bam_rate,
    net_margin_transport_usd,
    net_margin_customs_pct,
    net_margin_marketing_bam,
    net_margin_space_bam,
    net_margin_labor_bam,
    net_margin_accounting_bam,
    net_margin_utilities_bam,
    kalkulacije_folder,
    KALK_PROCESSED_FILES,
    KALK_LAST_OFFSETS,
) = load_settings()
if prefix_map.get("PR") == "Perike":
    prefix_map["PR"] = "Premium perike"
if prefix_map.get("P0") == "Perike P0":
    prefix_map["P0"] = "Klasične perike"
if "Perike" in MP_PRICE_BY_CATEGORY and "Premium perike" not in MP_PRICE_BY_CATEGORY:
    MP_PRICE_BY_CATEGORY["Premium perike"] = MP_PRICE_BY_CATEGORY["Perike"]
if "Perike P0" in MP_PRICE_BY_CATEGORY and "Klasične perike" not in MP_PRICE_BY_CATEGORY:
    MP_PRICE_BY_CATEGORY["Klasične perike"] = MP_PRICE_BY_CATEGORY["Perike P0"]
CUSTOM_SKU_SET = {s.upper() for s in CUSTOM_SKU_LIST}

# -------------------------------------------------------------------------
# MP CIJENOVNIK (KM) — OVO JE PUN DICT IZ ŠIFRARNIKA
# -------------------------------------------------------------------------

# -------------------------------------------------------------------------
# MP CIJENOVNIK (KM) — učitavanje iz JSON fajla
# -------------------------------------------------------------------------

try:
    with open(MP_CJENOVNIK_PATH, "r", encoding="utf-8") as f:
        mp_cijene_raw = json.load(f)
        mp_cijene = {str(k).upper(): v for k, v in mp_cijene_raw.items()}
except FileNotFoundError:
    mp_cijene = {}
    print("Upozorenje: mp_cjenovnik.json nije pronađen! MP cijene će biti 0.")

for sku in CUSTOM_SKU_SET:
    mp_cijene.setdefault(sku, 0.0)


def save_mp_cjenovnik():
    data = {str(k).upper(): float(v) for k, v in mp_cijene.items()}
    try:
        with open(MP_CJENOVNIK_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=True)
    except Exception as e:
        messagebox.showerror("Greska", f"Ne mogu sacuvati mp_cjenovnik.json: {e}")
# -------------------------------------------------------------------------
# TEŽINE PO PREFIKSIMA (grami)
# -------------------------------------------------------------------------

tezine_prefiksi = {
    "RR-": 155,
    "OPK-": 155,
    "AF-": 155,
    "KRR-": 140,
    "KRO-": 140,
    "KRA-": 140,
    "AR-": 140,
    "TRK-": 155,
    "DR-": 180,
    "U-": 240,
    "BD-": 240,
    "BDR-": 200,
    "EKS-": 170,
    "EKSOPK": 170,
    "P0": 235,
    "PR": 210,
}

DEFAULT_TEZINA = 200  # ako prefiks nije poznat

# -------------------------------------------------------------------------
# VOLUMEN (CBM) PO KOMADU
# Dimenzije: 25 × 25 × 1 cm
# -------------------------------------------------------------------------

CBM_PO_KOMADU = 0.25 * 0.25 * 0.01  # = 0.000625 m3

# -------------------------------------------------------------------------
# HELPER FUNKCIJE: PREFIKSI, CIJENE, TEŽINE, MP CIJENE
# -------------------------------------------------------------------------


def sifra_to_prefix(sifra: str) -> str:
    """Pronalaženje najdužeg prefiksa koji odgovara šifri."""
    if not isinstance(sifra, str):
        return ""
    sifra = sifra.strip().upper()
    candidates = [p for p in prefix_map.keys() if sifra.startswith(p)]
    if not candidates:
        return ""
    return max(candidates, key=len)


def kategorija_za_sifru(sifra: str, allow_custom: bool = True) -> str:
    """Pronalaženje kategorije prema prefiksu."""
    if not isinstance(sifra, str):
        return "Ostalo"
    sku = sifra.strip().upper()
    if sku in SKU_CATEGORY_OVERRIDES:
        return SKU_CATEGORY_OVERRIDES[sku]
    if allow_custom and sku in CUSTOM_SKU_SET:
        return "Custom"
    pref = sifra_to_prefix(sifra)
    return prefix_map.get(pref, "Ostalo")


def cijena_za_sifru(sifra: str) -> float:
    """Nabavna cijena artikla prema prefiksu."""
    if not isinstance(sifra, str):
        return 0.0
    sku = sifra.strip().upper()
    if sku in KALK_NABAVNE_BY_SKU:
        return float(KALK_NABAVNE_BY_SKU.get(sku, 0.0))
    if sku in PRICE_BY_SKU:
        return float(PRICE_BY_SKU.get(sku, 0.0))
    pref = sifra_to_prefix(sifra)
    return float(PRICE_BY_PREFIX.get(pref, 0.0))


def nabavna_value_and_currency(sifra: str):
    """Vrati (vrijednost, is_km)."""
    if not isinstance(sifra, str):
        return 0.0, False
    sku = sifra.strip().upper()
    if sku in PRICE_BY_SKU:
        return float(PRICE_BY_SKU.get(sku, 0.0)), False
    if sku in KALK_NABAVNE_BY_SKU:
        return float(KALK_NABAVNE_BY_SKU.get(sku, 0.0)), True
    pref = sifra_to_prefix(sifra)
    return float(PRICE_BY_PREFIX.get(pref, 0.0)), False


def mp_cijena_za_sifru(sifra: str) -> float:
    """Maloprodajna (KM) cijena artikla prema sifrarniku."""
    if not isinstance(sifra, str):
        return 0.0
    sku = sifra.strip().upper()
    if sku in KALK_MPC_W_BY_SKU:
        return float(KALK_MPC_W_BY_SKU.get(sku, 0.0))
    cat = kategorija_za_sifru(sku)
    if cat in MP_PRICE_BY_CATEGORY:
        return float(MP_PRICE_BY_CATEGORY.get(cat, 0.0))
    return float(mp_cijene.get(sku, 0.0))


def tezina_za_sifru(sifra: str) -> float:
    """Težina artikla u kg (grami / 1000)."""
    pref = sifra_to_prefix(sifra)
    g = tezine_prefiksi.get(pref, DEFAULT_TEZINA)
    return float(g) / 1000.0


def cbm_za_kolicinu(qty: int) -> float:
    """Volumen u CBM za zadani broj komada."""
    return float(qty) * CBM_PO_KOMADU


# -------------------------------------------------------------------------
# VALIDACIJE I UTILITY FUNKCIJE
# -------------------------------------------------------------------------


def fetch_usd_bam_rate() -> float:
    """Vraca srednji kurs USD->KM sa CBBiH."""
    url = "https://www.cbbh.ba/CurrencyExchange/"
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        html = resp.read().decode("utf-8", errors="ignore")

    rows = re.findall(r"<tr[^>]*>.*?</tr>", html, flags=re.DOTALL | re.IGNORECASE)
    for row in rows:
        if re.search(r'currcircle">\s*USD\s*<', row, re.IGNORECASE):
            m = re.search(
                r"middle-column[^>]*>\s*([0-9]+(?:[.,][0-9]+)?)\s*<",
                row,
            )
            if m:
                rate_str = m.group(1).replace(",", ".")
                return float(rate_str)
    raise ValueError("USD kurs nije pronadjen.")


def safe_float(val, default=0.0):
    """Sigurna konverzija u float."""
    try:
        return float(val)
    except:
        return default


def safe_int(val, default=0):
    """Sigurna konverzija u int."""
    try:
        return int(val)
    except:
        return default


def auto_format_sifra(s: str) -> str:
    """Čišćenje i normalizacija šifre."""
    if not isinstance(s, str):
        return ""
    return s.strip().upper()


def _normalize_col_name(value) -> str:
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("?", "dj")
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Preimenuje osnovne kolone na konzistentna ASCII imena."""
    rename = {}
    for col in df.columns:
        norm = _normalize_col_name(col)
        if norm == "sifra" or ("sifra" in norm and "dobavlj" in norm):
            rename[col] = "Sifra"
        elif "kolicina" in norm:
            rename[col] = "Kolicina"
        elif "stanje" in norm or "zaliha" in norm:
            rename[col] = "Stanje"
        elif "datum" in norm:
            rename[col] = "Datum"
        elif "vrijednost" in norm:
            rename[col] = "Vrijednost"
    if rename:
        df = df.rename(columns=rename)
    return df


# -------------------------------------------------------------------------
# KALKULACIJE -> NABAVNE CIJENE
# -------------------------------------------------------------------------


def parse_calc_excel_to_rows(
    excel_path: str, vat_default: float = 17.0, col_offsets=None, sheet_name=0
):
    if not os.path.isfile(excel_path):
        return False, f"Excel fajl ne postoji: {excel_path}", []

    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    except Exception as e:
        return False, f"Greska pri citanju Excela: {e}", []

    return parse_calc_df_to_rows(
        df, excel_path, vat_default=vat_default, col_offsets=col_offsets
    )


def parse_calc_df_to_rows(df, excel_path: str, vat_default: float = 17.0, col_offsets=None):
    mask_ident = df.apply(
        lambda row: row.astype(str).str.contains("Ident", case=False, na=False),
        axis=1,
    ).any(axis=1)
    if not mask_ident.any():
        mask_ident = df.apply(
            lambda row: row.astype(str).str.contains("Red. br.", case=False, na=False),
            axis=1,
        ).any(axis=1)
    header_indices = sorted(df.index[mask_ident].tolist())

    if not header_indices:
        return False, "Nisam nasao header red sa 'Ident'", []

    header_idx = header_indices[0]
    header_row = df.loc[header_idx]

    def normalize_header(value):
        if pd.isna(value):
            return ""
        text = str(value).replace("\n", " ").strip()
        text = " ".join(text.split())
        return (
            text.replace("č", "c")
            .replace("ć", "c")
            .replace("š", "s")
            .replace("ž", "z")
            .replace("đ", "d")
            .replace("Č", "C")
            .replace("Ć", "C")
            .replace("Š", "S")
            .replace("Ž", "Z")
            .replace("Đ", "D")
        )

    def find_col(header_row, keyword, from_right=False, contains=True):
        keyword = keyword.lower()
        keyword_comp = keyword.replace(".", "").replace(" ", "")
        cols = list(header_row.index)
        iterator = reversed(cols) if from_right else cols

        for col in iterator:
            h = normalize_header(header_row[col]).lower()
            if not h:
                continue
            if contains and keyword in h:
                return col
            if not contains and h == keyword:
                return col
            h_comp = h.replace(".", "").replace(" ", "")
            if contains and keyword_comp and keyword_comp in h_comp:
                return col
            if not contains and keyword_comp and h_comp == keyword_comp:
                return col
        return None

    col_red_br = find_col(header_row, "red.", contains=True)
    col_sku = find_col(header_row, "ident", contains=True)
    col_name = find_col(header_row, "naziv", contains=True)
    col_kolicina = find_col(header_row, "kolicina", contains=True)
    col_pdv_rate = find_col(header_row, "pdv %", from_right=True)
    col_prod_cijena_s_pdv = find_col(
        header_row, "prod. cijena s pdv", contains=True
    )
    if col_prod_cijena_s_pdv is None:
        col_prod_cijena_s_pdv = find_col(
            header_row, "cijena s pdv", contains=True
        )
    if col_prod_cijena_s_pdv is None:
        col_prod_cijena_s_pdv = find_col(
            header_row, "prodajna cijena s pdv", contains=True
        )
    if col_prod_cijena_s_pdv is None:
        col_prod_cijena_s_pdv = find_col(
            header_row, "cijena sa pdv", contains=True
        )
    if col_prod_cijena_s_pdv is None:
        col_prod_cijena_s_pdv = find_col(
            header_row, "prod. cijena sa pdv", contains=True
        )
    col_prod_cijena_bez_pdv = find_col(
        header_row, "prod. cijena bez pdv", contains=True
    )
    if col_prod_cijena_bez_pdv is None:
        col_prod_cijena_bez_pdv = find_col(
            header_row, "cijena bez pdv", contains=True
        )
    if col_prod_cijena_bez_pdv is None:
        col_prod_cijena_bez_pdv = find_col(
            header_row, "prodajna cijena bez pdv", contains=True
        )
    col_nabavna = find_col(header_row, "nabavna vrijednost", contains=True)
    if col_nabavna is None:
        col_nabavna = find_col(header_row, "nabavna", contains=True)

    if col_sku is None or col_prod_cijena_s_pdv is None:
        return False, "Fale kljucne kolone (Ident / Prod. cijena s PDV)", []

    col_list = list(df.columns)
    offsets = col_offsets or {}

    def apply_offset(col, offset, keyword=None):
        if col is None or offset in (None, 0):
            return col
        try:
            pos = col_list.index(col)
        except ValueError:
            return col
        new_pos = pos + int(offset)
        if new_pos < 0 or new_pos >= len(col_list):
            return col
        if keyword:
            key = keyword.lower()
            key_comp = key.replace(".", "").replace(" ", "")
            orig_h = normalize_header(header_row[col]).lower()
            tgt_h = normalize_header(header_row[col_list[new_pos]]).lower()

            def match(h):
                h_comp = h.replace(".", "").replace(" ", "")
                return (key in h) or (key_comp and key_comp in h_comp)

            if match(orig_h) and not match(tgt_h):
                return col
        return col_list[new_pos]

    col_sku = apply_offset(col_sku, offsets.get("sku"), keyword="ident")
    col_nabavna = apply_offset(col_nabavna, offsets.get("nabavna"), keyword="nabavna")
    col_prod_cijena_s_pdv = apply_offset(
        col_prod_cijena_s_pdv, offsets.get("mpc_w"), keyword="s pdv"
    )
    col_prod_cijena_bez_pdv = apply_offset(
        col_prod_cijena_bez_pdv, offsets.get("mpc_wo"), keyword="bez pdv"
    )
    col_pdv_rate = apply_offset(col_pdv_rate, offsets.get("pdv_rate"), keyword="pdv %")
    col_kolicina = apply_offset(col_kolicina, offsets.get("kolicina"), keyword="kolicina")

    def try_get_number(value):
        if value is None:
            return None
        if isinstance(value, (int, float, np.integer, np.floating)):
            if pd.isna(value):
                return None
            return float(value)
        s = str(value).strip().replace(" ", "").replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    basename = os.path.basename(excel_path)
    doc_no = os.path.splitext(basename)[0]
    doc_date = datetime.fromtimestamp(os.path.getmtime(excel_path)).strftime(
        "%Y-%m-%d"
    )

    rows_out = []
    all_block_bounds = header_indices + [len(df)]
    for i in range(len(header_indices)):
        start = header_indices[i] + 1
        end = all_block_bounds[i + 1]

        for idx in range(start, end):
            row = df.loc[idx]
            if row.isna().all():
                continue
            if col_red_br is not None:
                red_br_val = row[col_red_br]
                sku_probe = row[col_sku] if col_sku is not None else None
                if pd.isna(red_br_val) and (col_sku is None or pd.isna(sku_probe)):
                    continue
            sku = ""
            name = ""
            sku_raw = row[col_sku]
            if pd.isna(sku_raw):
                sku_raw = None
            if sku_raw is None and col_red_br is not None:
                red_br_val = row[col_red_br]
                if isinstance(red_br_val, str):
                    m = re.match(r"^\s*(\d+)\s+([A-Za-z0-9_.-]+)\s+(.*)$", red_br_val)
                    if m:
                        sku = m.group(2).strip().upper()
                        name = m.group(3).strip()
            if not sku:
                if sku_raw is None:
                    continue
                sku = str(sku_raw).strip()
                if not sku or sku.lower() == "nan":
                    continue
            name_raw = row[col_name] if col_name is not None else None
            if not name:
                name = "" if (name_raw is None or pd.isna(name_raw)) else str(name_raw).strip()

            qty = None
            if col_kolicina is not None:
                col_list = list(df.columns)
                col_pos = col_list.index(col_kolicina)
                search_cols = [col_kolicina]
                for offset in (1, 2, 3):
                    if col_pos - offset >= 0:
                        search_cols.append(col_list[col_pos - offset])
                    if col_pos + offset < len(col_list):
                        search_cols.append(col_list[col_pos + offset])
                for c in search_cols:
                    num = try_get_number(row[c])
                    if num is not None:
                        qty = num
                        break

            sale_with_vat = try_get_number(row[col_prod_cijena_s_pdv])
            if (
                sale_with_vat is None
                and col_prod_cijena_s_pdv is not None
                and offsets.get("mpc_w") in (None, 0)
            ):
                pos = col_list.index(col_prod_cijena_s_pdv)
                neighbor_cols = []
                if pos + 1 < len(col_list):
                    neighbor_cols.append(col_list[pos + 1])
                if pos > 0:
                    neighbor_cols.append(col_list[pos - 1])
                for c in neighbor_cols:
                    header_val = normalize_header(header_row[c])
                    if header_val:
                        continue
                    num = try_get_number(row[c])
                    if num is not None:
                        sale_with_vat = num
                        break
            vat_rate = None
            if col_pdv_rate is not None:
                vat_rate = try_get_number(row[col_pdv_rate])
            if vat_rate is None:
                vat_rate = vat_default

            sale_no_vat = None
            if col_prod_cijena_bez_pdv is not None:
                sale_no_vat = try_get_number(row[col_prod_cijena_bez_pdv])
            if sale_no_vat is None and sale_with_vat is not None and vat_rate is not None:
                try:
                    sale_no_vat = round(sale_with_vat / (1.0 + vat_rate / 100.0), 4)
                except ZeroDivisionError:
                    sale_no_vat = None
            if sale_with_vat is None and sale_no_vat is not None and vat_rate is not None:
                try:
                    sale_with_vat = round(sale_no_vat * (1.0 + vat_rate / 100.0), 4)
                except ZeroDivisionError:
                    sale_with_vat = None

            purchase_price = None
            if col_nabavna is not None:
                nab_pos = col_list.index(col_nabavna)
                if offsets.get("nabavna") not in (None, 0):
                    candidate_cols = [col_nabavna]
                else:
                    # Some calc sheets have the "Nabavna vrijednost" value shifted by +/-1.
                    candidate_cols = [col_nabavna]
                    if nab_pos + 1 < len(col_list):
                        candidate_cols.append(col_list[nab_pos + 1])
                    if nab_pos > 0:
                        candidate_cols.append(col_list[nab_pos - 1])

                for c in candidate_cols:
                    num = try_get_number(row[c])
                    if num is not None:
                        purchase_price = num
                        break

            rows_out.append(
                {
                    "sku": sku,
                    "name_hint": name,
                    "doc_no": doc_no,
                    "doc_date": doc_date,
                    "qty": qty,
                    "nab_unit": purchase_price,
                    "mpc_wo_vat_unit": sale_no_vat,
                    "mpc_w_vat_unit": sale_with_vat,
                    "vat_rate": vat_rate,
                    "notes": "excel_calc",
                }
            )

    if not rows_out:
        return False, "Nisam nasao nijednu stavku (SKU) u kalkulaciji", []

    return True, f"OK - {len(rows_out)} stavki", rows_out


def find_excel_files(root_dir):
    excel_files = []
    for dirpath, _, filenames in os.walk(root_dir):
        for name in filenames:
            if name.lower().endswith(".xlsx"):
                excel_files.append(os.path.join(dirpath, name))
    def _natural_key(path):
        base = os.path.basename(path).lower()
        parts = re.split(r"(\d+)", base)
        key = []
        for part in parts:
            if part.isdigit():
                key.append(int(part))
            else:
                key.append(part)
        return key

    return sorted(excel_files, key=_natural_key)


def parse_float_safe(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date_safe(s):
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def load_kalk_json(path):
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}
    items = data.get("items")
    if not isinstance(items, dict):
        return {}
    return items


def _accumulate_kalk_rows(rows, sku_data, file_name):
    for row in rows:
        sku = (row.get("sku") or "").strip().upper()
        if not sku:
            continue
        doc_no = (row.get("doc_no") or "").strip()
        doc_date_raw = (row.get("doc_date") or "").strip()
        nab_unit = parse_float_safe(row.get("nab_unit"))
        qty = parse_float_safe(row.get("qty"))
        mpc_wo = parse_float_safe(row.get("mpc_wo_vat_unit"))
        mpc_w = parse_float_safe(row.get("mpc_w_vat_unit"))
        vat_rate = parse_float_safe(row.get("vat_rate"))
        doc_date = parse_date_safe(doc_date_raw)

        sd = sku_data.setdefault(
            sku,
            {
                "sum_qty": 0.0,
                "sum_nab_value": 0.0,
                "nab_values": [],
                "sum_mpc_wo_value": 0.0,
                "sum_mpc_w_value": 0.0,
                "sum_vat_value": 0.0,
                "count_mpc_wo": 0,
                "count_mpc_w": 0,
                "count_vat": 0,
                "last_doc_date": None,
                "last_doc_no": "",
                "last_file_name": "",
                "last_nab_unit": 0.0,
                "last_mpc_wo": 0.0,
                "last_mpc_w": 0.0,
                "last_vat_rate": 0.0,
            },
        )

        if qty > 0:
            sd["sum_qty"] += qty
            if nab_unit > 0:
                sd["sum_nab_value"] += nab_unit * qty
                sd["nab_values"].append(nab_unit)
            if mpc_wo > 0:
                sd["count_mpc_wo"] += 1
                sd["sum_mpc_wo_value"] += mpc_wo * qty
            if mpc_w > 0:
                sd["count_mpc_w"] += 1
                sd["sum_mpc_w_value"] += mpc_w * qty
            if vat_rate > 0:
                sd["count_vat"] += 1
                sd["sum_vat_value"] += vat_rate * qty

        update_last = False
        if doc_date is not None:
            if sd["last_doc_date"] is None or doc_date > sd["last_doc_date"]:
                update_last = True
            elif doc_date == sd["last_doc_date"]:
                if doc_no and doc_no > sd["last_doc_no"]:
                    update_last = True
        elif sd["last_doc_date"] is None:
            update_last = True

        if update_last:
            sd["last_doc_date"] = doc_date
            sd["last_doc_no"] = doc_no
            sd["last_file_name"] = file_name
            sd["last_nab_unit"] = nab_unit
            sd["last_mpc_wo"] = mpc_wo
            sd["last_mpc_w"] = mpc_w
            sd["last_vat_rate"] = vat_rate


def _build_kalk_file_items(rows, file_name):
    sku_data = {}
    for row in rows:
        sku = (row.get("sku") or "").strip().upper()
        if not sku:
            continue
        doc_no = (row.get("doc_no") or "").strip()
        doc_date_raw = (row.get("doc_date") or "").strip()
        nab_unit = parse_float_safe(row.get("nab_unit"))
        qty = parse_float_safe(row.get("qty"))
        mpc_wo = parse_float_safe(row.get("mpc_wo_vat_unit"))
        mpc_w = parse_float_safe(row.get("mpc_w_vat_unit"))
        vat_rate = parse_float_safe(row.get("vat_rate"))
        doc_date = parse_date_safe(doc_date_raw)

        sd = sku_data.setdefault(
            sku,
            {
                "sum_qty": 0.0,
                "sum_nab_value": 0.0,
                "sum_mpc_wo_value": 0.0,
                "sum_mpc_w_value": 0.0,
                "sum_vat_value": 0.0,
                "count_mpc_wo": 0,
                "count_mpc_w": 0,
                "count_vat": 0,
                "nab_values_count": 0,
                "last_doc_date": "",
                "last_doc_no": "",
                "last_file_name": "",
                "last_nab_unit": 0.0,
                "last_mpc_wo": 0.0,
                "last_mpc_w": 0.0,
                "last_vat_rate": 0.0,
            },
        )

        if qty > 0:
            sd["sum_qty"] += qty
            if nab_unit > 0:
                sd["sum_nab_value"] += nab_unit * qty
                sd["nab_values_count"] += 1
            if mpc_wo > 0:
                sd["count_mpc_wo"] += 1
                sd["sum_mpc_wo_value"] += mpc_wo * qty
            if mpc_w > 0:
                sd["count_mpc_w"] += 1
                sd["sum_mpc_w_value"] += mpc_w * qty
            if vat_rate > 0:
                sd["count_vat"] += 1
                sd["sum_vat_value"] += vat_rate * qty

        update_last = False
        if doc_date is not None:
            last_doc_date = parse_date_safe(sd.get("last_doc_date", ""))
            if last_doc_date is None or doc_date > last_doc_date:
                update_last = True
            elif last_doc_date is not None and doc_date == last_doc_date:
                if doc_no and doc_no > sd.get("last_doc_no", ""):
                    update_last = True
        elif not sd.get("last_doc_date"):
            update_last = True

        if update_last:
            sd["last_doc_date"] = doc_date.isoformat() if doc_date else ""
            sd["last_doc_no"] = doc_no
            sd["last_file_name"] = file_name
            sd["last_nab_unit"] = nab_unit
            sd["last_mpc_wo"] = mpc_wo
            sd["last_mpc_w"] = mpc_w
            sd["last_vat_rate"] = vat_rate

    return sku_data


def _merge_kalk_file_cache(cache):
    sku_data = {}
    for entry in (cache or {}).values():
        items = entry.get("items", {})
        if not isinstance(items, dict):
            continue
        for sku, fd in items.items():
            sku_key = str(sku).strip().upper()
            if not sku_key:
                continue
            sd = sku_data.setdefault(
                sku_key,
                {
                    "sum_qty": 0.0,
                    "sum_nab_value": 0.0,
                    "sum_mpc_wo_value": 0.0,
                    "sum_mpc_w_value": 0.0,
                    "sum_vat_value": 0.0,
                    "count_mpc_wo": 0,
                    "count_mpc_w": 0,
                    "count_vat": 0,
                    "nab_values_count": 0,
                    "last_doc_date": None,
                    "last_doc_no": "",
                    "last_file_name": "",
                    "last_nab_unit": 0.0,
                    "last_mpc_wo": 0.0,
                    "last_mpc_w": 0.0,
                    "last_vat_rate": 0.0,
                },
            )
            sd["sum_qty"] += parse_float_safe(fd.get("sum_qty"))
            sd["sum_nab_value"] += parse_float_safe(fd.get("sum_nab_value"))
            sd["sum_mpc_wo_value"] += parse_float_safe(fd.get("sum_mpc_wo_value"))
            sd["sum_mpc_w_value"] += parse_float_safe(fd.get("sum_mpc_w_value"))
            sd["sum_vat_value"] += parse_float_safe(fd.get("sum_vat_value"))
            sd["count_mpc_wo"] += int(parse_float_safe(fd.get("count_mpc_wo")))
            sd["count_mpc_w"] += int(parse_float_safe(fd.get("count_mpc_w")))
            sd["count_vat"] += int(parse_float_safe(fd.get("count_vat")))
            sd["nab_values_count"] += int(parse_float_safe(fd.get("nab_values_count")))

            doc_date = parse_date_safe(fd.get("last_doc_date", ""))
            doc_no = str(fd.get("last_doc_no") or "")
            update_last = False
            if doc_date is not None:
                if sd["last_doc_date"] is None or doc_date > sd["last_doc_date"]:
                    update_last = True
                elif doc_date == sd["last_doc_date"] and doc_no > sd["last_doc_no"]:
                    update_last = True
            elif sd["last_doc_date"] is None:
                update_last = True

            if update_last:
                sd["last_doc_date"] = doc_date
                sd["last_doc_no"] = doc_no
                sd["last_file_name"] = str(fd.get("last_file_name") or "")
                sd["last_nab_unit"] = parse_float_safe(fd.get("last_nab_unit"))
                sd["last_mpc_wo"] = parse_float_safe(fd.get("last_mpc_wo"))
                sd["last_mpc_w"] = parse_float_safe(fd.get("last_mpc_w"))
                sd["last_vat_rate"] = parse_float_safe(fd.get("last_vat_rate"))

    return sku_data


def _kalk_finalize_and_write(sku_data, out_path):
    items = {}
    for sku, sd in sku_data.items():
        avg_nabavna = 0.0
        if sd["sum_qty"] > 0:
            avg_nabavna = sd["sum_nab_value"] / sd["sum_qty"]
        avg_mpc_wo = 0.0
        if sd["sum_qty"] > 0 and sd["sum_mpc_wo_value"] > 0:
            avg_mpc_wo = sd["sum_mpc_wo_value"] / sd["sum_qty"]
        avg_mpc_w = 0.0
        if sd["sum_qty"] > 0 and sd["sum_mpc_w_value"] > 0:
            avg_mpc_w = sd["sum_mpc_w_value"] / sd["sum_qty"]
        avg_vat = 0.0
        if sd["sum_qty"] > 0 and sd["sum_vat_value"] > 0:
            avg_vat = sd["sum_vat_value"] / sd["sum_qty"]
        items[sku] = {
            "nabavna_cijena": float(f"{sd.get('last_nab_unit', 0.0):.6f}"),
            "avg_nabavna": float(f"{avg_nabavna:.6f}"),
            "avg_mpc_wo_vat": float(f"{avg_mpc_wo:.6f}"),
            "avg_mpc_w_vat": float(f"{avg_mpc_w:.6f}"),
            "avg_vat_rate": float(f"{avg_vat:.6f}"),
            "sum_qty": float(f"{sd.get('sum_qty', 0.0):.4f}"),
            "count": int(
                sd.get(
                    "nab_values_count",
                    len(sd.get("nab_values", [])),
                )
            ),
            "count_mpc_wo": sd.get("count_mpc_wo", 0),
            "count_mpc_w": sd.get("count_mpc_w", 0),
            "count_vat": sd.get("count_vat", 0),
            "last_doc_no": sd.get("last_doc_no", ""),
            "last_doc_date": sd["last_doc_date"].isoformat()
            if isinstance(sd.get("last_doc_date"), date)
            else "",
            "last_file_name": sd.get("last_file_name", ""),
            "last_mpc_wo": float(f"{sd.get('last_mpc_wo', 0.0):.6f}"),
            "last_mpc_w": float(f"{sd.get('last_mpc_w', 0.0):.6f}"),
            "last_vat_rate": float(f"{sd.get('last_vat_rate', 0.0):.6f}"),
        }

    data = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": "FEMMA12",
        "items": items,
    }

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    with open(out_path, "w", encoding="utf-8") as f_json:
        json.dump(data, f_json, ensure_ascii=False, indent=2)

    return {sku: v["avg_nabavna"] for sku, v in items.items()}


def build_kalk_nabavne_json(kalk_folder, out_path):
    excel_files = find_excel_files(kalk_folder)
    if not excel_files:
        return {}

    sku_data = {}
    for path in excel_files:
        try:
            ok, _msg, rows = parse_calc_excel_to_rows(path)
        except Exception:
            ok = False
            rows = []
        if not ok:
            continue
        _accumulate_kalk_rows(rows, sku_data, os.path.basename(path))

    return _kalk_finalize_and_write(sku_data, out_path)


def build_kalk_nabavne_json_with_review(
    kalk_folder, out_path, review_cb, excel_files=None
):
    excel_files = excel_files or find_excel_files(kalk_folder)
    if excel_files:
        seen = set()
        unique_files = []
        for p in excel_files:
            try:
                key = os.path.normcase(os.path.abspath(p))
            except Exception:
                key = p
            if key in seen:
                continue
            seen.add(key)
            unique_files.append(p)
        excel_files = unique_files
    if not excel_files:
        return {}, {
            "total_rows": 0,
            "qty_found": 0,
            "qty_missing": 0,
            "accepted_files": 0,
            "total_files": 0,
            "accepted_paths": [],
        }

    sku_data = {}
    total_rows = 0
    qty_found = 0
    accepted_files = 0
    accepted_paths = []
    override_manual_skus = set()
    for path in excel_files:
        try:
            ok, _msg, rows = parse_calc_excel_to_rows(path)
        except Exception:
            ok = False
            rows = []
        if not ok:
            continue
        if review_cb:
            review_result = review_cb(path, rows)
            if isinstance(review_result, dict):
                ok = bool(review_result.get("ok", False))
                rows = review_result.get("rows", rows)
                override_manual_skus.update(review_result.get("override_skus", []))
            else:
                ok = bool(review_result)
            if not ok:
                continue
        accepted_files += 1
        accepted_paths.append(path)
        for row in rows:
            sku = (row.get("sku") or "").strip()
            if not sku:
                continue
            total_rows += 1
            qty_val = parse_float_safe(row.get("qty"))
            if qty_val > 0:
                qty_found += 1
        _accumulate_kalk_rows(rows, sku_data, os.path.basename(path))

    if accepted_files == 0:
        return {}, {
            "total_rows": 0,
            "qty_found": 0,
            "qty_missing": 0,
            "accepted_files": 0,
            "total_files": len(excel_files),
            "accepted_paths": [],
            "override_manual_skus": [],
        }

    items = _kalk_finalize_and_write(sku_data, out_path)
    qty_missing = total_rows - qty_found
    return items, {
        "total_rows": total_rows,
        "qty_found": qty_found,
        "qty_missing": qty_missing,
        "accepted_files": accepted_files,
        "total_files": len(excel_files),
        "accepted_paths": accepted_paths,
        "override_manual_skus": sorted(override_manual_skus),
    }


def ensure_kalk_nabavne_loaded(kalk_folder):
    global KALK_NABAVNE_BY_SKU, KALK_MPC_WO_BY_SKU, KALK_MPC_W_BY_SKU, KALK_VAT_BY_SKU
    if not kalk_folder:
        KALK_NABAVNE_BY_SKU = {}
        KALK_MPC_WO_BY_SKU = {}
        KALK_MPC_W_BY_SKU = {}
        KALK_VAT_BY_SKU = {}
        return
    data = load_kalk_json(KALK_NABAVNE_PATH)
    if not data:
        if os.path.isdir(kalk_folder) and find_excel_files(kalk_folder):
            build_kalk_nabavne_json(kalk_folder, KALK_NABAVNE_PATH)
            data = load_kalk_json(KALK_NABAVNE_PATH)
    KALK_NABAVNE_BY_SKU = {}
    KALK_MPC_WO_BY_SKU = {}
    KALK_MPC_W_BY_SKU = {}
    KALK_VAT_BY_SKU = {}
    for sku, item in (data or {}).items():
        key = str(sku).strip().upper()
        try:
            KALK_NABAVNE_BY_SKU[key] = float(
                item.get("avg_nabavna", item.get("nabavna_cijena", 0.0))
            )
        except Exception:
            pass
        try:
            KALK_MPC_WO_BY_SKU[key] = float(
                item.get("avg_mpc_wo_vat", item.get("last_mpc_wo", 0.0))
            )
        except Exception:
            pass
        try:
            KALK_MPC_W_BY_SKU[key] = float(
                item.get("avg_mpc_w_vat", item.get("last_mpc_w", 0.0))
            )
        except Exception:
            pass
        try:
            KALK_VAT_BY_SKU[key] = float(
                item.get("avg_vat_rate", item.get("last_vat_rate", 0.0))
            )
        except Exception:
            pass


def calc_kalk_qty_stats(kalk_folder):
    total_rows = 0
    qty_found = 0
    excel_files = find_excel_files(kalk_folder)
    for path in excel_files:
        try:
            ok, _msg, rows = parse_calc_excel_to_rows(path)
        except Exception:
            ok = False
            rows = []
        if not ok:
            continue
        for row in rows:
            sku = (row.get("sku") or "").strip()
            if not sku:
                continue
            total_rows += 1
            if row.get("qty") is not None:
                qty_found += 1
    qty_missing = total_rows - qty_found
    pct = (qty_found / total_rows * 100.0) if total_rows else 0.0
    return total_rows, qty_found, qty_missing, pct


ensure_kalk_nabavne_loaded(kalkulacije_folder)


def ensure_default_fills(wb) -> None:
    """Ensure Excel default fills include gray125 to avoid dotted pattern."""
    fills = getattr(wb, "_fills", None)
    if not fills:
        return
    has_gray = any(getattr(f, "patternType", None) == "gray125" for f in fills)
    if not has_gray:
        fills.insert(1, fills_mod.DEFAULT_GRAY_FILL)


def round_custom_qty(val) -> int:
    """Za custom proizvode zaokruži na najbližih 10, najmanje 50."""
    qty = safe_float(val, 0.0)
    return int(max(50, math.ceil(qty / 10.0) * 10))


def find_stanje_df(wb) -> pd.DataFrame | None:
    """Pronađi sheet koji sadrži kolone Sifra i Stanje (bilo koji sheet)."""
    for name in wb.sheetnames:
        ws = wb[name]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        df = pd.DataFrame(data)
        df.columns = df.iloc[0]
        df = df[1:]
        df = normalize_columns(df)
        if "Sifra" in df.columns and "Stanje" in df.columns:
            df["Sifra"] = df["Sifra"].astype(str).apply(auto_format_sifra)
            df["Stanje"] = df["Stanje"].apply(safe_float)
            return df
    return None


def find_sales_df(wb) -> pd.DataFrame | None:
    """Pronadji sheet koji sadrzi kolone Sifra i Kolicina (bilo koji sheet)."""
    for name in wb.sheetnames:
        ws = wb[name]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        df = pd.DataFrame(data)
        df.columns = df.iloc[0]
        df = df[1:]
        df = normalize_columns(df)
        if "Sifra" in df.columns and "Kolicina" in df.columns:
            df["Sifra"] = df["Sifra"].astype(str).apply(auto_format_sifra)
            df["Kolicina"] = df["Kolicina"].apply(safe_float)
            return df
    return None


def _promet_normalize_header(value):
    if pd.isna(value):
        return ""
    text = str(value).replace("\n", " ").strip()
    text = " ".join(text.split())
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = (
        text.replace("š", "s")
        .replace("Š", "S")
        .replace("č", "c")
        .replace("Č", "C")
        .replace("ć", "c")
        .replace("Ć", "C")
        .replace("đ", "d")
        .replace("Đ", "D")
        .replace("ž", "z")
        .replace("Ž", "Z")
    )
    return text.lower()


def _promet_find_header_indices(df, keywords_required):
    indices = []
    for idx in df.index:
        row = df.loc[idx]
        if row.isna().all():
            continue
        found = {key: False for key in keywords_required}
        for cell in row:
            h = _promet_normalize_header(cell)
            if not h:
                continue
            for key, words in keywords_required.items():
                if found[key]:
                    continue
                for w in words:
                    w_norm = w.lower()
                    w_comp = w_norm.replace(".", "").replace(" ", "")
                    h_comp = h.replace(".", "").replace(" ", "")
                    if w_norm in h or (w_comp and w_comp in h_comp):
                        found[key] = True
                        break
        if all(found.values()):
            indices.append(idx)
    return indices


def parse_promet_df_to_rows(df, kind: str, col_offsets=None, add_vat=True):
    if df is None or df.empty or df.isna().all().all():
        return True, "Prazan sheet", []

    if kind == "sales":
        keywords_required = {
            "sifra": ["sifra", "ident", "sifra dobavljaca"],
            "kolicina": ["kolicina"],
        }
    else:
        keywords_required = {
            "sifra": ["sifra", "ident", "sifra dobavljaca"],
            "stanje": ["stanje", "zaliha"],
        }

    header_indices = _promet_find_header_indices(df, keywords_required)
    if not header_indices:
        return False, "Nisam nasao header red sa trazenim kolonama.", []

    col_list = list(df.columns)
    offsets = col_offsets or {}

    def find_col_any(header_row, keywords, from_right=False):
        cols = list(header_row.index)
        iterator = reversed(cols) if from_right else cols
        for col in iterator:
            h = _promet_normalize_header(header_row[col])
            if not h:
                continue
            h_comp = h.replace(".", "").replace(" ", "")
            for w in keywords:
                w_norm = w.lower()
                w_comp = w_norm.replace(".", "").replace(" ", "")
                if w_norm in h or (w_comp and w_comp in h_comp):
                    return col
        return None

    def apply_offset(col, offset, keywords=None):
        if col is None or offset in (None, 0):
            return col
        try:
            pos = col_list.index(col)
        except ValueError:
            return col
        new_pos = pos + int(offset)
        if new_pos < 0 or new_pos >= len(col_list):
            return col
        if keywords:
            orig_h = _promet_normalize_header(header_row[col])
            tgt_h = _promet_normalize_header(header_row[col_list[new_pos]])
            for w in keywords:
                w_norm = w.lower()
                w_comp = w_norm.replace(".", "").replace(" ", "")
                orig_match = (w_norm in orig_h) or (w_comp and w_comp in orig_h.replace(".", "").replace(" ", ""))
                tgt_match = (w_norm in tgt_h) or (w_comp and w_comp in tgt_h.replace(".", "").replace(" ", ""))
                if orig_match and not tgt_match:
                    return col
        return col_list[new_pos]

    rows_out = []
    all_block_bounds = header_indices + [len(df)]
    for i in range(len(header_indices)):
        header_idx = header_indices[i]
        header_row = df.loc[header_idx]

        col_sifra = find_col_any(header_row, ["sifra", "ident", "sifra dobavljaca"])
        col_kolicina = (
            find_col_any(header_row, ["kolicina"]) if kind == "sales" else None
        )
        col_stanje = (
            find_col_any(header_row, ["stanje", "zaliha"]) if kind != "sales" else None
        )
        col_vrijednost = find_col_any(header_row, ["vrijednost", "iznos"])

        col_sifra = apply_offset(col_sifra, offsets.get("sifra"), keywords=["sifra", "ident", "sifra dobavljaca"])
        if kind == "sales":
            col_kolicina = apply_offset(col_kolicina, offsets.get("kolicina"), keywords=["kolicina"])
        else:
            col_stanje = apply_offset(col_stanje, offsets.get("stanje"), keywords=["stanje", "zaliha"])
        col_vrijednost = apply_offset(col_vrijednost, offsets.get("vrijednost"), keywords=["vrijednost", "iznos"])

        if col_sifra is None or (kind == "sales" and col_kolicina is None) or (
            kind != "sales" and col_stanje is None
        ):
            return False, "Fale kljucne kolone (Sifra i Kolicina/Stanje).", []

        def neighbor_numeric(row, col_idx, prefer="any"):
            try:
                pos = col_list.index(col_idx)
            except ValueError:
                return None
            offsets = [1, 2]
            if prefer == "left":
                order = [(-o) for o in offsets] + [o for o in offsets]
            elif prefer == "right":
                order = [o for o in offsets] + [(-o) for o in offsets]
            else:
                order = [(-o) for o in offsets] + [o for o in offsets]

            candidates = []
            for delta in order:
                neighbor_pos = pos + delta
                if neighbor_pos < 0 or neighbor_pos >= len(col_list):
                    continue
                neighbor_col = col_list[neighbor_pos]
                header_val = _promet_normalize_header(header_row[neighbor_col])
                if header_val:
                    continue
                num = safe_float(row[neighbor_col], 0.0)
                if num != 0:
                    candidates.append(num)
                    if prefer in ("left", "right"):
                        return num
            if not candidates:
                return None
            return candidates[0]

        start = header_idx + 1
        end = all_block_bounds[i + 1]

        def numeric_count(col_idx):
            if col_idx is None:
                return 0
            count = 0
            for ridx in range(start, end):
                row = df.loc[ridx]
                val = row[col_idx]
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    continue
                if isinstance(val, str):
                    val = val.replace(",", ".")
                try:
                    num = float(val)
                except Exception:
                    continue
                if not pd.isna(num):
                    count += 1
            return count

        def pick_best_numeric_col(
            base_col,
            max_delta=5,
            exclude_keywords=None,
            allow_header=False,
            prefer="any",
        ):
            if base_col is None:
                return base_col
            try:
                pos = col_list.index(base_col)
            except ValueError:
                return base_col
            base_count = numeric_count(base_col)
            if base_count > 0:
                return base_col
            best_col = base_col
            best_count = base_count
            if prefer == "right":
                order_signs = (1, -1)
            elif prefer == "left":
                order_signs = (-1, 1)
            else:
                order_signs = (-1, 1)
            for delta in range(1, max_delta + 1):
                for sign in order_signs:
                    cand_pos = pos + sign * delta
                    if cand_pos < 0 or cand_pos >= len(col_list):
                        continue
                    cand_col = col_list[cand_pos]
                    header_val = _promet_normalize_header(header_row[cand_col])
                    if header_val and not allow_header:
                        continue
                    if header_val and exclude_keywords:
                        header_comp = header_val.replace(".", "").replace(" ", "")
                        if any(
                            (k in header_val) or (k.replace(".", "").replace(" ", "") in header_comp)
                            for k in exclude_keywords
                        ):
                            continue
                    cnt = numeric_count(cand_col)
                    if cnt > best_count:
                        best_count = cnt
                        best_col = cand_col
            return best_col

        if kind != "sales":
            col_stanje = pick_best_numeric_col(
                col_stanje,
                exclude_keywords=["vrijednost", "iznos", "cijena", "mp", "m.p", "mpcijena", "pdv", "valuta"],
                allow_header=True,
            )
            col_vrijednost = pick_best_numeric_col(
                col_vrijednost,
                exclude_keywords=["zaliha", "stanje", "kolicina", "mj", "valuta", "mp", "m.p", "mpcijena", "cijena"],
                allow_header=False,
                prefer="right",
            )

        def numeric_count(col_idx):
            if col_idx is None:
                return 0
            count = 0
            for ridx in range(start, end):
                row = df.loc[ridx]
                val = row[col_idx]
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    continue
                if isinstance(val, str):
                    val = val.replace(",", ".")
                try:
                    num = float(val)
                except Exception:
                    continue
                if not pd.isna(num):
                    count += 1
            return count

        def pick_best_numeric_col(base_col, max_delta=5):
            if base_col is None:
                return base_col
            try:
                pos = col_list.index(base_col)
            except ValueError:
                return base_col
            base_count = numeric_count(base_col)
            if base_count > 0:
                return base_col
            best_col = base_col
            best_count = base_count
            for delta in range(1, max_delta + 1):
                for sign in (-1, 1):
                    cand_pos = pos + sign * delta
                    if cand_pos < 0 or cand_pos >= len(col_list):
                        continue
                    cand_col = col_list[cand_pos]
                    header_val = _promet_normalize_header(header_row[cand_col])
                    if header_val:
                        continue
                    cnt = numeric_count(cand_col)
                    if cnt > best_count:
                        best_count = cnt
                        best_col = cand_col
            return best_col

        if kind != "sales":
            col_stanje = pick_best_numeric_col(col_stanje)
            col_vrijednost = pick_best_numeric_col(col_vrijednost)
        for idx in range(start, end):
            row = df.loc[idx]
            if row.isna().all():
                continue
            sku_raw = row[col_sifra]
            if sku_raw is None or (isinstance(sku_raw, float) and pd.isna(sku_raw)):
                continue
            sku = str(sku_raw).strip().upper()
            if not sku or sku.lower() == "nan":
                continue
            if sku in ("UKUPNO", "TOTAL", "UKUPNA", "UKUPNO:"):
                continue

            qty = None
            if kind == "sales":
                raw_qty = row[col_kolicina]
                qty = safe_float(raw_qty, 0.0)
                if (raw_qty is None or pd.isna(raw_qty)) and (
                    qty is None or pd.isna(qty) or qty <= 0
                ):
                    alt = neighbor_numeric(row, col_kolicina, prefer="left")
                    if alt is not None:
                        qty = alt
                if qty < 0:
                    qty = 0.0
            else:
                raw_qty = row[col_stanje]
                qty = safe_float(raw_qty, 0.0)
                if (raw_qty is None or pd.isna(raw_qty)) and (
                    qty is None or pd.isna(qty) or qty <= 0
                ):
                    alt = neighbor_numeric(row, col_stanje, prefer="left")
                    if alt is None:
                        alt = neighbor_numeric(row, col_stanje, prefer="right")
                    if alt is not None:
                        qty = alt
                if qty < 0:
                    qty = 0.0

            value = None
            if col_vrijednost is not None:
                raw_val = row[col_vrijednost]
                value = safe_float(raw_val, 0.0)
                if (raw_val is None or pd.isna(raw_val)) and (
                    value is None or pd.isna(value) or value <= 0
                ):
                    alt = neighbor_numeric(row, col_vrijednost, prefer="right")
                    if alt is not None:
                        value = alt
                if value is not None and value < 0:
                    value = 0.0
                if kind == "sales" and add_vat and value:
                    value = value * 1.17

            rows_out.append(
                {
                    "Sifra": auto_format_sifra(sku),
                    "Kolicina": qty if kind == "sales" else None,
                    "Stanje": qty if kind != "sales" else None,
                    "Vrijednost": value,
                }
            )

    return True, "", rows_out


def show_promet_review(path, kind: str):
    parent = tk._default_root
    if parent is None:
        messagebox.showerror("Greska", "Ne mogu otvoriti pregled bez aktivnog prozora.")
        return {"ok": False, "rows": []}

    result = {"ok": False, "rows": []}
    sheet_names = []
    df_cache = {}
    try:
        xls = pd.ExcelFile(path)
        sheet_names = list(xls.sheet_names)
    except Exception:
        sheet_names = []

    def load_sheet_df(name):
        if name in df_cache:
            return df_cache[name]
        df = pd.read_excel(path, sheet_name=name, header=None)
        df_cache[name] = df
        return df

    active_sheet = tk.StringVar(value=sheet_names[0]) if sheet_names else None
    offsets_by_sheet = (
        {name: {"sifra": 0, "kolicina": 0, "stanje": 0, "vrijednost": 0} for name in sheet_names}
        if sheet_names
        else {0: {"sifra": 0, "kolicina": 0, "stanje": 0, "vrijednost": 0}}
    )
    updating_vars = {"active": False}
    visited_sheets = {sheet_names[0]} if sheet_names else set()

    def active_key():
        return active_sheet.get() if active_sheet else 0

    top = ctk.CTkToplevel(parent)
    title = "Provjera ukupne prodaje" if kind == "sales" else "Provjera trenutnog stanja"
    top.title(title)
    top.geometry("980x640")
    top.grab_set()
    top.focus_set()

    header = ctk.CTkLabel(
        top,
        text=title,
        font=ctk.CTkFont(size=14, weight="bold"),
    )
    header.pack(anchor="w", padx=12, pady=(12, 4))

    info = ctk.CTkLabel(
        top,
        text=f"Fajl: {os.path.basename(path)} | Stavki: 0",
        text_color=TEXT_MUTED,
    )
    info.pack(anchor="w", padx=12, pady=(0, 6))

    total_info = ctk.CTkLabel(top, text="", text_color=TEXT_MUTED)
    total_info.pack(anchor="w", padx=12, pady=(0, 6))

    if sheet_names and len(sheet_names) > 1:
        sheet_frame = ctk.CTkFrame(top)
        sheet_frame.pack(fill="x", padx=10, pady=(0, 6))
        ctk.CTkLabel(sheet_frame, text="Sheet:").pack(
            side="left", padx=(10, 6), pady=6
        )
        sheet_combo = ctk.CTkComboBox(
            sheet_frame, values=sheet_names, width=200, variable=active_sheet
        )
        sheet_combo.pack(side="left", padx=6, pady=6)
        ctk.CTkLabel(
            sheet_frame,
            text=f"Ukupno sheetova: {len(sheet_names)}",
            text_color=TEXT_MUTED,
        ).pack(side="left", padx=10, pady=6)

    offsets_frame = ctk.CTkFrame(top)
    offsets_frame.pack(fill="x", padx=10, pady=(0, 6))
    ctk.CTkLabel(
        offsets_frame,
        text="Pomjeri kolone:",
        text_color=TEXT_MUTED,
    ).grid(row=0, column=0, padx=10, pady=6, sticky="w")

    offset_options = [("Lijevo", -1), ("Header", 0), ("Desno", 1)]
    for n in range(2, 101):
        offset_options.insert(0, (f"Lijevo+{n}", -n))
        offset_options.append((f"Desno+{n}", n))
    offset_labels = [label for label, _ in offset_options]
    offset_map = {label: val for label, val in offset_options}

    def make_offset_combo(label_text, column, default="Header"):
        ctk.CTkLabel(offsets_frame, text=label_text).grid(
            row=0, column=column, padx=6, pady=6, sticky="w"
        )
        var = tk.StringVar(value=default)
        combo = ctk.CTkComboBox(
            offsets_frame, values=offset_labels, width=120, variable=var
        )
        combo.grid(row=0, column=column + 1, padx=6, pady=6, sticky="w")
        return var

    def offset_to_label(offset):
        for label, val in offset_options:
            if val == offset:
                return label
        return "Header"

    initial_offsets = offsets_by_sheet.get(active_key(), {})
    var_sifra = make_offset_combo(
        "Sifra", 1, offset_to_label(initial_offsets.get("sifra", 0))
    )
    if kind == "sales":
        var_qty = make_offset_combo(
            "Kolicina", 3, offset_to_label(initial_offsets.get("kolicina", 0))
        )
    else:
        var_qty = make_offset_combo(
            "Stanje", 3, offset_to_label(initial_offsets.get("stanje", 0))
        )
    var_val = make_offset_combo(
        "Vrijednost", 5, offset_to_label(initial_offsets.get("vrijednost", 0))
    )

    move_frame = ctk.CTkFrame(top)
    move_frame.pack(fill="x", padx=10, pady=(0, 6))

    ctk.CTkLabel(move_frame, text="Pomjeri kolonu:").pack(
        side="left", padx=(10, 6), pady=6
    )
    field_options = [("Sifra", "sifra")]
    field_options.append(("Kolicina" if kind == "sales" else "Stanje", "kolicina" if kind == "sales" else "stanje"))
    field_options.append(("Vrijednost", "vrijednost"))
    field_labels = [label for label, _ in field_options]
    field_map = {label: key for label, key in field_options}

    active_field = tk.StringVar(value=field_labels[0])
    field_combo = ctk.CTkComboBox(
        move_frame, values=field_labels, width=140, variable=active_field
    )
    field_combo.pack(side="left", padx=6, pady=6)

    ctk.CTkLabel(move_frame, text="Korak:").pack(
        side="left", padx=(10, 6), pady=6
    )
    step_var = tk.StringVar(value="1")
    step_combo = ctk.CTkComboBox(
        move_frame, values=["1", "2", "5", "10", "25", "50", "100"], width=70, variable=step_var
    )
    step_combo.pack(side="left", padx=6, pady=6)

    var_map = {
        "sifra": var_sifra,
        "kolicina": var_qty if kind == "sales" else None,
        "stanje": var_qty if kind != "sales" else None,
        "vrijednost": var_val,
    }

    def set_vars_from_offsets(offsets):
        updating_vars["active"] = True
        var_sifra.set(offset_to_label(offsets.get("sifra", 0)))
        if kind == "sales":
            var_qty.set(offset_to_label(offsets.get("kolicina", 0)))
        else:
            var_qty.set(offset_to_label(offsets.get("stanje", 0)))
        var_val.set(offset_to_label(offsets.get("vrijednost", 0)))
        updating_vars["active"] = False

    def update_offsets_from_vars():
        offsets = offsets_by_sheet.setdefault(active_key(), {"sifra": 0, "kolicina": 0, "stanje": 0, "vrijednost": 0})
        offsets["sifra"] = offset_map.get(var_sifra.get(), 0)
        if kind == "sales":
            offsets["kolicina"] = offset_map.get(var_qty.get(), 0)
        else:
            offsets["stanje"] = offset_map.get(var_qty.get(), 0)
        offsets["vrijednost"] = offset_map.get(var_val.get(), 0)

    def adjust_offset(direction):
        key = field_map.get(active_field.get(), "sifra")
        var = var_map.get(key, var_sifra)
        if var is None:
            return
        current = offset_map.get(var.get(), 0)
        try:
            step = int(step_var.get())
        except ValueError:
            step = 1
        new_offset = current + (direction * step)
        new_offset = max(-100, min(100, new_offset))
        var.set(offset_to_label(new_offset))

    btn_left = ctk.CTkButton(
        move_frame, text="Pomjeri lijevo", command=lambda: adjust_offset(-1)
    )
    btn_left.pack(side="left", padx=6, pady=6)

    btn_right = ctk.CTkButton(
        move_frame, text="Pomjeri desno", command=lambda: adjust_offset(1)
    )
    btn_right.pack(side="left", padx=6, pady=6)

    def reset_offsets_current_sheet():
        offsets_by_sheet[active_key()] = {"sifra": 0, "kolicina": 0, "stanje": 0, "vrijednost": 0}
        set_vars_from_offsets(offsets_by_sheet[active_key()])
        update_offsets_from_vars()
        refresh_preview()
        update_confirm_label()

    btn_reset = ctk.CTkButton(
        move_frame, text="Resetuj offsete", fg_color="#555555", hover_color="#444444", command=reset_offsets_current_sheet
    )
    btn_reset.pack(side="left", padx=6, pady=6)

    table_frame = ctk.CTkFrame(top)
    table_frame.pack(fill="both", expand=True, padx=10, pady=10)

    cols = ["Sifra", "Kolicina" if kind == "sales" else "Stanje", "Vrijednost"]
    tree = ttk.Treeview(table_frame, columns=cols, show="headings")
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=160, anchor="w")
    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True)
    vsb.pack(side="right", fill="y")

    count_label = ctk.CTkLabel(top, text="")
    count_label.pack(anchor="w", padx=12, pady=(0, 6))

    error_label = ctk.CTkLabel(top, text="", text_color="#ff6b6b")
    error_label.pack(anchor="w", padx=12, pady=(0, 6))

    btn_confirm_ref = {"btn": None}

    def update_confirm_label():
        btn = btn_confirm_ref["btn"]
        if btn is None:
            return
        if not sheet_names:
            btn.configure(text="Potvrdi")
            return
        unvisited = [s for s in sheet_names if s not in visited_sheets]
        btn.configure(text="Potvrdi" if not unvisited else "Dalje (sheet)")

    def refresh_preview(*_args):
        if updating_vars["active"]:
            return
        update_offsets_from_vars()
        if sheet_names:
            sheet_key = active_key()
            try:
                df_sheet = load_sheet_df(sheet_key)
            except Exception as e:
                error_label.configure(text=f"Greska pri citanju sheeta: {e}")
                return
            ok, msg, rows_new = parse_promet_df_to_rows(
                df_sheet, kind, col_offsets=offsets_by_sheet.get(sheet_key), add_vat=True
            )
        else:
            ok, msg, rows_new = parse_promet_df_to_rows(
                None, kind, col_offsets=offsets_by_sheet.get(0), add_vat=True
            )
        tree.delete(*tree.get_children())
        if ok:
            error_label.configure(text="")
            for row in rows_new:
                qty_val = row.get("Kolicina") if kind == "sales" else row.get("Stanje")
                tree.insert(
                    "",
                    "end",
                    values=(
                        row.get("Sifra", ""),
                        "" if qty_val is None else qty_val,
                        "" if row.get("Vrijednost") is None else f"{row.get('Vrijednost'):.2f}",
                    ),
                )
            info.configure(
                text=f"Fajl: {os.path.basename(path)} | Stavki: {len(rows_new)}"
            )
            count_label.configure(text=f"Prikazano {len(rows_new)} od {len(rows_new)} stavki.")
        else:
            error_label.configure(text=msg)
            info.configure(text=f"Fajl: {os.path.basename(path)} | Stavki: 0")
            count_label.configure(text="Prikazano 0 od 0 stavki.")

        total = 0
        for sheet_key in (sheet_names or [0]):
            try:
                df_sheet = load_sheet_df(sheet_key) if sheet_names else None
            except Exception:
                continue
            ok_total, _msg_total, rows_total = parse_promet_df_to_rows(
                df_sheet, kind, col_offsets=offsets_by_sheet.get(sheet_key), add_vat=True
            )
            if ok_total:
                total += len(rows_total)
        total_info.configure(text=f"Ukupno stavki (svi sheetovi): {total}")

    def on_sheet_change(*_args):
        if not sheet_names:
            return
        key = active_key()
        visited_sheets.add(key)
        set_vars_from_offsets(offsets_by_sheet.get(key, {}))
        refresh_preview()
        update_confirm_label()

    if active_sheet is not None:
        active_sheet.trace_add("write", on_sheet_change)

    for var in (var_sifra, var_qty, var_val):
        var.trace_add("write", refresh_preview)

    btn_frame = ctk.CTkFrame(top)
    btn_frame.pack(fill="x", padx=10, pady=(0, 10))

    def on_confirm():
        if sheet_names:
            unvisited = [s for s in sheet_names if s not in visited_sheets]
            if unvisited:
                active_sheet.set(unvisited[0])
                return
        combined_rows = []
        for sheet_key in (sheet_names or [0]):
            df_sheet = load_sheet_df(sheet_key) if sheet_names else None
            ok, msg, rows_new = parse_promet_df_to_rows(
                df_sheet, kind, col_offsets=offsets_by_sheet.get(sheet_key), add_vat=True
            )
            if not ok:
                error_label.configure(text=msg)
                if sheet_names:
                    active_sheet.set(sheet_key)
                    set_vars_from_offsets(offsets_by_sheet.get(sheet_key, {}))
                return
            combined_rows.extend(rows_new)
        result["ok"] = True
        result["rows"] = combined_rows
        top.destroy()

    def on_skip():
        top.destroy()

    btn_skip = ctk.CTkButton(
        btn_frame,
        text="Preskoci",
        fg_color="#666666",
        hover_color="#555555",
        command=on_skip,
    )
    btn_skip.pack(side="right", padx=10, pady=8)

    btn_confirm = ctk.CTkButton(
        btn_frame,
        text="Potvrdi",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=on_confirm,
    )
    btn_confirm.pack(side="right", padx=10, pady=8)
    btn_confirm_ref["btn"] = btn_confirm
    update_confirm_label()

    refresh_preview()
    top.protocol("WM_DELETE_WINDOW", on_skip)
    top.wait_window()
    return result


def _natural_sku_key(value: str):
    s = str(value).strip().upper()
    parts = re.split(r"(\d+)", s)
    key = []
    for part in parts:
        if part.isdigit():
            key.append(int(part))
        else:
            key.append(part)
    return key


def get_sheet_by_name_ci(wb, sheet_name: str):
    """Vrati sheet bez obzira na velika/mala slova, ili None."""
    target = sheet_name.lower()
    for name in wb.sheetnames:
        if name.lower() == target:
            return wb[name]
    return None


# -------------------------------------------------------------------------
# FUNKCIJE ZA PROCJENU TRANSPORTA
# -------------------------------------------------------------------------


def izracunaj_transport(
    total_kg,
    total_cbm,
    total_kom,
    avion_kg,
    avion_cbm,
    avion_kom,
    brod_kg,
    brod_cbm,
    transport_tip,
    metoda,
):
    """
    Rada logika transporta za v11.0:
    - transport_tip = "avion" ili "brod"
    - metoda = "kg", "cbm" ili "kom"
    """
    if transport_tip == "avion":
        if metoda == "kg":
            return total_kg * avion_kg
        elif metoda == "cbm":
            return total_cbm * avion_cbm
        elif metoda == "kom":
            return total_kom * avion_kom

    elif transport_tip == "brod":
        if metoda == "kg":
            return total_kg * brod_kg
        elif metoda == "cbm":
            return total_cbm * brod_cbm

    return 0.0


# -------------------------------------------------------------------------
# PANTHEON KONVERZIJA  —  .XLS (XML/HTML) → ČISTI .XLSX
# -------------------------------------------------------------------------


def choose_pantheon_file():
    """Odabir Pantheon .xls fajla."""
    global pantheon_xls_path
    p = filedialog.askopenfilename(
        title="Odaberi Pantheon .xls",
        filetypes=[("Pantheon XLS", "*.xls"), ("All files", "*.*")],
    )
    if p:
        pantheon_xls_path = p
        pantheon_file_label.configure(text=os.path.basename(p))


# ---------------------------------------------------------------
# HTML-like XLS čitanje (fallback način)
# ---------------------------------------------------------------


def _read_html_like_xls(path: str) -> pd.DataFrame:
    """Čitanje .xls fajla koji je ustvari HTML tabela."""
    try:
        tables = pd.read_html(path, header=None)
        if not tables:
            raise ValueError("Nema HTML tabela u fajlu.")
        df = max(tables, key=lambda d: d.shape[0] * d.shape[1])
        return df.fillna("").astype(str)
    except Exception as e:
        raise ValueError(f"Neuspjelo HTML čitanje: {e}")


# ---------------------------------------------------------------
# XML Spreadsheet ML čitanje (originalni Pantheon format)
# ---------------------------------------------------------------


def _read_spreadsheetml(path: str) -> pd.DataFrame:
    """Čitanje Pantheon XLS XML SpreadsheetML formata."""
    try:
        tree = ET.parse(path)
        root = tree.getroot()
    except Exception as e:
        raise ValueError(f"Nije validan SpreadsheetML: {e}")

    # XML namespace
    ss_ns = "{" + "urn:schemas-microsoft-com:office:spreadsheet" + "}"
    table = root.find(f".//{ss_ns}Table")
    if table is None:
        raise ValueError("Neuspjelo parsiranje Pantheon XML.")

    rows_data = []
    max_cols = 0

    for row in table.findall(f"{ss_ns}Row"):
        row_vals = []
        current_col = 0
        for cell in row.findall(f"{ss_ns}Cell"):

            idx_attr = cell.attrib.get(f"{ss_ns}Index") or cell.attrib.get("ss:Index")
            if idx_attr:
                try:
                    target = int(idx_attr) - 1
                    while current_col < target:
                        row_vals.append("")
                        current_col += 1
                except:
                    pass

            data_elem = cell.find(f"{ss_ns}Data")
            val = (
                str(data_elem.text).strip()
                if (data_elem is not None and data_elem.text)
                else ""
            )
            row_vals.append(val)
            current_col += 1

        if row_vals:
            max_cols = max(max_cols, len(row_vals))
            rows_data.append(row_vals)

    for r in rows_data:
        if len(r) < max_cols:
            r += [""] * (max_cols - len(r))

    return pd.DataFrame(rows_data)


# ---------------------------------------------------------------
# AUTODETEKCIJA KOLONA SIFRA / KOLICINA
# ---------------------------------------------------------------


def _score_code_column(series: pd.Series):
    """Provjera vjerovatnoće da je kolona sifra artikla."""
    count = 0
    for s in series.astype(str):
        if re.match(r"[A-Za-z]{1,4}-?\d+", s.strip()):
            count += 1
    return count / len(series)


def _score_qty_column(series: pd.Series):
    """Provjera vjerovatnoće da je kolona kolicina (numerička)."""
    count = 0
    for s in series:
        try:
            if float(s) != 0:
                count += 1
        except:
            pass
    return count / len(series)


def _autodetect_columns(df: pd.DataFrame):
    """Automatsko prepoznavanje kolone šifre i kolone količine."""
    best_code, best_qty = 0, 0
    code_col, qty_col = 0, 1

    for i in range(df.shape[1]):
        sc = _score_code_column(df.iloc[:, i])
        sq = _score_qty_column(df.iloc[:, i])
        if sc > best_code:
            best_code = sc
            code_col = i
        if sq > best_qty:
            best_qty = sq
            qty_col = i

    return code_col, qty_col


# ---------------------------------------------------------------
# KONVERZIJA PANTHEON → ČISTI .XLSX
# ---------------------------------------------------------------


def convert_pantheon():
    """Glavna funkcija konverzije Pantheon .xls → .xlsx (prodaja + stanje)."""
    global file_path, pantheon_xls_path

    if not pantheon_xls_path:
        messagebox.showerror("Greška", "Prvo odaberite Pantheon .xls fajl.")
        return

    try:
        # 1) probaj XML Spreadsheet ML
        try:
            df_raw = _read_spreadsheetml(pantheon_xls_path)
            src = "xml"
        except:
            df_raw = _read_html_like_xls(pantheon_xls_path)
            src = "html"

        df_raw = df_raw.fillna("").astype(str)

        # autodetekcija kolona (default E/O)
        code_idx, qty_idx = 4, 14
        if not (
            _score_code_column(df_raw.iloc[:, code_idx]) > 0.25
            and _score_qty_column(df_raw.iloc[:, qty_idx]) > 0.50
        ):
            code_idx, qty_idx = _autodetect_columns(df_raw)

        df = df_raw.iloc[:, [code_idx, qty_idx]].copy()
        df.columns = ["Sifra", "Kolicina_raw"]
        df["Sifra"] = df["Sifra"].astype(str).apply(auto_format_sifra)

        # čišćenje nepotrebnih redova
        df = df[df["Sifra"].str.strip() != ""]
        df = df[df["Sifra"].str.lower() != "ident"]

        # Pantheon količine ×100 → realne količine
        df["Kolicina"] = df["Kolicina_raw"].apply(lambda x: safe_float(x) / 100.0)
        df = df[["Sifra", "Kolicina"]]

        # kreiranje stanje sheet-a (sve stavke sa 0)
        df_stanje = df.copy()
        df_stanje["Stanje"] = 0

        # output .xlsx
        out = pantheon_xls_path.replace(".xls", "_clean.xlsx")
        wb = Workbook()
        ws1 = wb.active
        if ws1 is None:
            raise RuntimeError("Ne mogu dobiti aktivni sheet.")
        ws1.title = "prodaja"
        ws1.append(["Sifra", "Kolicina"])

        for _, r in df.iterrows():
            ws1.append([r["Sifra"], float(r["Kolicina"])])

        ws2 = wb.create_sheet("stanje")
        ws2.append(["Sifra", "Stanje"])
        for _, r in df_stanje.iterrows():
            ws2.append([r["Sifra"], 0])

        wb.save(out)
        file_path = out
        messagebox.showinfo("Uspjeh", f"Konverzija završena.\nSačuvan fajl: {out}")

    except Exception as e:
        messagebox.showerror("Greška", str(e))


# -------------------------------------------------------------------------
# ANALIZA PRODAJE — "Ukupna prodaja" ZA V11.0
# -------------------------------------------------------------------------


def analyze_sales():
    global file_path

    if not file_path:
        messagebox.showerror("Greška", "Prvo otvorite fajl za analizu.")
        return

    if not os.path.exists(file_path):
        messagebox.showerror("Greska", f"Fajl ne postoji: {file_path}")
        return
    try:
        wb = load_workbook(file_path)
    except PermissionError:
        messagebox.showerror(
            "Greška",
            "Ne mogu pristupiti file-u, ako je već otvoren u excel-u molim izgasite ga",
        )
        return
    except Exception:
        messagebox.showerror("Greška", "Ne mogu otvoriti fajl.")
        return

    ws = get_sheet_by_name_ci(wb, "prodaja")
    if ws is None:
        messagebox.showerror("Greška", "Nedostaje sheet 'prodaja'.")
        return
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]

    df = pd.DataFrame(data, columns=headers)
    df = normalize_columns(df)
    df["Sifra"] = df["Sifra"].astype(str).apply(auto_format_sifra)
    df["Kolicina"] = df["Kolicina"].apply(safe_float)

    # ukloni prazne
    df = df[df["Sifra"] != ""]
    df = df[df["Kolicina"] > 0]

    df_stanje = find_stanje_df(wb)
    stanje_map = {}
    if df_stanje is not None and not df_stanje.empty:
        df_stanje = df_stanje.groupby("Sifra", as_index=False)["Stanje"].max()
        stanje_map = dict(zip(df_stanje["Sifra"], df_stanje["Stanje"]))
    df["Stanje"] = df["Sifra"].map(stanje_map).fillna(0.0).apply(safe_float)

    # Filter datuma (ako postoji kolona Datum) + spremi period za nabavku
    try:
        start_date = pd.to_datetime(sales_date_from.get())
        end_date = pd.to_datetime(sales_date_to.get())
        period_days = max(1, (end_date - start_date).days + 1)
        if "Datum" in df.columns:
            df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce")
            df = df[(df["Datum"] >= start_date) & (df["Datum"] <= end_date)]
    except:
        messagebox.showerror("Greska", "Neispravan period (datum OD/DO).")
        return

    if df.empty:
        messagebox.showinfo("Info", "Nema prodaje u fajlu.")
        return

    global usd_bam_rate
    rate = 0.0
    if "usd_bam_entry" in globals():
        rate = safe_float(usd_bam_entry.get(), 0.0)
    if rate <= 0:
        rate = safe_float(usd_bam_rate, 0.0)
    if rate <= 0:
        messagebox.showerror(
            "Greska", "Unesite USD->KM kurs ili ucitajte sa CBBiH."
        )
        return
    usd_bam_rate = rate
    save_settings()

    # kategorije (u analizi prodaje custom artikli ostaju u originalnim kategorijama)
    df["Kategorija"] = df["Sifra"].apply(
        lambda s: kategorija_za_sifru(s, allow_custom=False)
    )

    added_mp = 0
    for sku in df["Sifra"].unique():
        if sku in mp_cijene:
            continue
        cat = kategorija_za_sifru(sku, allow_custom=False)
        if cat in MP_PRICE_BY_CATEGORY:
            mp_cijene[sku] = float(MP_PRICE_BY_CATEGORY.get(cat, 0.0))
            added_mp += 1
    if added_mp:
        save_mp_cjenovnik()

    def _build_category_mp_fallback():
        cat_prices = {}
        for sku, price in mp_cijene.items():
            p = safe_float(price, 0.0)
            if p <= 0:
                continue
            cat = kategorija_za_sifru(sku, allow_custom=False)
            cat_prices.setdefault(cat, []).append(p)
        return {cat: sum(vals) / len(vals) for cat, vals in cat_prices.items()}

    cat_mp_fallback = _build_category_mp_fallback()

    def _mp_price_for_sku(s):
        cat = kategorija_za_sifru(s, allow_custom=False)
        sku = s.strip().upper()
        kalk_price = KALK_MPC_W_BY_SKU.get(sku)
        if kalk_price is not None:
            kalk_price = safe_float(kalk_price, 0.0)
            if kalk_price > 0:
                return float(kalk_price)
        if cat in MP_PRICE_BY_CATEGORY:
            return float(MP_PRICE_BY_CATEGORY.get(cat, 0.0))
        if sku in mp_cijene:
            return float(mp_cijene.get(sku, 0.0))
        return float(cat_mp_fallback.get(cat, 0.0))

    df["MP_cijena"] = df["Sifra"].apply(_mp_price_for_sku)
    df["Vrijednost"] = df["Kolicina"] * df["MP_cijena"]
    df[["Nabavna_val", "Nabavna_is_km"]] = df["Sifra"].apply(
        lambda s: pd.Series(nabavna_value_and_currency(s))
    )
    df["Nabavna_KM"] = df.apply(
        lambda r: float(r["Nabavna_val"]) if r["Nabavna_is_km"] else float(r["Nabavna_val"]) * usd_bam_rate,
        axis=1,
    )
    df["Nabavna_usd"] = df.apply(
        lambda r: float(r["Nabavna_val"]) / usd_bam_rate if r["Nabavna_is_km"] and usd_bam_rate else float(r["Nabavna_val"]),
        axis=1,
    )
    df["COGS_KM"] = df["Kolicina"] * df["Nabavna_KM"]
    df["VAT_rate"] = df["Sifra"].apply(
        lambda s: KALK_VAT_BY_SKU.get(str(s).strip().upper(), 17.0)
    )
    def _mp_wo_vat_for_row(r):
        sku = str(r["Sifra"]).strip().upper()
        vat_rate = float(r["VAT_rate"])
        fallback = (
            float(r["MP_cijena"]) / (1.0 + vat_rate / 100.0)
            if vat_rate > 0
            else float(r["MP_cijena"])
        )
        kalk_wo = KALK_MPC_WO_BY_SKU.get(sku)
        if kalk_wo is None:
            return fallback
        kalk_wo = safe_float(kalk_wo, 0.0)
        return float(kalk_wo) if kalk_wo > 0 else fallback

    df["MP_wo_vat"] = df.apply(_mp_wo_vat_for_row, axis=1)
    df["Neto_prihod"] = df["Kolicina"] * df["MP_wo_vat"]

    # -----------------------------------------------------------------
    # UPIS U EXCEL - sheet "Kalkulacije cijene" (prosjek nabavnih cijena)
    # -----------------------------------------------------------------

    if "Kalkulacije cijene" in wb.sheetnames:
        del wb["Kalkulacije cijene"]
    ws_kalk = wb.create_sheet("Kalkulacije cijene")
    ws_kalk.append(
        [
            "Sifra",
            "Avg nabavna (KM)",
            "Avg MPC bez PDV (KM)",
            "Avg MPC s PDV (KM)",
            "Avg PDV (%)",
            "Sum qty",
            "Count",
            "Last doc date",
            "Last doc no",
            "Last file",
        ]
    )

    kalk_items = load_kalk_json(KALK_NABAVNE_PATH)
    for sku, item in sorted(kalk_items.items()):
        ws_kalk.append(
            [
                str(sku).strip().upper(),
                float(item.get("avg_nabavna", 0.0)),
                float(item.get("avg_mpc_wo_vat", 0.0)),
                float(item.get("avg_mpc_w_vat", 0.0)),
                float(item.get("avg_vat_rate", 0.0)),
                float(item.get("sum_qty", 0.0)),
                int(item.get("count", 0)),
                str(item.get("last_doc_date", "")),
                str(item.get("last_doc_no", "")),
                str(item.get("last_file_name", "")),
            ]
        )

    # -----------------------------------------------------------------
    # UPIS U EXCEL - sheet "Kalkulacije nedostaju" (SKU bez kalk. nabavne)
    # -----------------------------------------------------------------

    if "Kalkulacije nedostaju" in wb.sheetnames:
        del wb["Kalkulacije nedostaju"]
    ws_kalk_missing = wb.create_sheet("Kalkulacije nedostaju")
    ws_kalk_missing.append(
        ["Sifra", "Kategorija", "Status", "U JSON", "Nabavna definisana"]
    )

    kalk_skus = set(KALK_NABAVNE_BY_SKU.keys())
    missing_skus = sorted(
        sku for sku in df["Sifra"].unique() if sku not in kalk_skus
    )
    for sku in missing_skus:
        is_manual = sku in PRICE_BY_SKU
        status = (
            "Rucno dodano - nedostaje u kalkulaciji"
            if is_manual
            else "Nedostaje u kalkulaciji"
        )
        in_json = "DA" if sku in mp_cijene else "NE"
        nab_val, _is_km = nabavna_value_and_currency(sku)
        nab_defined = "DA" if nab_val and nab_val > 0 else "NE"
        ws_kalk_missing.append(
            [
                sku,
                kategorija_za_sifru(sku, allow_custom=False),
                status,
                in_json,
                nab_defined,
            ]
        )

    use_net_margin = (
        net_margin_transport_usd > 0
        or net_margin_customs_pct > 0
        or net_margin_marketing_bam > 0
        or net_margin_space_bam > 0
        or net_margin_labor_bam > 0
        or net_margin_accounting_bam > 0
        or net_margin_utilities_bam > 0
    )
    if use_net_margin:
        # Transport is already included in nabavna for sales analysis.
        transport_bam = 0.0
        transport_usd = 0.0
        base_bam = df["Nabavna_KM"]
        customs_bam_usd = pd.Series(0.0, index=df.index)
        customs_bam_km = pd.Series(0.0, index=df.index)
        customs_bam = pd.Series(
            np.where(df["Nabavna_is_km"], customs_bam_km, customs_bam_usd),
            index=df.index,
        )
        marketing_bam = np.where(
            df["Kategorija"].astype(str).str.strip().str.lower() == "ostalo",
            0.0,
            net_margin_marketing_bam,
        )
        total_fixed_month = (
            net_margin_space_bam
            + net_margin_labor_bam
            + net_margin_accounting_bam
            + net_margin_utilities_bam
        )
        total_fixed_period = total_fixed_month * (period_days / 30.0)
        non_ostalo_mask = (
            df["Kategorija"].astype(str).str.strip().str.lower() != "ostalo"
        )
        total_qty_non_ostalo = float(df.loc[non_ostalo_mask, "Kolicina"].sum())
        overhead_per_piece = (
            total_fixed_period / total_qty_non_ostalo
            if total_qty_non_ostalo > 0
            else 0.0
        )
        overhead_bam = np.where(non_ostalo_mask, overhead_per_piece, 0.0)
        net_cost_per_piece = (
            base_bam + transport_bam + customs_bam + marketing_bam + overhead_bam
        )
        df["Neto_trosak_KM"] = df["Kolicina"] * net_cost_per_piece

    # grupisanje po kategoriji
    df_group = (
        df.groupby("Kategorija")
        .agg(
            Kolicina=("Kolicina", "sum"),
            Vrijednost=("Vrijednost", "sum"),
            COGS_KM=("COGS_KM", "sum"),
            Neto_prihod=("Neto_prihod", "sum"),
        )
        .reset_index()
    )
    df_group["Bruto_marza"] = df_group["Vrijednost"] - df_group["COGS_KM"]
    df_group["Bruto_marza_pct"] = df_group["Bruto_marza"] / df_group[
        "Vrijednost"
    ].replace(0, 1)
    if use_net_margin:
        df_net_by_cat = (
            df.groupby("Kategorija")
            .agg(Neto_trosak_KM=("Neto_trosak_KM", "sum"))
            .reset_index()
        )
        df_group = df_group.merge(df_net_by_cat, on="Kategorija", how="left")
        net_revenue = df_group["Neto_prihod"]
        df_group["Neto_marza_pct"] = (
            (net_revenue - df_group["Neto_trosak_KM"])
            / net_revenue.replace(0, 1)
        )
    else:
        df_group["Neto_marza_pct"] = None
    total_net_margin = None
    total_net_margin_pct = None
    if use_net_margin:
        total_revenue = df["Neto_prihod"].sum()
        total_net_cost = float(df["Neto_trosak_KM"].sum())
        total_net_margin = total_revenue - total_net_cost
        total_net_margin_pct = (
            total_net_margin / total_revenue if total_revenue else 0.0
        )

    # % udjela
    total_k = df_group["Kolicina"].sum()
    total_v = df_group["Vrijednost"].sum()

    df_group["pct_kolicine"] = df_group["Kolicina"] / total_k
    df_group["pct_vrijednosti"] = df_group["Vrijednost"] / total_v

    # Sortiraj po vrijednosti (KM) opadajuce
    df_group["Vrijednost"] = pd.to_numeric(df_group["Vrijednost"], errors="coerce").fillna(0)
    df_group = df_group.sort_values("Vrijednost", ascending=False)

    # -----------------------------------------------------------------
    # UPIS U EXCEL — sheet "Ukupna prodaja"
    # -----------------------------------------------------------------

    if "Ukupna prodaja" in wb.sheetnames:
        del wb["Ukupna prodaja"]

    ws_out = wb.create_sheet("Ukupna prodaja")

    headers_out = [
        "Kategorija",
        "Kolicina",
        "Udio po količini (%)",
        "Vrijednost (KM)",
        "Udio po vrijednosti (%)",
        "COGS (KM)",
        "Bruto marza (KM)",
        "Bruto marza (%)",
        "Neto marza (%)",
    ]

    ws_out.append(headers_out)

    for _, r in df_group.iterrows():
        ws_out.append(
            [
                r["Kategorija"],
                int(r["Kolicina"]),
                float(r["pct_kolicine"]),
                round(float(r["Vrijednost"]), 2),
                float(r["pct_vrijednosti"]),
                round(float(r["COGS_KM"]), 2),
                round(float(r["Bruto_marza"]), 2),
                float(r["Bruto_marza_pct"]),
                float(r["Neto_marza_pct"]) if r["Neto_marza_pct"] is not None else None,
            ]
        )

    # format procenata
    for row in range(2, 2 + len(df_group)):
        ws_out.cell(row=row, column=3).number_format = "0.00%"
        ws_out.cell(row=row, column=5).number_format = "0.00%"
        ws_out.cell(row=row, column=8).number_format = "0.00%"
        ws_out.cell(row=row, column=9).number_format = "0.00%"

    # Neto marza red uklonjen (racuna se u koloni Neto marza (%))

    # -----------------------------------------------------------------
    # GRAFIKON — vrijednosti
    # -----------------------------------------------------------------

    chart = BarChart()
    chart.type = "col"
    chart.title = "Vrijednost prodaje po kategorijama"

    data_ref = Reference(ws_out, min_col=4, min_row=1, max_row=len(df_group) + 1)
    cats_ref = Reference(ws_out, min_col=1, min_row=2, max_row=len(df_group) + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.y_axis.title = "KM"
    chart.x_axis.title = "Kategorije"

    ws_out.add_chart(chart, "G10")

    write_meta_tmp(start_date, end_date, period_days)

    # -----------------------------------------------------------------
    # UPIS U EXCEL - sheet "Prodaja artikli" (po sifri)
    # -----------------------------------------------------------------

    if "Prodaja artikli" in wb.sheetnames:
        del wb["Prodaja artikli"]
    ws_items = wb.create_sheet("Prodaja artikli")

    agg_items = {
        "Kolicina": ("Kolicina", "sum"),
        "Stanje": ("Stanje", "max"),
        "MP_cijena": ("MP_cijena", "max"),
        "Vrijednost": ("Vrijednost", "sum"),
        "Nabavna_usd": ("Nabavna_usd", "max"),
        "COGS_KM": ("COGS_KM", "sum"),
    }
    if use_net_margin:
        agg_items["Neto_trosak_KM"] = ("Neto_trosak_KM", "sum")
        agg_items["Neto_prihod"] = ("Neto_prihod", "sum")
    df_items = df.groupby(["Kategorija", "Sifra"]).agg(**agg_items).reset_index()
    total_kolicina = df_items["Kolicina"].sum()
    total_vrijednost = df_items["Vrijednost"].sum() or 1
    df_items["Udio_kolicine"] = df_items["Kolicina"] / total_kolicina
    df_items["Udio_vrijednosti"] = df_items["Vrijednost"] / total_vrijednost
    df_items["Bruto_marza"] = df_items["Vrijednost"] - df_items["COGS_KM"]
    df_items["Bruto_marza_pct"] = df_items["Bruto_marza"] / df_items[
        "Vrijednost"
    ].replace(0, 1)
    if use_net_margin:
        net_revenue_items = df_items["Neto_prihod"]
        df_items["Neto_marza_pct"] = (
            (net_revenue_items - df_items["Neto_trosak_KM"])
            / net_revenue_items.replace(0, 1)
        )
    else:
        df_items["Neto_marza_pct"] = None

    df_items = df_items.sort_values(["Kategorija", "Vrijednost"], ascending=[True, False])

    headers_items = [
        "Kategorija",
        "Sifra",
        "Kolicina",
        "Trenutni lager",
        "Udio kolicine (%)",
        "MP cijena",
        "Vrijednost",
        "Udio vrijednosti (%)",
        "COGS (KM)",
        "Bruto marza (KM)",
        "Bruto marza (%)",
        "Neto marza (%)",
    ]
    ws_items.append(headers_items)
    ws_items.sheet_view.showGridLines = True

    ensure_default_fills(wb)
    custom_fill = PatternFill(
        start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid"
    )

    for _, r in df_items.iterrows():
        row_idx = ws_items.max_row + 1
        ws_items.append(
            [
                r["Kategorija"],
                r["Sifra"],
                int(r["Kolicina"]),
                int(r["Stanje"]),
                float(r["Udio_kolicine"]),
                float(r["MP_cijena"]),
                round(float(r["Vrijednost"]), 2),
                float(r["Udio_vrijednosti"]),
                round(float(r["COGS_KM"]), 2),
                round(float(r["Bruto_marza"]), 2),
                float(r["Bruto_marza_pct"]),
                float(r["Neto_marza_pct"]) if r["Neto_marza_pct"] is not None else None,
            ]
        )
        ws_items.cell(row=row_idx, column=5).number_format = "0.00%"
        ws_items.cell(row=row_idx, column=8).number_format = "0.00%"
        ws_items.cell(row=row_idx, column=11).number_format = "0.00%"
        ws_items.cell(row=row_idx, column=12).number_format = "0.00%"
        if r["Sifra"].upper() in CUSTOM_SKU_SET:
            for col in range(1, 13):
                ws_items.cell(row=row_idx, column=col).fill = custom_fill

    # -----------------------------------------------------------------
    # UPIS U EXCEL - sheet "Prodaja po grupama" (SKU udio unutar kategorije)
    # -----------------------------------------------------------------

    if "Prodaja po grupama" in wb.sheetnames:
        del wb["Prodaja po grupama"]
    ws_groups = wb.create_sheet("Prodaja po grupama")

    df_group_items = df_items.copy()
    cat_totals = df_group_items.groupby("Kategorija").agg(
        total_k=("Kolicina", "sum"), total_v=("Vrijednost", "sum"),
    )
    df_group_items = df_group_items.merge(
        cat_totals, left_on="Kategorija", right_index=True
    )
    df_group_items["Udio_kolicine_grupa"] = df_group_items["Kolicina"] / df_group_items["total_k"].replace(0, 1)
    df_group_items["Udio_vrijednosti_grupa"] = df_group_items["Vrijednost"] / df_group_items["total_v"].replace(0, 1)
    # sortiraj artikle unutar svake grupe po % kolicine (najvece -> najmanje)
    df_group_items = df_group_items.sort_values(
        ["Kategorija", "Udio_kolicine_grupa", "Kolicina"],
        ascending=[True, False, False],
    )

    headers_groups = [
        "Kategorija",
        "Sifra",
        "Kolicina",
        "Trenutni lager",
        "Udio kolicine u grupi (%)",
        "Vrijednost",
        "Udio vrijednosti u grupi (%)",
        "MP cijena",
        "Neto marza (%)",
    ]
    ws_groups.append(headers_groups)
    ws_groups.sheet_view.showGridLines = True

    net_pct_by_cat = {}
    if use_net_margin:
        net_pct_by_cat = dict(
            zip(df_group["Kategorija"], df_group["Neto_marza_pct"])
        )
    last_cat = None
    for _, r in df_group_items.iterrows():
        row_idx = ws_groups.max_row + 1
        net_pct = None
        if use_net_margin and r["Kategorija"] != last_cat:
            net_pct = net_pct_by_cat.get(r["Kategorija"])
        ws_groups.append(
            [
                r["Kategorija"],
                r["Sifra"],
                int(r["Kolicina"]),
                int(r["Stanje"]),
                float(r["Udio_kolicine_grupa"]),
                round(float(r["Vrijednost"]), 2),
                float(r["Udio_vrijednosti_grupa"]),
                float(r["MP_cijena"]),
                float(net_pct) if net_pct is not None else None,
            ]
        )
        ws_groups.cell(row=row_idx, column=5).number_format = "0.00%"
        ws_groups.cell(row=row_idx, column=7).number_format = "0.00%"
        ws_groups.cell(row=row_idx, column=9).number_format = "0.00%"
        if r["Sifra"].upper() in CUSTOM_SKU_SET:
            for col in range(1, 10):
                ws_groups.cell(row=row_idx, column=col).fill = custom_fill
        last_cat = r["Kategorija"]


    # -----------------------------------------------------------------
    # SNIMI FAJL
    # -----------------------------------------------------------------

    save_path = output_file_path or file_path
    try:
        wb.save(save_path)
    except PermissionError:
        messagebox.showerror(
            "Greška",
            "Ne mogu pristupiti file-u, ako je već otvoren u excel-u molim izgasite ga",
        )
        return
    except Exception:
        messagebox.showerror("Greška", "Ne mogu sačuvati fajl.")
        return

    messagebox.showinfo("Uspjeh", "Analiza uspješno završena.\nSheet: Ukupna prodaja")


# -------------------------------------------------------------------------
# PRORACUN PROMETA I ZALIHA
# -------------------------------------------------------------------------


def _build_category_mp_fallback():
    cat_prices = {}
    for sku, price in mp_cijene.items():
        p = safe_float(price, 0.0)
        if p <= 0:
            continue
        cat = kategorija_za_sifru(sku, allow_custom=False)
        cat_prices.setdefault(cat, []).append(p)
    return {cat: sum(vals) / len(vals) for cat, vals in cat_prices.items()}


def generate_promet_zaliha():
    global promet_sales_path, promet_stanje_path, promet_output_path

    if not promet_sales_path or not os.path.exists(promet_sales_path):
        messagebox.showerror("Greska", "Odaberi validan fajl ukupne prodaje.")
        return
    if not promet_stanje_path or not os.path.exists(promet_stanje_path):
        messagebox.showerror("Greska", "Odaberi validan fajl trenutnog stanja.")
        return

    ensure_kalk_nabavne_loaded(kalkulacije_folder)
    kalk_items = load_kalk_json(KALK_NABAVNE_PATH)
    kalk_qty_map = {
        str(sku).strip().upper(): safe_float(item.get("sum_qty", 0.0))
        for sku, item in (kalk_items or {}).items()
    }

    try:
        review_sales = show_promet_review(promet_sales_path, "sales")
    except Exception:
        messagebox.showerror("Greska", "Ne mogu otvoriti fajl prodaje.")
        return
    if not review_sales.get("ok"):
        return
    df_sales = pd.DataFrame(review_sales.get("rows", []))
    if df_sales.empty or "Sifra" not in df_sales.columns or "Kolicina" not in df_sales.columns:
        messagebox.showerror("Greska", "Nedostaju kolone Sifra i Kolicina u prodaji.")
        return
    if "Vrijednost" not in df_sales.columns:
        messagebox.showerror("Greska", "Nedostaje kolona Vrijednost u prodaji.")
        return
    df_sales["Sifra"] = df_sales["Sifra"].astype(str).apply(auto_format_sifra)
    df_sales["Kolicina"] = df_sales["Kolicina"].apply(safe_float)
    df_sales["Vrijednost"] = df_sales["Vrijednost"].apply(safe_float)
    df_sales = (
        df_sales.groupby("Sifra", as_index=False)[["Kolicina", "Vrijednost"]].sum()
    )
    sales_map = dict(zip(df_sales["Sifra"], df_sales["Kolicina"]))
    sales_value_map = dict(zip(df_sales["Sifra"], df_sales["Vrijednost"]))
    sales_avg_map = {
        str(sku).strip().upper(): (safe_float(val) / safe_float(qty) if safe_float(qty) > 0 else 0.0)
        for sku, qty, val in zip(df_sales["Sifra"], df_sales["Kolicina"], df_sales["Vrijednost"])
    }

    try:
        review_stanje = show_promet_review(promet_stanje_path, "stanje")
    except Exception:
        messagebox.showerror("Greska", "Ne mogu otvoriti fajl stanja.")
        return
    if not review_stanje.get("ok"):
        return
    df_stanje = pd.DataFrame(review_stanje.get("rows", []))
    if df_stanje.empty or "Sifra" not in df_stanje.columns or "Stanje" not in df_stanje.columns:
        messagebox.showerror("Greska", "Nedostaju kolone Sifra i Stanje u stanju.")
        return
    df_stanje["Sifra"] = df_stanje["Sifra"].astype(str).apply(auto_format_sifra)
    df_stanje["Stanje"] = df_stanje["Stanje"].apply(safe_float)
    df_stanje = df_stanje.groupby("Sifra", as_index=False)["Stanje"].max()
    stanje_map = dict(zip(df_stanje["Sifra"], df_stanje["Stanje"]))

    cat_mp_fallback = _build_category_mp_fallback()

    def _mp_price_for_sku(s):
        cat = kategorija_za_sifru(s, allow_custom=False)
        sku = s.strip().upper()
        kalk_price = KALK_MPC_W_BY_SKU.get(sku)
        if kalk_price is not None:
            kalk_price = safe_float(kalk_price, 0.0)
            if kalk_price > 0:
                return float(kalk_price)
        if cat in MP_PRICE_BY_CATEGORY:
            return float(MP_PRICE_BY_CATEGORY.get(cat, 0.0))
        if sku in mp_cijene:
            return float(mp_cijene.get(sku, 0.0))
        return float(cat_mp_fallback.get(cat, 0.0))

    all_skus = set(kalk_qty_map.keys()) | set(sales_map.keys()) | set(stanje_map.keys())
    if not all_skus:
        messagebox.showinfo("Info", "Nema podataka za proracun.")
        return

    rows = []
    for sku in sorted(all_skus, key=_natural_sku_key):
        kalk_qty = safe_float(kalk_qty_map.get(sku, 0.0))
        sales_qty = safe_float(sales_map.get(sku, 0.0))
        expected_qty = kalk_qty - sales_qty
        current_qty = safe_float(stanje_map.get(sku, 0.0))
        diff_qty = current_qty - expected_qty
        mp_price = safe_float(_mp_price_for_sku(sku), 0.0)
        diff_value = diff_qty * mp_price
        panth_avg_price = safe_float(sales_avg_map.get(sku, 0.0), 0.0)
        panth_diff_value = diff_qty * panth_avg_price
        status = "OK"
        if diff_qty < 0:
            status = "MANJAK"
        elif diff_qty > 0:
            status = "VISAK"
        note = ""
        if sku not in kalk_qty_map:
            note = "Nedostaje u kalkulacijama"
        category = kategorija_za_sifru(sku, allow_custom=True)
        rows.append(
            {
                "sku": sku,
                "kalk_qty": kalk_qty,
                "sales_qty": sales_qty,
                "expected_qty": expected_qty,
                "current_qty": current_qty,
                "diff_qty": diff_qty,
                "mp_price": mp_price,
                "diff_value": diff_value,
                "panth_avg_price": panth_avg_price,
                "panth_diff_value": panth_diff_value,
                "status": status,
                "note": note,
                "category": category,
                "is_custom": category == "Custom",
            }
        )

    base_path = promet_sales_path or promet_stanje_path
    out_dir = os.path.dirname(base_path) if base_path else app_base_dir()
    out_name = _default_promet_output_name(base_path)
    save_path = promet_output_path or os.path.join(out_dir, out_name)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Ne mogu dobiti aktivni sheet.")
    ws.title = "Analiza"

    header = [
        "Sifra",
        "Stanje iz kalkulacija",
        "Ukupna prodaja",
        "Ocekivano stanje",
        "Trenutno Pantheon stanje",
        "Razlika (manjak-/visak+)",
        "MP cijena",
        "MP vrijednost razlike",
        "Panth_srednja_cijena",
        "Panth vrijednost razlike",
        "Status",
        "Napomena",
    ]

    def write_sheet(target_ws, data_rows):
        total_diff_qty = 0.0
        total_diff_value = 0.0
        total_panth_value = 0.0
        target_ws.append(header)
        for r in data_rows:
            target_ws.append(
                [
                    r["sku"],
                    r["kalk_qty"],
                    r["sales_qty"],
                    r["expected_qty"],
                    r["current_qty"],
                    r["diff_qty"],
                    r["mp_price"],
                    r["diff_value"],
                    r["panth_avg_price"],
                    r["panth_diff_value"],
                    r["status"],
                    r["note"],
                ]
            )
            if r.get("is_custom"):
                for cell in target_ws[target_ws.max_row]:
                    cell.font = Font(color="0000FF")
            total_diff_qty += r["diff_qty"]
            total_diff_value += r["diff_value"]
            total_panth_value += r["panth_diff_value"]
        target_ws.append([])
        target_ws.append(
            [
                "UKUPNO",
                "",
                "",
                "",
                "",
                total_diff_qty,
                "",
                total_diff_value,
                "",
                total_panth_value,
                "",
                "",
            ]
        )

    write_sheet(ws, rows)

    ws_kosa = wb.create_sheet("Kosa")
    rows_kosa = [r for r in rows if r["category"] != "Ostalo"]
    write_sheet(ws_kosa, rows_kosa)

    ws_ostalo = wb.create_sheet("Ostalo")
    rows_ostalo = [r for r in rows if r["category"] == "Ostalo"]
    write_sheet(ws_ostalo, rows_ostalo)

    try:
        wb.save(save_path)
    except PermissionError:
        messagebox.showerror(
            "Greska",
            "Ne mogu pristupiti file-u, ako je otvoren u excel-u molim zatvorite ga",
        )
        return
    except Exception:
        messagebox.showerror("Greska", "Ne mogu sacuvati fajl.")
        return

    messagebox.showinfo("Uspjeh", f"Fajl sacuvan: {save_path}")


# -------------------------------------------------------------------------
# NABAVKA V11.0 – BUDŽET + POPUST + KG/CBM TRANSPORT (AVION/BROD)
# -------------------------------------------------------------------------


def calculate_procurement():
    global file_path

    if not file_path:
        messagebox.showerror("Greška", "Prvo učitaj fajl za analizu.")
        return

    load_path = output_file_path or file_path
    try:
        wb = load_workbook(load_path)
    except PermissionError:
        messagebox.showerror(
            "Greška",
            "Ne mogu pristupiti file-u, ako je već otvoren u excel-u molim izgasite ga",
        )
        return
    except Exception:
        messagebox.showerror("Greška", "Ne mogu otvoriti fajl.")
        return

    if (
        "Meta" not in wb.sheetnames
        and output_file_path
        and file_path
        and os.path.abspath(load_path) != os.path.abspath(file_path)
    ):
        try:
            wb_alt = load_workbook(file_path)
        except Exception:
            wb_alt = None
        if wb_alt is not None and "Meta" in wb_alt.sheetnames:
            wb = wb_alt

    ws_prodaja = get_sheet_by_name_ci(wb, "prodaja")
    if ws_prodaja is None:
        messagebox.showerror("Greška", "Nedostaje sheet prodaja.")
        return

    # ---------------------------------------------------------------------
    # 1) UČITAVANJE PODATAKA
    # ---------------------------------------------------------------------

    df_p = pd.DataFrame(
        ws_prodaja.iter_rows(values_only=True),
    )
    df_p.columns = df_p.iloc[0]
    df_p = df_p[1:]
    df_p = normalize_columns(df_p)
    df_p["Sifra"] = df_p["Sifra"].astype(str).apply(auto_format_sifra)
    df_p["Kolicina"] = df_p["Kolicina"].apply(safe_float)
    df_p = df_p[df_p["Kolicina"] > 0]

    df_s = find_stanje_df(wb)
    if df_s is None:
        messagebox.showerror(
            "Greška", "Nije pronađena kolona 'Stanje' ni u jednom sheet-u."
        )
        return

    if df_p.empty:
        messagebox.showinfo("Info", "Nema prodaje u fajlu.")
        return

    # kategorije
    df_p["Kategorija"] = df_p["Sifra"].apply(kategorija_za_sifru)

    # filtriranje po kategorijama
    rep_enabled = p_cb_rep_var.get()
    per_enabled = p_cb_per_var.get()
    custom_enabled = p_cb_custom_var.get()

    allowed = []
    if rep_enabled:
        allowed += [
            "Afro rep",
            "Ravni rep",
            "Repovi OPK",
            "Ariana repovi",
            "Kratki repovi",
            "Repovi trakica",
            "U klipse",
            "Ekstenzije",
            "Blowdry klipse",
            "Dugi repovi",
        ]
    if per_enabled:
        allowed += ["Premium perike", "Klasične perike"]
    if custom_enabled:
        allowed += ["Custom"]

    if not allowed:
        messagebox.showerror("Greška", "Odaberite barem jednu kategoriju.")
        return

    df_p = df_p[df_p["Kategorija"].isin(allowed)]

    # period (preuzimamo iz privremenog fajla ili Meta sheeta)
    meta_tmp = read_meta_tmp()
    if meta_tmp:
        try:
            start_date = pd.to_datetime(meta_tmp.get("SalesPeriodStart"))
            end_date = pd.to_datetime(meta_tmp.get("SalesPeriodEnd"))
            period_days = int(meta_tmp.get("SalesPeriodDays"))
        except Exception:
            start_date = end_date = period_days = None
    else:
        start_date = end_date = period_days = None

    if start_date is None or end_date is None or period_days is None:
        if "Meta" not in wb.sheetnames:
            messagebox.showerror(
                "Greska",
                "Nedostaje period prodaje. Prvo uradi Ukupna prodaja.",
            )
            return
        ws_meta = wb["Meta"]
        try:
            start_date = pd.to_datetime(ws_meta.cell(row=1, column=2).value)
            end_date = pd.to_datetime(ws_meta.cell(row=2, column=2).value)
            period_days = int(ws_meta.cell(row=3, column=2).value)
        except Exception:
            messagebox.showerror("Greska", "Neispravan period prodaje.")
            return
    if period_days <= 0:
        messagebox.showerror("Greska", "Neispravan broj dana perioda.")
        return
    if "Datum" in df_p.columns:
        df_p["Datum"] = pd.to_datetime(df_p["Datum"], errors="coerce")
        df_p = df_p[(df_p["Datum"] >= start_date) & (df_p["Datum"] <= end_date)]

    # ---------------------------------------------------------------------
    # 2) AGREGACIJA I IZRAČUNI
    # ---------------------------------------------------------------------

    df_group = df_p.groupby("Sifra").agg(Kolicina=("Kolicina", "sum")).reset_index()

    # stanje merge
    df_s["key"] = df_s["Sifra"].str.lower()
    df_group["key"] = df_group["Sifra"].str.lower()
    df_group = df_group.merge(df_s[["key", "Stanje"]], on="key", how="left")
    df_group["Stanje"] = df_group["Stanje"].fillna(0)

    # kategorije
    df_group["Kategorija"] = df_group["Sifra"].apply(kategorija_za_sifru)
    custom_mask = df_group["Sifra"].str.upper().isin(CUSTOM_SKU_SET)

    # dnevna prodaja (na osnovu ukupne količine / 60 dana)
    df_group["Dnevna_prodaja"] = df_group["Kolicina"] / float(period_days)

    # % udjela po količini
    total_k = df_group["Kolicina"].sum()
    if total_k <= 0:
        messagebox.showerror("Greška", "Ukupna prodaja je 0.")
        return
    df_group["pct"] = df_group["Kolicina"] / total_k

    # nabavne cijene (kalkulacije su u KM, ostalo u USD)
    df_group[["Nabavna_val", "Nabavna_is_km"]] = df_group["Sifra"].apply(
        lambda s: pd.Series(nabavna_value_and_currency(s))
    )
    if df_group["Nabavna_is_km"].any() and usd_bam_rate <= 0:
        messagebox.showerror(
            "Greska",
            "Postavite USD->KM kurs da bi se KM cijene iz kalkulacija pretvorile u USD.",
        )
        return
    df_group["Cijena_artikla"] = df_group.apply(
        lambda r: float(r["Nabavna_val"]) / usd_bam_rate
        if r["Nabavna_is_km"]
        else float(r["Nabavna_val"]),
        axis=1,
    )

    # popust
    try:
        popust = safe_float(p_discount_entry.get(), 0.0)
    except:
        popust = 0.0

    if popust < 0:
        popust = 0
    if popust > 90:
        popust = 90

    df_group["Cijena_sa_popustom"] = (
        df_group["Cijena_artikla"] * (1 - popust / 100.0)
    ).round(2)

    # trajanje
    try:
        trajanje = safe_int(p_target_days_entry.get(), 60)
    except:
        trajanje = 60

    if trajanje < 0:
        trajanje = 0
    dani_dostave = 35

    # preporuka nabavke po prodaji (bez ogranicenja 50)
    df_group["Nabavka_po_prodaji"] = (
        trajanje * df_group["Dnevna_prodaja"]
        + dani_dostave * df_group["Dnevna_prodaja"]
        - df_group["Stanje"]
    ).apply(lambda x: int(max(x, 0)))

    df_group["Minimalna_nabavka"] = (
        trajanje * df_group["Dnevna_prodaja"]
        + dani_dostave * df_group["Dnevna_prodaja"]
        - df_group["Stanje"]
    ).apply(lambda x: int(max(x, 0)))
    df_group.loc[custom_mask, "Minimalna_nabavka"] = df_group.loc[
        custom_mask, "Minimalna_nabavka"
    ].apply(round_custom_qty)

    # budžet
    try:
        budzet = safe_float(p_budget_entry.get(), 0.0)
    except:
        budzet = 0.0

    if budzet < 0:
        budzet = 0

    ignore_budget = p_ignore_budget_var.get()

    if ignore_budget:
        df_group["Kolicina_po_budzetu"] = df_group["Minimalna_nabavka"]
        df_group["Kolicina_final"] = df_group["Minimalna_nabavka"]
    else:
        df_group["Kolicina_po_budzetu_float"] = (
            budzet * df_group["pct"] / df_group["Cijena_sa_popustom"]
        ).fillna(0.0)

        df_group["Kolicina_po_budzetu"] = (
            df_group["Kolicina_po_budzetu_float"].round().astype(int)
        )

        df_group["Kolicina_final"] = df_group[
            ["Minimalna_nabavka", "Kolicina_po_budzetu"]
        ].max(axis=1)

    # za custom zaokruživanje i min 50
    df_group.loc[custom_mask, "Kolicina_po_budzetu"] = df_group.loc[
        custom_mask, "Kolicina_po_budzetu"
    ].apply(round_custom_qty)
    df_group.loc[custom_mask, "Kolicina_final"] = df_group.loc[
        custom_mask, "Kolicina_final"
    ].apply(round_custom_qty)

    # ---------------------------------------------------------
    # 3) ±5% prilagođavanje budžeta
    # ---------------------------------------------------------

    lower = budzet * 0.95
    upper = budzet * 1.05

    def cost_articles(df_loc):
        return float((df_loc["Kolicina_final"] * df_loc["Cijena_sa_popustom"]).sum())

    cost = cost_articles(df_group)
    order_desc = df_group.sort_values("pct", ascending=False).index.tolist()

    if not ignore_budget:
        # smanji ako treba
        safety = 0
        while cost > upper and safety < 10000:
            changed = False
            for idx in order_desc:
                cur = int(df_group.at[idx, "Kolicina_final"])
                minq = int(df_group.at[idx, "Minimalna_nabavka"])
                if cur > minq:
                    df_group.at[idx, "Kolicina_final"] = cur - 1
                    changed = True
                    cost = cost_articles(df_group)
                    if cost <= upper:
                        break
            if not changed:
                break
            safety += 1

        # poveća
        safety = 0
        while cost < lower and safety < 10000:
            for idx in order_desc:
                df_group.at[idx, "Kolicina_final"] = (
                    int(df_group.at[idx, "Kolicina_final"]) + 1
                )
                cost = cost_articles(df_group)
                if cost >= lower:
                    break
            safety += 1
            if lower <= cost <= upper:
                break

        # osiguraj minimalnu kolicinu za custom proizvode
        df_group.loc[custom_mask, "Kolicina_final"] = df_group.loc[
            custom_mask, "Kolicina_final"
        ].apply(round_custom_qty)

    # ---------------------------------------------------------
    # 4) IZRAČUN TEŽINE I CBM-a
    # ---------------------------------------------------------

    df_group["Tezina_kg"] = (
        df_group["Sifra"].apply(tezina_za_sifru) * df_group["Kolicina_final"]
    )
    df_group["CBM"] = df_group["Kolicina_final"] * CBM_PO_KOMADU

    # podijeli na perike i ostalo (ostalo dodatno na main i custom)
    perike_categories = {"Premium perike", "Klasične perike"}
    perike_mask = df_group["Kategorija"].isin(perike_categories)
    df_perike = df_group[perike_mask].copy()
    df_rest = df_group[~perike_mask].copy()

    custom_mask_rest = df_rest["Sifra"].str.upper().isin(CUSTOM_SKU_SET)
    df_main = df_rest[~custom_mask_rest].copy()
    df_custom = df_rest[custom_mask_rest].copy()

    # re-izracun pct za main da bude relativan na non-custom
    total_k_main = df_main["Kolicina"].sum()
    if total_k_main > 0:
        df_main["pct"] = df_main["Kolicina"] / total_k_main
    total_k_perike = df_perike["Kolicina"].sum()
    if total_k_perike > 0:
        df_perike["pct"] = df_perike["Kolicina"] / total_k_perike

    total_kg = float(df_main["Tezina_kg"].sum())
    total_cbm = float(df_main["CBM"].sum())
    total_kom = int(df_main["Kolicina_final"].sum())
    total_kg_custom = float(df_custom["Tezina_kg"].sum())
    total_cbm_custom = float(df_custom["CBM"].sum())
    total_kom_custom = int(df_custom["Kolicina_final"].sum())
    total_kg_perike = float(df_perike["Tezina_kg"].sum())
    total_cbm_perike = float(df_perike["CBM"].sum())
    total_kom_perike = int(df_perike["Kolicina_final"].sum())

    # ---------------------------------------------------------
    # 5) TRANSPORT OPCIJE (A2/B2 — KG i CBM)
    # ---------------------------------------------------------

    transport_tip = transport_var.get()  # avion / brod
    metoda = metoda_var.get()  # kg / cbm / kom

    avion_kg = safe_float(avion_kg_entry.get(), 0.0)
    avion_cbm = safe_float(avion_cbm_entry.get(), 0.0)
    avion_kom = safe_float(avion_kom_entry.get(), 1.30)
    brod_kg = safe_float(brod_kg_entry.get(), 0.0)
    brod_cbm = safe_float(brod_cbm_entry.get(), 0.0)

    if metoda == "kom" and transport_tip != "avion":
        messagebox.showerror("Greska", "Metoda 'po kom' je dostupna samo za avion.")
        return

    ukupno_transport = izracunaj_transport(
        total_kg,
        total_cbm,
        total_kom,
        avion_kg,
        avion_cbm,
        avion_kom,
        brod_kg,
        brod_cbm,
        transport_tip,
        metoda,
    )
    ukupno_transport_custom = izracunaj_transport(
        total_kg_custom,
        total_cbm_custom,
        total_kom_custom,
        avion_kg,
        avion_cbm,
        avion_kom,
        brod_kg,
        brod_cbm,
        transport_tip,
        metoda,
    )
    ukupno_transport_perike = izracunaj_transport(
        total_kg_perike,
        total_cbm_perike,
        total_kom_perike,
        avion_kg,
        avion_cbm,
        avion_kom,
        brod_kg,
        brod_cbm,
        transport_tip,
        metoda,
    )

    # ---------------------------------------------------------
    # 6) FINALNE VRIJEDNOSTI
    # ---------------------------------------------------------

    suma_artikli = cost_articles(df_main)
    suma_artikli_custom = cost_articles(df_custom)
    suma_artikli_perike = cost_articles(df_perike)
    total_val = suma_artikli + ukupno_transport
    total_val_custom = suma_artikli_custom + ukupno_transport_custom
    total_val_perike = suma_artikli_perike + ukupno_transport_perike
    total_artikli = total_kom
    total_artikli_perike = total_kom_perike

    # ---------------------------------------------------------
    # 7) SNIMI U EXCEL (sheet Nabavka)
    # ---------------------------------------------------------

    if "Nabavka" in wb.sheetnames:
        del wb["Nabavka"]

    ws = wb.create_sheet("Nabavka")

    red_fill = PatternFill("solid", fgColor="FFC7CE")

    ws.append(
        [
            "Sifra",
            "% prodaje",
            "Dnevna_prodaja",
            "Ukupna prodaja za period",
            "Stanje",
            "Cijena_sa_popustom",
            "Nabavka_po_prodaji",
            "Minimalna_nabavka",
            "Kolicina_po_budzetu",
            "Kolicina_za_nabavku",
            "Težina_kg",
            "Transport_tip",
        ]
    )
    # oboji cijelu kolonu Nabavka_po_prodaji
    for col in range(1, len(ws[1]) + 1):
        if ws.cell(row=1, column=col).value == "Nabavka_po_prodaji":
            ws.cell(row=1, column=col).fill = red_fill

    for _, r in df_main.iterrows():
        ws.append(
            [
                r["Sifra"],
                float(r["pct"]),
                float(r["Dnevna_prodaja"]),
                float(r["Kolicina"]),
                int(r["Stanje"]),
                float(r["Cijena_sa_popustom"]),
                int(r["Nabavka_po_prodaji"]),
                int(r["Minimalna_nabavka"]),
                int(r["Kolicina_po_budzetu"]),
                int(r["Kolicina_final"]),
                float(r["Tezina_kg"]),
                transport_tip,
            ]
        )
        # custom red fill u koloni Nabavka_po_prodaji
        if r["Sifra"].upper() in CUSTOM_SKU_SET:
            for col in range(1, len(ws[1]) + 1):
                if ws.cell(row=1, column=col).value == "Nabavka_po_prodaji":
                    ws.cell(ws.max_row, column=col).fill = red_fill

    for row in range(2, 2 + len(df_main)):
        ws.cell(row=row, column=2).number_format = "0.00%"
        # highlight kolone Nabavka_po_prodaji za custom (ovdje nema custom jer su izdvojeni)

    start = 2 + len(df_main) + 2

    ws.cell(start, 1, "Ukupna cijena artikala")
    ws.cell(start, 2, round(suma_artikli, 2))

    ws.cell(start + 1, 1, "Ukupan transport")
    ws.cell(start + 1, 2, round(ukupno_transport, 2))

    ws.cell(start + 2, 1, "TOTAL (artikli + transport)")
    ws.cell(start + 2, 2, round(total_val, 2))

    ws.cell(start + 3, 1, "Total komada")
    ws.cell(start + 3, 2, total_artikli)

    ws.cell(start + 4, 1, "Ukupna težina (kg)")
    ws.cell(start + 4, 2, round(total_kg, 2))

    ws.cell(start + 5, 1, "Ukupan volumen (CBM)")
    ws.cell(start + 5, 2, round(total_cbm, 4))

    ws.cell(start + 7, 1, "Transport metoda")
    if metoda == "kg":
        metoda_label = "Po KG"
    elif metoda == "cbm":
        metoda_label = "Po CBM"
    else:
        metoda_label = "Po kom (avion)"
    ws.cell(start + 7, 2, metoda_label)

    ws.cell(start + 8, 1, "Cijene transporta (Avion/Brod)")
    ws.cell(
        start + 8,
        2,
        f"Avion: {avion_kg} USD/kg, {avion_cbm} USD/CBM, {avion_kom} USD/kom; Brod: {brod_kg} USD/kg, {brod_cbm} USD/CBM",
    )

    # ---------------------------------------------------------
    # 7a) SNIMI U EXCEL (sheet Perike)
    # ---------------------------------------------------------

    if "Perike" in wb.sheetnames:
        del wb["Perike"]

    ws_perike = wb.create_sheet("Perike")
    ws_perike.append(
        [
            "Sifra",
            "% prodaje",
            "Dnevna_prodaja",
            "Ukupna prodaja za period",
            "Stanje",
            "Cijena_sa_popustom",
            "Nabavka_po_prodaji",
            "Minimalna_nabavka",
            "Kolicina_po_budzetu",
            "Kolicina_za_nabavku",
            "Te_ina_kg",
            "Transport_tip",
        ]
    )

    for col in range(1, len(ws_perike[1]) + 1):
        if ws_perike.cell(row=1, column=col).value == "Nabavka_po_prodaji":
            ws_perike.cell(row=1, column=col).fill = red_fill

    for _, r in df_perike.iterrows():
        ws_perike.append(
            [
                r["Sifra"],
                float(r["pct"]),
                float(r["Dnevna_prodaja"]),
                float(r["Kolicina"]),
                int(r["Stanje"]),
                float(r["Cijena_sa_popustom"]),
                int(r["Nabavka_po_prodaji"]),
                int(r["Minimalna_nabavka"]),
                int(r["Kolicina_po_budzetu"]),
                int(r["Kolicina_final"]),
                float(r["Tezina_kg"]),
                transport_tip,
            ]
        )

    for row in range(2, 2 + len(df_perike)):
        ws_perike.cell(row=row, column=2).number_format = "0.00%"

    start_p = 2 + len(df_perike) + 2

    ws_perike.cell(start_p, 1, "Ukupna cijena artikala")
    ws_perike.cell(start_p, 2, round(suma_artikli_perike, 2))

    ws_perike.cell(start_p + 1, 1, "Ukupan transport")
    ws_perike.cell(start_p + 1, 2, round(ukupno_transport_perike, 2))

    ws_perike.cell(start_p + 2, 1, "TOTAL (artikli + transport)")
    ws_perike.cell(start_p + 2, 2, round(total_val_perike, 2))

    ws_perike.cell(start_p + 3, 1, "Total komada")
    ws_perike.cell(start_p + 3, 2, total_artikli_perike)

    ws_perike.cell(start_p + 4, 1, "Ukupna tezina (kg)")
    ws_perike.cell(start_p + 4, 2, round(total_kg_perike, 2))

    ws_perike.cell(start_p + 5, 1, "Ukupan volumen (CBM)")
    ws_perike.cell(start_p + 5, 2, round(total_cbm_perike, 4))

    ws_perike.cell(start_p + 7, 1, "Transport metoda")
    if metoda == "kg":
        metoda_label = "Po KG"
    elif metoda == "cbm":
        metoda_label = "Po CBM"
    else:
        metoda_label = "Po kom (avion)"
    ws_perike.cell(start_p + 7, 2, metoda_label)

    ws_perike.cell(start_p + 8, 1, "Cijene transporta (Avion/Brod)")
    ws_perike.cell(
        start_p + 8,
        2,
        f"Avion: {avion_kg} USD/kg, {avion_cbm} USD/CBM, {avion_kom} USD/kom; Brod: {brod_kg} USD/kg, {brod_cbm} USD/CBM",
    )

    if "Custom" in wb.sheetnames:
        del wb["Custom"]
    df_custom = df_custom.copy()
    if not df_custom.empty:
        ws_custom = wb.create_sheet("Custom")
        ws_custom.append(
            [
                "Sifra",
                "% prodaje",
                "Dnevna_prodaja",
                "Ukupna prodaja za period",
                "Stanje",
                "Cijena_sa_popustom",
                "Nabavka_po_prodaji",
                "Minimalna_nabavka",
                "Kolicina_po_budzetu",
                "Kolicina_za_nabavku",
                "Tezina_kg",
                "Napomena",
            ]
        )
        for _, r in df_custom.iterrows():
            ws_custom.append(
                [
                    r["Sifra"],
                    float(r["pct"]),
                    float(r["Dnevna_prodaja"]),
                    float(r["Kolicina"]),
                    int(r["Stanje"]),
                    float(r["Cijena_sa_popustom"]),
                    int(r["Nabavka_po_prodaji"]),
                    int(r["Minimalna_nabavka"]),
                    int(r["Kolicina_po_budzetu"]),
                    int(r["Kolicina_final"]),
                    float(r["Tezina_kg"]),
                    "Custom (min 50 kom)",
                ]
            )
        for row in range(2, 2 + len(df_custom)):
            ws_custom.cell(row=row, column=2).number_format = "0.00%"
            # highlight Nabavka_po_prodaji za custom (svi redovi ovdje su custom)
            for col in range(1, len(ws_custom[1]) + 1):
                if ws_custom.cell(row=1, column=col).value == "Nabavka_po_prodaji":
                    ws_custom.cell(row=row, column=col).fill = red_fill

        # rezime za custom
        start_c = ws_custom.max_row + 2
        ws_custom.cell(start_c, 1, "Ukupna cijena artikala")
        ws_custom.cell(start_c, 2, round(cost_articles(df_custom), 2))

        ws_custom.cell(start_c + 1, 1, "Ukupan transport")
        ws_custom.cell(start_c + 1, 2, round(ukupno_transport_custom, 2))

        ws_custom.cell(start_c + 2, 1, "TOTAL (artikli + transport)")
        ws_custom.cell(start_c + 2, 2, round(cost_articles(df_custom) + ukupno_transport_custom, 2))

        ws_custom.cell(start_c + 3, 1, "Total komada")
        ws_custom.cell(start_c + 3, 2, int(df_custom["Kolicina_final"].sum()))

        ws_custom.cell(start_c + 4, 1, "Ukupna tezina (kg)")
        ws_custom.cell(start_c + 4, 2, round(float(df_custom["Tezina_kg"].sum()), 2))

        ws_custom.cell(start_c + 5, 1, "Ukupan volumen (CBM)")
        ws_custom.cell(start_c + 5, 2, round(float(df_custom["CBM"].sum()), 4))

        ws_custom.cell(start_c + 7, 1, "Transport metoda")
        ws_custom.cell(start_c + 7, 2, "Po kom (custom izdvojeno)")

        ws_custom.cell(start_c + 8, 1, "Cijene transporta (Avion/Brod)")
        ws_custom.cell(
            start_c + 8,
            2,
            f"Avion: {avion_kg} USD/kg, {avion_cbm} USD/CBM, {avion_kom} USD/kom; Brod: {brod_kg} USD/kg, {brod_cbm} USD/CBM",
        )

    ws.cell(start + 10, 1, "💰 Made by FEMMA finance department").font = Font(
        bold=True, color="FF5252"
    )

    save_path = output_file_path or file_path
    try:
        wb.save(save_path)
    except PermissionError:
        messagebox.showerror(
            "Greška",
            "Ne mogu pristupiti file-u, ako je već otvoren u excel-u molim izgasite ga",
        )
        return
    except Exception:
        messagebox.showerror("Greška", "Ne mogu sačuvati fajl.")
        return

    messagebox.showinfo(
        "Uspjeh",
        f"Nabavka sačuvana.\n"
        f"Artikli: {suma_artikli:.2f}\n"
        f"Transport: {ukupno_transport:.2f}\n"
        f"TOTAL: {total_val:.2f}\n"
        f"Težina: {total_kg:.2f} kg\n"
        f"Volumen: {total_cbm:.4f} CBM",
    )


# -------------------------------------------------------------------------
# GUI – SIDEBAR I NAVIGACIJA
# -------------------------------------------------------------------------


def clear_pages(app):
    """Skriva sve glavne stranice."""
    app.page_pantheon.pack_forget()
    app.page_sales.pack_forget()
    app.page_promet.pack_forget()
    app.page_proc.pack_forget()
    app.page_settings.pack_forget()


def show_pantheon(self):
    clear_pages(self)
    self.page_pantheon.pack(fill="both", expand=True)


def show_sales(self):
    clear_pages(self)
    self.page_sales.pack(fill="both", expand=True)


def show_proc(self):
    clear_pages(self)
    self.page_proc.pack(fill="both", expand=True)


# Proracun prometa i zaliha
def show_promet(self):
    clear_pages(self)
    self.page_promet.pack(fill="both", expand=True)


# Bind metode u klasu
ctk.CTk.show_pantheon = show_pantheon
ctk.CTk.show_sales = show_sales
ctk.CTk.show_proc = show_proc
ctk.CTk.show_promet = show_promet


def show_settings(self):
    clear_pages(self)
    self.page_settings.pack(fill="both", expand=True)


ctk.CTk.show_settings = show_settings


# -------------------------------------------------------------------------
# SIDEBAR
# -------------------------------------------------------------------------


def build_sidebar(self):

    sidebar = ctk.CTkFrame(self, width=200, corner_radius=0)
    sidebar.pack(side="left", fill="y")

    title = ctk.CTkLabel(
        sidebar, text="FEMMA 12.0", font=ctk.CTkFont(size=18, weight="bold")
    )
    title.pack(pady=(20, 10))

    # Buttons
    btn1 = ctk.CTkButton(
        sidebar, text="Pantheon konverzija", command=self.show_pantheon
    )
    btn1.pack(pady=8, fill="x", padx=10)

    btn2 = ctk.CTkButton(sidebar, text="Ukupna prodaja", command=self.show_sales)
    btn2.pack(pady=8, fill="x", padx=10)

    btn3 = ctk.CTkButton(
        sidebar, text="Proracun prometa i zaliha", command=self.show_promet
    )
    btn3.pack(pady=8, fill="x", padx=10)

    btn4 = ctk.CTkButton(sidebar, text="Nabavka", command=self.show_proc)
    btn4.pack(pady=8, fill="x", padx=10)

    btn5 = ctk.CTkButton(sidebar, text="Podesavanja", command=self.show_settings)
    btn5.pack(pady=8, fill="x", padx=10)

    footer = ctk.CTkLabel(
        sidebar, text="💰 FEMMA finance dept.", font=ctk.CTkFont(size=12, weight="bold")
    )
    footer.pack(side="bottom", pady=20)


# Bind sidebar u klasu
ctk.CTk.build_sidebar = build_sidebar
# -------------------------------------------------------------------------
# GUI – PANTHEON KONVERZIJA
# -------------------------------------------------------------------------


def build_pantheon_page(self, parent):

    frame = ctk.CTkFrame(parent)
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    header = ctk.CTkLabel(
        frame,
        text="Pantheon konverzija (.xls → čisti .xlsx)",
        font=ctk.CTkFont(size=22, weight="bold"),
    )
    header.pack(pady=(10, 15))

    # Odaberi fajl
    global pantheon_file_label
    pantheon_file_label = ctk.CTkLabel(
        frame, text="Nijedan fajl nije odabran", text_color=TEXT_MUTED
    )
    pantheon_file_label.pack(pady=6)

    btn_choose = ctk.CTkButton(
        frame, text="Odaberi Pantheon .xls fajl", command=choose_pantheon_file
    )
    btn_choose.pack(pady=12)

    btn_convert = ctk.CTkButton(
        frame,
        text="Konvertuj u čisti Excel",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=convert_pantheon,
    )
    btn_convert.pack(pady=20)

    # Info Footer
    info = ctk.CTkLabel(
        frame,
        text="* Konverzija automatski kreira sheet 'prodaja' i sheet 'stanje'.",
        text_color=TEXT_MUTED,
    )
    info.pack(pady=10)


# Bind this into class
ctk.CTk._build_pantheon_page = build_pantheon_page
# -------------------------------------------------------------------------
# GUI – UKUPNA PRODAJA
# -------------------------------------------------------------------------


def build_sales_page(self, parent):

    frame = ctk.CTkFrame(parent)
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    header = ctk.CTkLabel(
        frame,
        text="Ukupna prodaja — analiza MP vrijednosti",
        font=ctk.CTkFont(size=22, weight="bold"),
    )
    header.pack(pady=(10, 20))

    # Aktivni fajl
    global sales_file_label
    sales_file_label = ctk.CTkLabel(
        frame, text="Nije odabran fajl za analizu.", text_color=TEXT_MUTED
    )
    sales_file_label.pack(pady=4)

    # Dugme za učitavanje postojećeg .xlsx
    def choose_sales_file():
        global file_path
        p = filedialog.askopenfilename(
            title="Odaberi fajl za analizu",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if p:
            file_path = p
            sales_file_label.configure(text=os.path.basename(p))

    btn_choose = ctk.CTkButton(
        frame,
        text="Učitaj file za analizu",
        fg_color="#1f6aa5",
        hover_color="#185a8f",
        command=choose_sales_file,
    )
    btn_choose.pack(pady=10)

    # Output fajl (opcionalno)
    global output_file_label_sales
    output_file_label_sales = ctk.CTkLabel(
        frame, text="Output fajl: (isti kao input)", text_color=TEXT_MUTED
    )
    output_file_label_sales.pack(pady=4)

    def choose_output_file():
        p = filedialog.asksaveasfilename(
            title="Sacuvaj output fajl",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=os.path.dirname(file_path) if file_path else None,
            initialfile=_default_output_name(),
        )
        if p:
            _set_output_file_path(p)

    btn_output = ctk.CTkButton(
        frame,
        text="Odaberi output fajl",
        fg_color="#2b8a3e",
        hover_color="#237133",
        command=choose_output_file,
    )
    btn_output.pack(pady=6)
    # ===============================
    #   FILTER DATUMA — OD / DO
    # ===============================

    date_frame = ctk.CTkFrame(frame)
    date_frame.pack(fill="x", padx=20, pady=(5, 15))

    ctk.CTkLabel(
        date_frame,
        text="Period analize (OD - DO):",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).grid(row=0, column=0, columnspan=4, sticky="w", padx=5, pady=(5, 3))

    ctk.CTkLabel(date_frame, text="OD:").grid(row=1, column=0, sticky="e", padx=5)
    ctk.CTkLabel(date_frame, text="DO:").grid(row=1, column=2, sticky="e", padx=5)

    global sales_date_from, sales_date_to
    sales_date_from = DateEntry(date_frame, width=12)
    sales_date_to = DateEntry(date_frame, width=12)

    sales_date_from.grid(row=1, column=1, sticky="w", padx=5)
    sales_date_to.grid(row=1, column=3, sticky="w", padx=5)

    # Kurs USD -> KM (za COGS / bruto marzu)
    rate_frame = ctk.CTkFrame(frame)
    rate_frame.pack(fill="x", padx=20, pady=(0, 10))

    ctk.CTkLabel(
        rate_frame, text="USD -> KM kurs:", font=ctk.CTkFont(size=14, weight="bold")
    ).grid(row=0, column=0, sticky="w", padx=5, pady=6)

    global usd_bam_entry
    usd_bam_entry = ctk.CTkEntry(rate_frame, width=120)
    if usd_bam_rate > 0:
        usd_bam_entry.insert(0, f"{usd_bam_rate:.6f}")
    usd_bam_entry.grid(row=0, column=1, sticky="w", padx=5, pady=6)

    def fetch_rate():
        global usd_bam_rate
        try:
            rate = fetch_usd_bam_rate()
        except Exception as e:
            messagebox.showerror("Greska", f"Ne mogu ucitati kurs: {e}")
            return
        usd_bam_rate = rate
        usd_bam_entry.delete(0, tk.END)
        usd_bam_entry.insert(0, f"{rate:.6f}")
        save_settings()

    btn_rate = ctk.CTkButton(
        rate_frame,
        text="Ucitaj kurs (CBBiH)",
        fg_color="#5b2d90",
        hover_color="#4b2576",
        command=fetch_rate,
    )
    btn_rate.grid(row=0, column=2, padx=10, pady=6)

    # Dugme za analizu
    btn_analyze = ctk.CTkButton(
        frame,
        text="Analiziraj prodaju",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=analyze_sales,
    )
    btn_analyze.pack(pady=20)

    footer = ctk.CTkLabel(
        frame,
        text="* Kreira sheet 'Ukupna prodaja' u aktivnom fajlu.",
        text_color=TEXT_MUTED,
    )
    footer.pack(pady=8)


# Bind into class
ctk.CTk._build_sales_page = build_sales_page
# -------------------------------------------------------------------------
# GUI – PRORACUN PROMETA I ZALIHA
# -------------------------------------------------------------------------


def build_promet_page(self, parent):
    frame = ctk.CTkFrame(parent)
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    header = ctk.CTkLabel(
        frame,
        text="Proracun prometa i zaliha",
        font=ctk.CTkFont(size=22, weight="bold"),
    )
    header.pack(pady=(10, 20))

    # Prodaja file
    global promet_sales_label
    promet_sales_label = ctk.CTkLabel(
        frame, text="Nije odabran fajl ukupne prodaje.", text_color=TEXT_MUTED
    )
    promet_sales_label.pack(pady=4)

    def choose_promet_sales_file():
        global promet_sales_path
        p = filedialog.askopenfilename(
            title="Odaberi fajl ukupne prodaje",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if p:
            promet_sales_path = p
            promet_sales_label.configure(text=os.path.basename(p))

    btn_sales = ctk.CTkButton(
        frame,
        text="Odaberi fajl ukupne prodaje",
        fg_color="#1f6aa5",
        hover_color="#185a8f",
        command=choose_promet_sales_file,
    )
    btn_sales.pack(pady=10)

    # Stanje file
    global promet_stanje_label
    promet_stanje_label = ctk.CTkLabel(
        frame, text="Nije odabran fajl trenutnog stanja.", text_color=TEXT_MUTED
    )
    promet_stanje_label.pack(pady=4)

    def choose_promet_stanje_file():
        global promet_stanje_path
        p = filedialog.askopenfilename(
            title="Odaberi fajl trenutnog stanja",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if p:
            promet_stanje_path = p
            promet_stanje_label.configure(text=os.path.basename(p))

    btn_stanje = ctk.CTkButton(
        frame,
        text="Odaberi fajl trenutnog stanja",
        fg_color="#1f6aa5",
        hover_color="#185a8f",
        command=choose_promet_stanje_file,
    )
    btn_stanje.pack(pady=10)

    # Output file (opcionalno)
    global output_file_label_promet
    output_file_label_promet = ctk.CTkLabel(
        frame, text="Output fajl: (automatski)", text_color=TEXT_MUTED
    )
    output_file_label_promet.pack(pady=4)

    def choose_promet_output_file():
        base_path = promet_sales_path or promet_stanje_path
        p = filedialog.asksaveasfilename(
            title="Sacuvaj output fajl",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=os.path.dirname(base_path) if base_path else None,
            initialfile=_default_promet_output_name(base_path),
        )
        if p:
            _set_promet_output_path(p)

    btn_output = ctk.CTkButton(
        frame,
        text="Odaberi output fajl (opcionalno)",
        fg_color="#2b8a3e",
        hover_color="#237133",
        command=choose_promet_output_file,
    )
    btn_output.pack(pady=6)

    sep = ttk.Separator(frame, orient="horizontal")
    sep.pack(fill="x", padx=10, pady=(12, 6))

    # Historija artikla (kalkulacije)
    history_frame = ctk.CTkFrame(frame)
    history_frame.pack(fill="x", pady=(16, 6), padx=10)

    ctk.CTkLabel(
        history_frame,
        text="Historija artikla (kalkulacije):",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).grid(row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(8, 4))

    ctk.CTkLabel(history_frame, text="SKU:").grid(
        row=1, column=0, sticky="e", padx=10, pady=6
    )
    sku_entry = ctk.CTkEntry(history_frame, width=200)
    sku_entry.grid(row=1, column=1, sticky="w", padx=6, pady=6)

    def _doc_no_from_path(p):
        base = os.path.basename(p)
        m = re.search(r"FEMMA_(\d{2})-160-(\d{6})", base, re.IGNORECASE)
        if m:
            return f"{m.group(1)}-160-{m.group(2)}"
        return os.path.splitext(base)[0]

    def _fmt_num(val):
        try:
            num = float(val)
        except (TypeError, ValueError):
            return ""
        return f"{num:.4f}".rstrip("0").rstrip(".")

    def show_sku_history():
        sku = sku_entry.get().strip().upper()
        if not sku:
            messagebox.showerror("Greska", "Unesi SKU.")
            return
        cache = load_kalk_file_cache()
        if not cache:
            messagebox.showerror(
                "Greska",
                "Nema sacuvanih kalkulacija. Prvo ucitaj kalkulacije.",
            )
            return
        rows = []
        for path, entry in cache.items():
            items = entry.get("items", {})
            if not isinstance(items, dict):
                continue
            data = items.get(sku)
            if not data:
                continue
            sum_qty = safe_float(data.get("sum_qty", 0.0))
            if sum_qty <= 0:
                continue
            sum_nab_value = safe_float(data.get("sum_nab_value", 0.0))
            sum_mpc_w_value = safe_float(data.get("sum_mpc_w_value", 0.0))
            avg_nab = sum_nab_value / sum_qty if sum_qty > 0 else 0.0
            avg_mpc_w = sum_mpc_w_value / sum_qty if sum_qty > 0 else 0.0
            doc_no = _doc_no_from_path(path)
            try:
                mtime = entry.get("mtime")
                if mtime:
                    dt = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d")
                else:
                    dt = ""
            except Exception:
                dt = ""
            rows.append(
                (
                    doc_no,
                    sum_qty,
                    avg_nab,
                    avg_mpc_w,
                    dt,
                    os.path.basename(path),
                )
            )
        if not rows:
            messagebox.showinfo(
                "Info",
                f"Nema zapisa za SKU: {sku}",
            )
            return

        top = ctk.CTkToplevel(frame)
        top.title(f"Historija artikla - {sku}")
        top.geometry("900x520")
        top.grab_set()
        top.focus_set()

        header_lbl = ctk.CTkLabel(
            top,
            text=f"Historija artikla: {sku}",
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        header_lbl.pack(anchor="w", padx=12, pady=(12, 6))

        table_frame = ctk.CTkFrame(top)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = ("Dokument", "Kolicina", "Avg nabavna", "MP s PDV", "Datum", "Fajl")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            anchor = "center" if col != "Fajl" else "w"
            tree.column(col, width=140 if col != "Fajl" else 260, anchor=anchor)
        tree.tag_configure("separator", background="#202020")
        tree.tag_configure("total", background="#1f1f1f", foreground="#FFFFFF")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        for doc_no, qty, avg_nab, avg_mpc_w, dt, fname in sorted(rows):
            tree.insert(
                "",
                "end",
                values=(
                    doc_no,
                    _fmt_num(qty),
                    _fmt_num(avg_nab),
                    _fmt_num(avg_mpc_w),
                    dt,
                    fname,
                ),
            )

        total_qty = sum(qty for _, qty, *_ in rows)
        total_nab_value = sum(qty * avg_nab for _, qty, avg_nab, *_ in rows)
        total_mpc_value = sum(qty * avg_mpc_w for _, qty, _, avg_mpc_w, *_ in rows)
        avg_nab_total = total_nab_value / total_qty if total_qty > 0 else 0.0
        avg_mpc_total = total_mpc_value / total_qty if total_qty > 0 else 0.0
        tree.insert("", "end", values=("", "", "", "", "", ""), tags=("separator",))
        tree.insert(
            "",
            "end",
            values=(
                "UKUPNO",
                _fmt_num(total_qty),
                _fmt_num(avg_nab_total),
                _fmt_num(avg_mpc_total),
                "",
                "",
            ),
            tags=("total",),
        )

    btn_history = ctk.CTkButton(
        history_frame,
        text="Historija SKU",
        fg_color="#5b2d90",
        hover_color="#4b2576",
        command=show_sku_history,
    )
    btn_history.grid(row=1, column=2, sticky="w", padx=10, pady=6)

    # Dugme za generisanje izvjestaja
    btn_generate = ctk.CTkButton(
        frame,
        text="Generisi izvjestaj",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=generate_promet_zaliha,
    )
    btn_generate.pack(pady=20)

    info = ctk.CTkLabel(
        frame,
        text="* Output fajl se kreira kao novi Excel sa sheet-om 'Analiza'.",
        text_color=TEXT_MUTED,
    )
    info.pack(pady=8)


# Bind into class
ctk.CTk._build_promet_page = build_promet_page
# -------------------------------------------------------------------------
# GUI – NABAVKA (v11.0)
# -------------------------------------------------------------------------


def build_proc_page(self, parent):
    container = ctk.CTkFrame(parent)
    container.pack(fill="both", expand=True, padx=20, pady=20)

    frame = ctk.CTkFrame(container)
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    header = ctk.CTkLabel(
        frame,
        text="Nabavka - budzet + popust + transport",
        font=ctk.CTkFont(size=22, weight="bold"),
    )
    header.pack(pady=(10, 15))

    # ===============================
    #  SEKCIJA 1 — KATEGORIJE
    # ===============================

    sec_cat = ctk.CTkFrame(frame)
    sec_cat.pack(fill="x", pady=10, padx=8)

    ctk.CTkLabel(
        sec_cat, text="Odaberi kategorije:", font=ctk.CTkFont(size=14, weight="bold")
    ).pack(anchor="w", padx=10, pady=(10, 5))

    # Output fajl (opcionalno)
    global output_file_label_proc
    output_file_label_proc = ctk.CTkLabel(
        frame, text="Output fajl: (isti kao input)", text_color=TEXT_MUTED
    )
    output_file_label_proc.pack(pady=4)

    def choose_output_file_proc():
        p = filedialog.asksaveasfilename(
            title="Sacuvaj output fajl",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=os.path.dirname(file_path) if file_path else None,
            initialfile=_default_output_name(),
        )
        if p:
            _set_output_file_path(p)

    btn_output_proc = ctk.CTkButton(
        frame,
        text="Odaberi output fajl",
        fg_color="#2b8a3e",
        hover_color="#237133",
        command=choose_output_file_proc,
    )
    btn_output_proc.pack(pady=6)

    global p_cb_rep_var, p_cb_per_var, p_cb_custom_var
    p_cb_rep_var = tk.BooleanVar(value=True)
    p_cb_per_var = tk.BooleanVar(value=True)
    p_cb_custom_var = tk.BooleanVar(value=True)

    ctk.CTkCheckBox(
        sec_cat, text="Repovi / Klipse / Ekstenzije", variable=p_cb_rep_var
    ).pack(anchor="w", padx=14, pady=2)
    ctk.CTkCheckBox(
        sec_cat, text="Perike (Klasične / Premium)", variable=p_cb_per_var
    ).pack(
        anchor="w", padx=14, pady=2
    )
    ctk.CTkCheckBox(
        sec_cat, text="Custom proizvodi (min 50 kom)", variable=p_cb_custom_var
    ).pack(anchor="w", padx=14, pady=2)

    # ===============================
    #  SEKCIJA 2 — BUDŽET / POPUST
    # ===============================

    sec_form = ctk.CTkFrame(frame)
    sec_form.pack(fill="x", pady=10, padx=8)

    # Budžet
    ctk.CTkLabel(sec_form, text="Budžet (USD):").grid(
        row=0, column=0, padx=10, pady=8, sticky="e"
    )
    global p_budget_entry
    p_budget_entry = ctk.CTkEntry(sec_form, width=120)
    p_budget_entry.insert(0, "0")
    p_budget_entry.grid(row=0, column=1, padx=10, pady=8)

    # Popust
    ctk.CTkLabel(sec_form, text="Popust dobavljača (%):").grid(
        row=0, column=2, padx=10, pady=8, sticky="e"
    )
    global p_discount_entry
    p_discount_entry = ctk.CTkEntry(sec_form, width=120)
    p_discount_entry.insert(0, "0")
    p_discount_entry.grid(row=0, column=3, padx=10, pady=8)

    # Trajanje zaliha
    ctk.CTkLabel(sec_form, text="Cilj zaliha (dana):").grid(
        row=1, column=0, padx=10, pady=8, sticky="e"
    )
    global p_target_days_entry
    p_target_days_entry = ctk.CTkEntry(sec_form, width=120)
    p_target_days_entry.insert(0, "60")
    p_target_days_entry.grid(row=1, column=1, padx=10, pady=8)

    # Ignorisi budzet
    global p_ignore_budget_var
    p_ignore_budget_var = tk.BooleanVar(value=False)
    ctk.CTkCheckBox(
        sec_form, text="Ignorisi budzet (nabavi za cilj zaliha)", variable=p_ignore_budget_var
    ).grid(row=1, column=2, columnspan=2, padx=10, pady=8, sticky="w")

    # ===============================
    #  SEKCIJA 3 — TRANSPORT OPCIJE
    # ===============================

    sec_tr = ctk.CTkFrame(frame)
    sec_tr.pack(fill="x", pady=10, padx=8)

    ctk.CTkLabel(
        sec_tr, text="Transport opcije:", font=ctk.CTkFont(size=14, weight="bold")
    ).grid(row=0, column=0, columnspan=2, padx=10, pady=(10, 3), sticky="w")
    # Avion cijene
    ctk.CTkLabel(sec_tr, text="Avion (USD/kg):").grid(
        row=1, column=0, padx=10, pady=4, sticky="e"
    )
    global avion_kg_entry
    avion_kg_entry = ctk.CTkEntry(sec_tr, width=90)
    avion_kg_entry.insert(0, "0")
    avion_kg_entry.grid(row=1, column=1, padx=10, pady=4)

    ctk.CTkLabel(sec_tr, text="Avion (USD/CBM):").grid(
        row=1, column=2, padx=10, pady=4, sticky="e"
    )
    global avion_cbm_entry
    avion_cbm_entry = ctk.CTkEntry(sec_tr, width=90)
    avion_cbm_entry.insert(0, "0")
    avion_cbm_entry.grid(row=1, column=3, padx=10, pady=4)

    ctk.CTkLabel(sec_tr, text="Avion (USD/kom):").grid(
        row=2, column=0, padx=10, pady=4, sticky="e"
    )
    global avion_kom_entry
    avion_kom_entry = ctk.CTkEntry(sec_tr, width=90)
    avion_kom_entry.insert(0, "1.30")
    avion_kom_entry.grid(row=2, column=1, padx=10, pady=4)

    # Brod cijene
    ctk.CTkLabel(sec_tr, text="Brod (USD/kg):").grid(
        row=3, column=0, padx=10, pady=4, sticky="e"
    )
    global brod_kg_entry
    brod_kg_entry = ctk.CTkEntry(sec_tr, width=90)
    brod_kg_entry.insert(0, "0")
    brod_kg_entry.grid(row=3, column=1, padx=10, pady=4)

    ctk.CTkLabel(sec_tr, text="Brod (USD/CBM):").grid(
        row=3, column=2, padx=10, pady=4, sticky="e"
    )
    global brod_cbm_entry
    brod_cbm_entry = ctk.CTkEntry(sec_tr, width=90)
    brod_cbm_entry.insert(0, "0")
    brod_cbm_entry.grid(row=3, column=3, padx=10, pady=4)

    # Radio buttons (metoda: KG / CBM / KOM)
    ctk.CTkLabel(sec_tr, text="Metoda obracuna:", font=ctk.CTkFont(size=13)).grid(
        row=4, column=0, padx=10, pady=(12, 4), sticky="e"
    )

    global metoda_var
    metoda_var = tk.StringVar(value="kg")

    ctk.CTkRadioButton(sec_tr, text="po KG", variable=metoda_var, value="kg").grid(
        row=4, column=1, pady=4, padx=10, sticky="w"
    )
    ctk.CTkRadioButton(sec_tr, text="po CBM", variable=metoda_var, value="cbm").grid(
        row=4, column=2, pady=4, padx=10, sticky="w"
    )
    ctk.CTkRadioButton(
        sec_tr, text="po kom (avion)", variable=metoda_var, value="kom"
    ).grid(row=4, column=3, pady=4, padx=10, sticky="w")

    # Transport tip: avion / brod
    ctk.CTkLabel(sec_tr, text="Transport tip:", font=ctk.CTkFont(size=13)).grid(
        row=5, column=0, padx=10, pady=(12, 4), sticky="e"
    )

    global transport_var
    transport_var = tk.StringVar(value="avion")

    ctk.CTkRadioButton(
        sec_tr, text="Avion", variable=transport_var, value="avion"
    ).grid(row=5, column=1, pady=4, padx=10, sticky="w")
    ctk.CTkRadioButton(sec_tr, text="Brod", variable=transport_var, value="brod").grid(
        row=5, column=2, pady=4, padx=10, sticky="w"
    )

    # ===============================
    #  SEKCIJA 4 — DUGME ZA RAČUNANJE
    # ===============================

    btn_calc = ctk.CTkButton(
        frame,
        text="Izračunaj nabavku",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        height=40,
        command=calculate_procurement,
    )
    btn_calc.pack(pady=25)

    footer = ctk.CTkLabel(
        frame,
        text="* Kreira sheet 'Nabavka' u aktivnom Excel fajlu.",
        text_color=TEXT_MUTED,
    )
    footer.pack(pady=8)



# -------------------------------------------------------------------------
# GUI - PODESAVANJA (v11.0)
# -------------------------------------------------------------------------


def build_settings_page(self, parent):
    frame = ctk.CTkFrame(parent)
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    header = ctk.CTkLabel(
        frame,
        text="Podesavanja - artikli i kategorije",
        font=ctk.CTkFont(size=20, weight="bold"),
    )
    header.pack(pady=(10, 15))

    tabview = ctk.CTkTabview(frame)
    tabview.pack(fill="both", expand=True, padx=10, pady=10)

    tab_items = tabview.add("Artikli i kategorije")
    tab_prices = tabview.add("Cijene")
    tab_mp_prices = tabview.add("Promjena MP cijena")
    tab_mp_prices_cat = tabview.add("Promjena MP cijena po grupama")
    tab_transport = tabview.add("Transport")
    tab_net_margin = tabview.add("Postavke neto marze")
    tab_excel_out = tabview.add("Excel izlaz")
    tab_excel_in = tabview.add("Excel ulaz")
    tab_kalk = tabview.add("Kalkulacije")

    def get_categories():
        cats = set(prefix_map.values())
        cats.update(SKU_CATEGORY_OVERRIDES.values())
        cats.add("Custom")
        cats.add("Ostalo")
        return sorted(cats)

    def refresh_category_options():
        cats = get_categories()
        if not cats:
            cats = ["Custom"]
        add_cat_combo.configure(values=cats)
        view_cat_combo.configure(values=cats)
        if add_cat_combo.get() not in cats:
            add_cat_combo.set(cats[0])
        if view_cat_combo.get() not in cats:
            view_cat_combo.set(cats[0])
        refresh_sku_list()
        refresh_mp_category_options()

    def refresh_sku_list():
        cat = view_cat_combo.get().strip()
        if not cat:
            return
        if cat == "Custom":
            skus = sorted(CUSTOM_SKU_SET)
        else:
            all_skus = set(mp_cijene.keys())
            all_skus.update(CUSTOM_SKU_SET)
            all_skus.update(SKU_CATEGORY_OVERRIDES.keys())
            skus = sorted(
                [
                    sku
                    for sku in all_skus
                    if kategorija_za_sifru(sku, allow_custom=False) == cat
                ]
            )
        sku_listbox.delete(0, tk.END)
        for sku in skus:
            sku_listbox.insert(tk.END, sku)
        sku_count.configure(text=f"SKU count: {len(skus)}")

    def add_sku_to_category():
        sku = sku_entry.get().strip().upper()
        cat = add_cat_combo.get().strip()
        if not sku:
            messagebox.showerror("Greska", "Unesi SKU.")
            return
        if not cat:
            messagebox.showerror("Greska", "Odaberi kategoriju.")
            return

        if cat == "Custom":
            CUSTOM_SKU_SET.add(sku)
        else:
            SKU_CATEGORY_OVERRIDES[sku] = cat
        CUSTOM_SKU_LIST[:] = sorted(CUSTOM_SKU_SET)

        if sku not in mp_cijene:
            if cat in MP_PRICE_BY_CATEGORY:
                mp_cijene[sku] = float(MP_PRICE_BY_CATEGORY.get(cat, 0.0))
            else:
                mp_cijene[sku] = 0.0
            save_mp_cjenovnik()
        save_settings()
        refresh_category_options()
        sku_entry.delete(0, tk.END)
        messagebox.showinfo("Uspjeh", f"SKU {sku} dodat u kategoriju {cat}.")

    def remove_selected_sku():
        selection = sku_listbox.curselection()
        if not selection:
            messagebox.showerror("Greska", "Odaberi SKU za brisanje.")
            return
        sku = sku_listbox.get(selection[0])
        cat = view_cat_combo.get().strip()
        if cat == "Custom":
            if sku in CUSTOM_SKU_SET:
                CUSTOM_SKU_SET.discard(sku)
        else:
            if sku in SKU_CATEGORY_OVERRIDES:
                del SKU_CATEGORY_OVERRIDES[sku]
            else:
                if sku in CUSTOM_SKU_SET:
                    messagebox.showerror(
                        "Greska",
                        "SKU je u Custom listi i po prefiksu pripada kategoriji. "
                        "Uklanjanje bi ga izbacilo iz fiksne grupe. "
                        "Ako zelis promijeniti kategoriju, dodaj override.",
                    )
                    return
                SKU_CATEGORY_OVERRIDES[sku] = "Ostalo"
        CUSTOM_SKU_LIST[:] = sorted(CUSTOM_SKU_SET)
        save_settings()
        refresh_category_options()

    def add_category_prefix():
        name = cat_entry.get().strip()
        pref = prefix_entry.get().strip().upper()
        if not name:
            messagebox.showerror("Greska", "Unesi naziv kategorije.")
            return
        if not pref:
            messagebox.showerror("Greska", "Unesi prefiks.")
            return
        prefix_map[pref] = name
        save_settings()
        refresh_category_options()
        cat_entry.delete(0, tk.END)
        prefix_entry.delete(0, tk.END)
        messagebox.showinfo("Uspjeh", f"Kategorija {name} dodana za prefiks {pref}.")

    def refresh_prefix_price_list():
        prefix_price_listbox.delete(0, tk.END)
        for pref, price in sorted(PRICE_BY_PREFIX.items()):
            prefix_price_listbox.insert(tk.END, f"{pref} = {price:.2f}")

    def refresh_sku_price_list():
        sku_price_listbox.delete(0, tk.END)
        for sku, price in sorted(PRICE_BY_SKU.items()):
            sku_price_listbox.insert(tk.END, f"{sku} = {price:.2f} (rucno)")

    def on_prefix_price_select(_event=None):
        selection = prefix_price_listbox.curselection()
        if not selection:
            return
        text = prefix_price_listbox.get(selection[0])
        if " = " not in text:
            return
        pref, price = text.split(" = ", 1)
        prefix_price_entry.delete(0, tk.END)
        prefix_price_entry.insert(0, pref)
        prefix_price_value_entry.delete(0, tk.END)
        prefix_price_value_entry.insert(0, price)

    def on_sku_price_select(_event=None):
        selection = sku_price_listbox.curselection()
        if not selection:
            return
        text = sku_price_listbox.get(selection[0])
        if " = " not in text:
            return
        sku, price = text.split(" = ", 1)
        sku_price_entry.delete(0, tk.END)
        sku_price_entry.insert(0, sku)
        sku_price_value_entry.delete(0, tk.END)
        sku_price_value_entry.insert(0, price)

    def add_or_update_prefix_price():
        pref = prefix_price_entry.get().strip().upper()
        if not pref:
            messagebox.showerror("Greska", "Unesi prefiks.")
            return
        try:
            price = float(prefix_price_value_entry.get().strip())
        except Exception:
            messagebox.showerror("Greska", "Unesi validnu cijenu.")
            return
        PRICE_BY_PREFIX[pref] = float(price)
        save_settings()
        refresh_prefix_price_list()
        messagebox.showinfo("Uspjeh", f"Cijena za prefiks {pref} sacuvana.")

    def remove_prefix_price():
        pref = prefix_price_entry.get().strip().upper()
        if not pref:
            messagebox.showerror("Greska", "Unesi prefiks.")
            return
        if pref not in PRICE_BY_PREFIX:
            messagebox.showerror("Greska", "Prefiks nije pronadjen.")
            return
        if not messagebox.askyesno("Potvrda", f"Ukloniti cijenu za prefiks {pref}?"):
            return
        del PRICE_BY_PREFIX[pref]
        save_settings()
        refresh_prefix_price_list()

    def add_or_update_sku_price():
        sku = sku_price_entry.get().strip().upper()
        if not sku:
            messagebox.showerror("Greska", "Unesi SKU.")
            return
        try:
            price = float(sku_price_value_entry.get().strip())
        except Exception:
            messagebox.showerror("Greska", "Unesi validnu cijenu.")
            return
        PRICE_BY_SKU[sku] = float(price)
        save_settings()
        refresh_sku_price_list()
        messagebox.showinfo("Uspjeh", f"Cijena za SKU {sku} sacuvana.")

    def remove_sku_price():
        sku = sku_price_entry.get().strip().upper()
        if not sku:
            messagebox.showerror("Greska", "Unesi SKU.")
            return
        if sku not in PRICE_BY_SKU:
            messagebox.showerror("Greska", "SKU nije pronadjen.")
            return
        if not messagebox.askyesno("Potvrda", f"Ukloniti cijenu za SKU {sku}?"):
            return
        del PRICE_BY_SKU[sku]
        save_settings()
        refresh_sku_price_list()

    mp_price_search_var = tk.StringVar(value="")

    def refresh_mp_price_list():
        mp_price_listbox.delete(0, tk.END)
        query = mp_price_search_var.get().strip().upper()
        for sku, price in sorted(mp_cijene.items()):
            if query and query not in sku:
                continue
            mp_price_listbox.insert(tk.END, f"{sku} = {float(price):.2f}")

    def on_mp_price_select(_event=None):
        selection = mp_price_listbox.curselection()
        if not selection:
            return
        text = mp_price_listbox.get(selection[0])
        if " = " not in text:
            return
        sku, price = text.split(" = ", 1)
        mp_price_entry.delete(0, tk.END)
        mp_price_entry.insert(0, sku)
        mp_price_value_entry.delete(0, tk.END)
        mp_price_value_entry.insert(0, price)

    def add_or_update_mp_price():
        sku = mp_price_entry.get().strip().upper()
        if not sku:
            messagebox.showerror("Greska", "Unesi SKU.")
            return
        try:
            price = float(mp_price_value_entry.get().strip())
        except Exception:
            messagebox.showerror("Greska", "Unesi validnu cijenu.")
            return
        mp_cijene[sku] = float(price)
        save_mp_cjenovnik()
        refresh_mp_price_list()
        messagebox.showinfo("Uspjeh", f"MP cijena za SKU {sku} sacuvana.")

    def remove_mp_price():
        sku = mp_price_entry.get().strip().upper()
        if not sku:
            messagebox.showerror("Greska", "Unesi SKU.")
            return
        if sku not in mp_cijene:
            messagebox.showerror("Greska", "SKU nije pronadjen.")
            return
        if not messagebox.askyesno("Potvrda", f"Ukloniti MP cijenu za SKU {sku}?"):
            return
        del mp_cijene[sku]
        save_mp_cjenovnik()
        refresh_mp_price_list()

    def refresh_mp_category_options():
        cats = get_categories()
        if not cats:
            cats = ["Custom"]
        mp_cat_combo.configure(values=cats)
        if mp_cat_combo.get() not in cats:
            mp_cat_combo.set(cats[0])
        refresh_mp_category_list()

    def refresh_mp_category_list():
        mp_cat_listbox.delete(0, tk.END)
        for cat, price in sorted(MP_PRICE_BY_CATEGORY.items()):
            mp_cat_listbox.insert(tk.END, f"{cat} = {float(price):.2f}")

    def on_mp_category_select(_event=None):
        selection = mp_cat_listbox.curselection()
        if not selection:
            return
        text = mp_cat_listbox.get(selection[0])
        if " = " not in text:
            return
        cat, price = text.split(" = ", 1)
        mp_cat_combo.set(cat)
        mp_cat_price_entry.delete(0, tk.END)
        mp_cat_price_entry.insert(0, price)

    def add_or_update_mp_category_price():
        cat = mp_cat_combo.get().strip()
        if not cat:
            messagebox.showerror("Greska", "Odaberi kategoriju.")
            return
        try:
            price = float(mp_cat_price_entry.get().strip())
        except Exception:
            messagebox.showerror("Greska", "Unesi validnu cijenu.")
            return
        MP_PRICE_BY_CATEGORY[cat] = float(price)
        save_settings()
        refresh_mp_category_list()
        messagebox.showinfo("Uspjeh", f"MP cijena za kategoriju {cat} sacuvana.")

    def remove_mp_category_price():
        cat = mp_cat_combo.get().strip()
        if not cat:
            messagebox.showerror("Greska", "Odaberi kategoriju.")
            return
        if cat not in MP_PRICE_BY_CATEGORY:
            messagebox.showerror("Greska", "Kategorija nije pronadjena.")
            return
        if not messagebox.askyesno("Potvrda", f"Ukloniti MP cijenu za {cat}?"):
            return
        del MP_PRICE_BY_CATEGORY[cat]
        save_settings()
        refresh_mp_category_list()

    # Layout - tab Artikli i kategorije
    top = ctk.CTkFrame(tab_items)
    top.pack(fill="x", padx=10, pady=10)

    ctk.CTkLabel(
        top, text="Dodaj SKU u kategoriju", font=ctk.CTkFont(size=14, weight="bold")
    ).grid(row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(top, text="SKU:").grid(row=1, column=0, padx=10, pady=6, sticky="e")
    sku_entry = ctk.CTkEntry(top, width=160)
    sku_entry.grid(row=1, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(top, text="Kategorija:").grid(
        row=1, column=2, padx=10, pady=6, sticky="e"
    )
    add_cat_combo = ctk.CTkComboBox(top, values=get_categories(), width=200)
    add_cat_combo.grid(row=1, column=3, padx=10, pady=6, sticky="w")

    btn_add_sku = ctk.CTkButton(
        top, text="Dodaj", fg_color=ROSEGOLD, hover_color=ROSEGOLD_DARK, command=add_sku_to_category
    )
    btn_add_sku.grid(row=2, column=3, padx=10, pady=(4, 10), sticky="e")

    mid = ctk.CTkFrame(tab_items)
    mid.pack(fill="both", expand=True, padx=10, pady=10)

    ctk.CTkLabel(
        mid, text="Pregled SKU po kategoriji", font=ctk.CTkFont(size=14, weight="bold")
    ).grid(row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(mid, text="Kategorija:").grid(
        row=1, column=0, padx=10, pady=6, sticky="e"
    )
    view_cat_combo = ctk.CTkComboBox(
        mid, values=get_categories(), width=220, command=lambda _: refresh_sku_list()
    )
    view_cat_combo.grid(row=1, column=1, padx=10, pady=6, sticky="w")

    sku_list_frame = ctk.CTkFrame(mid)
    sku_list_frame.grid(row=2, column=0, columnspan=3, padx=10, pady=(5, 10), sticky="nsew")
    mid.grid_rowconfigure(2, weight=1)
    mid.grid_columnconfigure(1, weight=1)

    sku_listbox = tk.Listbox(sku_list_frame, height=10, exportselection=False)
    sku_listbox.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
    scrollbar = tk.Scrollbar(sku_list_frame, command=sku_listbox.yview)
    scrollbar.pack(side="right", fill="y", pady=10, padx=(0, 10))
    sku_listbox.config(yscrollcommand=scrollbar.set)

    sku_count = ctk.CTkLabel(mid, text="SKU count: 0", text_color=TEXT_MUTED)
    sku_count.grid(row=3, column=0, columnspan=2, padx=10, pady=(0, 10), sticky="w")

    btn_remove = ctk.CTkButton(
        mid, text="Izbrisi odabrani", fg_color="#a33b3b", hover_color="#8a2f2f", command=remove_selected_sku
    )
    btn_remove.grid(row=3, column=2, padx=10, pady=(0, 10), sticky="e")

    ctk.CTkLabel(
        mid,
        text="Napomena: lista prikazuje sve SKU-ove; Custom je posebna oznaka.",
        text_color=TEXT_MUTED,
    ).grid(row=4, column=0, columnspan=3, padx=10, pady=(0, 10), sticky="w")

    bottom = ctk.CTkFrame(tab_items)
    bottom.pack(fill="x", padx=10, pady=10)

    ctk.CTkLabel(
        bottom, text="Nova kategorija / prefiks", font=ctk.CTkFont(size=14, weight="bold")
    ).grid(row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(bottom, text="Naziv kategorije:").grid(
        row=1, column=0, padx=10, pady=6, sticky="e"
    )
    cat_entry = ctk.CTkEntry(bottom, width=220)
    cat_entry.grid(row=1, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(bottom, text="Prefiks:").grid(
        row=1, column=2, padx=10, pady=6, sticky="e"
    )
    prefix_entry = ctk.CTkEntry(bottom, width=120)
    prefix_entry.grid(row=1, column=3, padx=10, pady=6, sticky="w")

    btn_add_cat = ctk.CTkButton(
        bottom, text="Dodaj kategoriju", fg_color=ROSEGOLD, hover_color=ROSEGOLD_DARK, command=add_category_prefix
    )
    btn_add_cat.grid(row=2, column=3, padx=10, pady=(4, 10), sticky="e")

    # Placeholder tabs
    prices_frame = ctk.CTkFrame(tab_prices)
    prices_frame.pack(fill="both", expand=True, padx=10, pady=10)

    ctk.CTkLabel(
        prices_frame,
        text="Nabavne cijene po prefiksu",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).grid(row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(prices_frame, text="Prefiks:").grid(
        row=1, column=0, padx=10, pady=6, sticky="e"
    )
    prefix_price_entry = ctk.CTkEntry(prices_frame, width=120)
    prefix_price_entry.grid(row=1, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(prices_frame, text="Cijena (USD):").grid(
        row=1, column=2, padx=10, pady=6, sticky="e"
    )
    prefix_price_value_entry = ctk.CTkEntry(prices_frame, width=120)
    prefix_price_value_entry.grid(row=1, column=3, padx=10, pady=6, sticky="w")

    btn_prefix_save = ctk.CTkButton(
        prices_frame,
        text="Dodaj / azuriraj",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=add_or_update_prefix_price,
    )
    btn_prefix_save.grid(row=2, column=3, padx=10, pady=(4, 10), sticky="e")

    btn_prefix_remove = ctk.CTkButton(
        prices_frame,
        text="Ukloni",
        fg_color="#a33b3b",
        hover_color="#8a2f2f",
        command=remove_prefix_price,
    )
    btn_prefix_remove.grid(row=2, column=2, padx=10, pady=(4, 10), sticky="e")

    prefix_list_frame = ctk.CTkFrame(prices_frame)
    prefix_list_frame.grid(
        row=3, column=0, columnspan=4, padx=10, pady=(5, 10), sticky="nsew"
    )
    prices_frame.grid_rowconfigure(3, weight=1)
    prices_frame.grid_columnconfigure(1, weight=1)

    prefix_price_listbox = tk.Listbox(prefix_list_frame, height=8, exportselection=False)
    prefix_price_listbox.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
    prefix_scroll = tk.Scrollbar(prefix_list_frame, command=prefix_price_listbox.yview)
    prefix_scroll.pack(side="right", fill="y", pady=10, padx=(0, 10))
    prefix_price_listbox.config(yscrollcommand=prefix_scroll.set)
    prefix_price_listbox.bind("<<ListboxSelect>>", on_prefix_price_select)

    ctk.CTkLabel(
        prices_frame,
        text="Nabavne cijene po SKU (override ima prioritet nad prefiksom)",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).grid(row=4, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(prices_frame, text="SKU:").grid(
        row=5, column=0, padx=10, pady=6, sticky="e"
    )
    sku_price_entry = ctk.CTkEntry(prices_frame, width=120)
    sku_price_entry.grid(row=5, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(prices_frame, text="Cijena (USD):").grid(
        row=5, column=2, padx=10, pady=6, sticky="e"
    )
    sku_price_value_entry = ctk.CTkEntry(prices_frame, width=120)
    sku_price_value_entry.grid(row=5, column=3, padx=10, pady=6, sticky="w")

    btn_sku_save = ctk.CTkButton(
        prices_frame,
        text="Dodaj / azuriraj",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=add_or_update_sku_price,
    )
    btn_sku_save.grid(row=6, column=3, padx=10, pady=(4, 10), sticky="e")

    btn_sku_remove = ctk.CTkButton(
        prices_frame,
        text="Ukloni",
        fg_color="#a33b3b",
        hover_color="#8a2f2f",
        command=remove_sku_price,
    )
    btn_sku_remove.grid(row=6, column=2, padx=10, pady=(4, 10), sticky="e")

    sku_list_frame = ctk.CTkFrame(prices_frame)
    sku_list_frame.grid(
        row=7, column=0, columnspan=4, padx=10, pady=(5, 10), sticky="nsew"
    )
    prices_frame.grid_rowconfigure(7, weight=1)
    prices_frame.grid_columnconfigure(1, weight=1)

    sku_price_listbox = tk.Listbox(sku_list_frame, height=8, exportselection=False)
    sku_price_listbox.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
    sku_scroll = tk.Scrollbar(sku_list_frame, command=sku_price_listbox.yview)
    sku_scroll.pack(side="right", fill="y", pady=10, padx=(0, 10))
    sku_price_listbox.config(yscrollcommand=sku_scroll.set)
    sku_price_listbox.bind("<<ListboxSelect>>", on_sku_price_select)

    refresh_prefix_price_list()
    refresh_sku_price_list()

    mp_frame = ctk.CTkFrame(tab_mp_prices)
    mp_frame.pack(fill="both", expand=True, padx=10, pady=10)

    ctk.CTkLabel(
        mp_frame,
        text="MP cijene (KM) po SKU",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).grid(row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(mp_frame, text="SKU:").grid(
        row=1, column=0, padx=10, pady=6, sticky="e"
    )
    mp_price_entry = ctk.CTkEntry(mp_frame, width=160)
    mp_price_entry.grid(row=1, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(mp_frame, text="MP cijena (KM):").grid(
        row=1, column=2, padx=10, pady=6, sticky="e"
    )
    mp_price_value_entry = ctk.CTkEntry(mp_frame, width=120)
    mp_price_value_entry.grid(row=1, column=3, padx=10, pady=6, sticky="w")

    btn_mp_save = ctk.CTkButton(
        mp_frame,
        text="Dodaj / azuriraj",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=add_or_update_mp_price,
    )
    btn_mp_save.grid(row=2, column=3, padx=10, pady=(4, 10), sticky="e")

    btn_mp_remove = ctk.CTkButton(
        mp_frame,
        text="Ukloni",
        fg_color="#a33b3b",
        hover_color="#8a2f2f",
        command=remove_mp_price,
    )
    btn_mp_remove.grid(row=2, column=2, padx=10, pady=(4, 10), sticky="e")

    ctk.CTkLabel(mp_frame, text="Pretraga:").grid(
        row=3, column=0, padx=10, pady=(0, 10), sticky="e"
    )
    mp_price_search_entry = ctk.CTkEntry(
        mp_frame, textvariable=mp_price_search_var, width=200
    )
    mp_price_search_entry.grid(
        row=3, column=1, columnspan=3, padx=10, pady=(0, 10), sticky="we"
    )
    mp_price_search_var.trace_add("write", lambda *_: refresh_mp_price_list())

    mp_list_frame = ctk.CTkFrame(mp_frame)
    mp_list_frame.grid(
        row=4, column=0, columnspan=4, padx=10, pady=(5, 10), sticky="nsew"
    )
    mp_frame.grid_rowconfigure(4, weight=1)
    mp_frame.grid_columnconfigure(1, weight=1)

    mp_price_listbox = tk.Listbox(mp_list_frame, height=12, exportselection=False)
    mp_price_listbox.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
    mp_scroll = tk.Scrollbar(mp_list_frame, command=mp_price_listbox.yview)
    mp_scroll.pack(side="right", fill="y", pady=10, padx=(0, 10))
    mp_price_listbox.config(yscrollcommand=mp_scroll.set)
    mp_price_listbox.bind("<<ListboxSelect>>", on_mp_price_select)

    refresh_mp_price_list()

    mp_cat_frame = ctk.CTkFrame(tab_mp_prices_cat)
    mp_cat_frame.pack(fill="both", expand=True, padx=10, pady=10)

    ctk.CTkLabel(
        mp_cat_frame,
        text="MP cijene (KM) po kategoriji (override nad SKU)",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).grid(row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(mp_cat_frame, text="Kategorija:").grid(
        row=1, column=0, padx=10, pady=6, sticky="e"
    )
    mp_cat_combo = ctk.CTkComboBox(mp_cat_frame, values=get_categories(), width=220)
    mp_cat_combo.grid(row=1, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(mp_cat_frame, text="MP cijena (KM):").grid(
        row=1, column=2, padx=10, pady=6, sticky="e"
    )
    mp_cat_price_entry = ctk.CTkEntry(mp_cat_frame, width=120)
    mp_cat_price_entry.grid(row=1, column=3, padx=10, pady=6, sticky="w")

    btn_mp_cat_save = ctk.CTkButton(
        mp_cat_frame,
        text="Dodaj / azuriraj",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=add_or_update_mp_category_price,
    )
    btn_mp_cat_save.grid(row=2, column=3, padx=10, pady=(4, 10), sticky="e")

    btn_mp_cat_remove = ctk.CTkButton(
        mp_cat_frame,
        text="Ukloni",
        fg_color="#a33b3b",
        hover_color="#8a2f2f",
        command=remove_mp_category_price,
    )
    btn_mp_cat_remove.grid(row=2, column=2, padx=10, pady=(4, 10), sticky="e")

    mp_cat_list_frame = ctk.CTkFrame(mp_cat_frame)
    mp_cat_list_frame.grid(
        row=3, column=0, columnspan=4, padx=10, pady=(5, 10), sticky="nsew"
    )
    mp_cat_frame.grid_rowconfigure(3, weight=1)
    mp_cat_frame.grid_columnconfigure(1, weight=1)

    mp_cat_listbox = tk.Listbox(mp_cat_list_frame, height=12, exportselection=False)
    mp_cat_listbox.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
    mp_cat_scroll = tk.Scrollbar(mp_cat_list_frame, command=mp_cat_listbox.yview)
    mp_cat_scroll.pack(side="right", fill="y", pady=10, padx=(0, 10))
    mp_cat_listbox.config(yscrollcommand=mp_cat_scroll.set)
    mp_cat_listbox.bind("<<ListboxSelect>>", on_mp_category_select)

    refresh_mp_category_options()

    ctk.CTkLabel(
        tab_transport,
        text="Transport podesavanja ce biti dodana kasnije.",
        text_color=TEXT_MUTED,
    ).pack(pady=20)

    # Postavke neto marze
    net_frame = ctk.CTkFrame(tab_net_margin)
    net_frame.pack(fill="both", expand=True, padx=10, pady=10)

    ctk.CTkLabel(
        net_frame,
        text="Postavke neto marze (po artiklu)",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).grid(row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(net_frame, text="Transport (USD/kom):").grid(
        row=1, column=0, padx=10, pady=6, sticky="e"
    )
    net_transport_entry = ctk.CTkEntry(net_frame, width=120)
    net_transport_entry.insert(0, f"{net_margin_transport_usd:.2f}")
    net_transport_entry.grid(row=1, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(net_frame, text="Carina (%):").grid(
        row=1, column=2, padx=10, pady=6, sticky="e"
    )
    net_customs_entry = ctk.CTkEntry(net_frame, width=120)
    net_customs_entry.insert(0, f"{net_margin_customs_pct:.2f}")
    net_customs_entry.grid(row=1, column=3, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(net_frame, text="Marketing (KM/kom):").grid(
        row=2, column=0, padx=10, pady=6, sticky="e"
    )
    net_marketing_entry = ctk.CTkEntry(net_frame, width=120)
    net_marketing_entry.insert(0, f"{net_margin_marketing_bam:.2f}")
    net_marketing_entry.grid(row=2, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(net_frame, text="Poslovni prostor (KM/mj):").grid(
        row=3, column=0, padx=10, pady=6, sticky="e"
    )
    net_space_entry = ctk.CTkEntry(net_frame, width=120)
    net_space_entry.insert(0, f"{net_margin_space_bam:.2f}")
    net_space_entry.grid(row=3, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(net_frame, text="Radna snaga (KM/mj):").grid(
        row=3, column=2, padx=10, pady=6, sticky="e"
    )
    net_labor_entry = ctk.CTkEntry(net_frame, width=120)
    net_labor_entry.insert(0, f"{net_margin_labor_bam:.2f}")
    net_labor_entry.grid(row=3, column=3, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(net_frame, text="Knjigovodstvo (KM/mj):").grid(
        row=4, column=0, padx=10, pady=6, sticky="e"
    )
    net_accounting_entry = ctk.CTkEntry(net_frame, width=120)
    net_accounting_entry.insert(0, f"{net_margin_accounting_bam:.2f}")
    net_accounting_entry.grid(row=4, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(net_frame, text="Komunalije (KM/mj):").grid(
        row=4, column=2, padx=10, pady=6, sticky="e"
    )
    net_utilities_entry = ctk.CTkEntry(net_frame, width=120)
    net_utilities_entry.insert(0, f"{net_margin_utilities_bam:.2f}")
    net_utilities_entry.grid(row=4, column=3, padx=10, pady=6, sticky="w")

    def save_net_margin_settings():
        global net_margin_transport_usd, net_margin_customs_pct, net_margin_marketing_bam
        global net_margin_space_bam, net_margin_labor_bam
        global net_margin_accounting_bam, net_margin_utilities_bam
        try:
            net_margin_transport_usd = float(net_transport_entry.get().strip() or 0.0)
            net_margin_customs_pct = float(net_customs_entry.get().strip() or 0.0)
            net_margin_marketing_bam = float(net_marketing_entry.get().strip() or 0.0)
            net_margin_space_bam = float(net_space_entry.get().strip() or 0.0)
            net_margin_labor_bam = float(net_labor_entry.get().strip() or 0.0)
            net_margin_accounting_bam = float(net_accounting_entry.get().strip() or 0.0)
            net_margin_utilities_bam = float(net_utilities_entry.get().strip() or 0.0)
        except Exception:
            messagebox.showerror("Greska", "Unesi validne vrijednosti.")
            return
        save_settings()
        messagebox.showinfo("Uspjeh", "Postavke neto marze sacuvane.")

    btn_save_net = ctk.CTkButton(
        net_frame,
        text="Sacuvaj",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=save_net_margin_settings,
    )
    btn_save_net.grid(row=5, column=3, padx=10, pady=(6, 10), sticky="e")

    # Kalkulacije - nabavne cijene
    kalk_frame = ctk.CTkFrame(tab_kalk)
    kalk_frame.pack(fill="both", expand=True, padx=10, pady=10)

    ctk.CTkLabel(
        kalk_frame,
        text="Kalkulacije - nabavne cijene",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).grid(row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(10, 5))

    kalk_path_label = ctk.CTkLabel(
        kalk_frame,
        text=f"Folder kalkulacija: {kalkulacije_folder or '(nije odabran)'}",
        text_color=TEXT_MUTED,
        wraplength=700,
        justify="left",
    )
    kalk_path_label.grid(row=1, column=0, columnspan=3, sticky="w", padx=10, pady=6)

    def _refresh_kalk_label():
        kalk_path_label.configure(
            text=f"Folder kalkulacija: {kalkulacije_folder or '(nije odabran)'}"
        )

    def choose_kalk_folder():
        global kalkulacije_folder
        p = filedialog.askdirectory(
            title="Odaberi folder kalkulacija",
            initialdir=kalkulacije_folder or DEFAULT_KALK_FOLDER,
        )
        if p:
            kalkulacije_folder = p
            save_settings()
            _refresh_kalk_label()

    last_offsets = KALK_LAST_OFFSETS

    def show_kalk_review(path, rows, initial_error=None):
        result = {"ok": False, "rows": rows, "override_skus": []}
        current_rows = list(rows)
        sheet_names = []
        df_cache = {}
        try:
            xls = pd.ExcelFile(path)
            sheet_names = list(xls.sheet_names)
        except Exception:
            sheet_names = []

        if not sheet_names:
            sheet_names = []

        def load_sheet_df(name):
            if name in df_cache:
                return df_cache[name]
            df = pd.read_excel(path, sheet_name=name, header=None)
            df_cache[name] = df
            return df

        try:
            df_preview = load_sheet_df(sheet_names[0]) if sheet_names else None
        except Exception:
            df_preview = None

        active_sheet = tk.StringVar(value=sheet_names[0]) if sheet_names else None
        offsets_by_sheet = (
            {name: last_offsets.copy() for name in sheet_names}
            if sheet_names
            else {0: last_offsets.copy()}
        )
        updating_vars = {"active": False}
        visited_sheets = {sheet_names[0]} if sheet_names else set()

        def active_key():
            return active_sheet.get() if active_sheet else 0

        top = ctk.CTkToplevel(self)
        top.title(f"Provjera kalkulacije - {os.path.basename(path)}")
        top.geometry("980x640")
        top.grab_set()
        top.focus_set()

        def ensure_window_fits():
            # Grow to fit new content without moving the window.
            top.update_idletasks()
            try:
                x, y = top.winfo_x(), top.winfo_y()
            except Exception:
                x, y = 0, 0
            cur_w, cur_h = top.winfo_width(), top.winfo_height()
            req_w, req_h = top.winfo_reqwidth(), top.winfo_reqheight()
            new_w = max(cur_w, req_w)
            new_h = max(cur_h, req_h)
            if new_w != cur_w or new_h != cur_h:
                top.geometry(f"{new_w}x{new_h}+{x}+{y}")
                top.minsize(new_w, new_h)

        header = ctk.CTkLabel(
            top,
            text="Provjera ucitanih vrijednosti",
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        header.pack(anchor="w", padx=12, pady=(12, 4))

        info = ctk.CTkLabel(
            top,
            text=f"Fajl: {os.path.basename(path)} | Stavki: {len(rows)}",
            text_color=TEXT_MUTED,
        )
        info.pack(anchor="w", padx=12, pady=(0, 6))
        total_info = ctk.CTkLabel(top, text="", text_color=TEXT_MUTED)
        total_info.pack(anchor="w", padx=12, pady=(0, 6))

        if sheet_names and len(sheet_names) > 1:
            sheet_frame = ctk.CTkFrame(top)
            sheet_frame.pack(fill="x", padx=10, pady=(0, 6))
            ctk.CTkLabel(sheet_frame, text="Sheet:").pack(
                side="left", padx=(10, 6), pady=6
            )
            sheet_combo = ctk.CTkComboBox(
                sheet_frame, values=sheet_names, width=200, variable=active_sheet
            )
            sheet_combo.pack(side="left", padx=6, pady=6)
            ctk.CTkLabel(
                sheet_frame,
                text=f"Ukupno sheetova: {len(sheet_names)}",
                text_color=TEXT_MUTED,
            ).pack(side="left", padx=10, pady=6)
        prev = KALK_PROCESSED_FILES.get(path)
        if isinstance(prev, dict):
            try:
                mtime = os.path.getmtime(path)
                size = os.path.getsize(path)
            except Exception:
                mtime = None
                size = None
            if prev.get("mtime") == mtime and prev.get("size") == size:
                note_text = "Napomena: Ovaj fajl je vec obradjen (nije mijenjan)."
            else:
                note_text = "Napomena: Ovaj fajl je vec obradjen (izmijenjen)."
            ctk.CTkLabel(top, text=note_text, text_color="#ffb84d").pack(
                anchor="w", padx=12, pady=(0, 6)
            )

        offsets_frame = ctk.CTkFrame(top)
        offsets_frame.pack(fill="x", padx=10, pady=(0, 6))

        ctk.CTkLabel(
            offsets_frame,
            text="Pomjeri kolone:",
            text_color=TEXT_MUTED,
        ).grid(row=0, column=0, padx=10, pady=6, sticky="w")

        offset_options = [
            ("Lijevo+2", -2),
            ("Lijevo", -1),
            ("Header", 0),
            ("Desno", 1),
            ("Desno+2", 2),
        ]
        offset_labels = [label for label, _ in offset_options]
        offset_map = {label: val for label, val in offset_options}

        def make_offset_combo(label_text, column, default="Header"):
            ctk.CTkLabel(offsets_frame, text=label_text).grid(
                row=0, column=column, padx=6, pady=6, sticky="w"
            )
            var = tk.StringVar(value=default)
            combo = ctk.CTkComboBox(
                offsets_frame, values=offset_labels, width=120, variable=var
            )
            combo.grid(row=0, column=column + 1, padx=6, pady=6, sticky="w")
            return var

        def offset_to_label(offset):
            for label, val in offset_options:
                if val == offset:
                    return label
            return "Header"

        initial_offsets = offsets_by_sheet.get(active_key(), last_offsets)
        var_sku = make_offset_combo(
            "SKU", 1, offset_to_label(initial_offsets.get("sku", 0))
        )
        var_nabavna = make_offset_combo(
            "Nabavna", 3, offset_to_label(initial_offsets.get("nabavna", 0))
        )
        var_mpc_wo = make_offset_combo(
            "MP bez PDV", 5, offset_to_label(initial_offsets.get("mpc_wo", 0))
        )
        var_mpc_w = make_offset_combo(
            "MP s PDV", 7, offset_to_label(initial_offsets.get("mpc_w", 0))
        )
        var_pdv = make_offset_combo(
            "PDV %", 9, offset_to_label(initial_offsets.get("pdv_rate", 0))
        )
        var_qty = make_offset_combo(
            "Kolicina", 11, offset_to_label(initial_offsets.get("kolicina", 0))
        )

        manual_frame = ctk.CTkFrame(top)
        manual_frame.pack(fill="x", padx=10, pady=(0, 6))

        manual_label = ctk.CTkLabel(manual_frame, text="", text_color=TEXT_MUTED)
        manual_label.pack(side="left", padx=10, pady=6)

        override_manual_var = tk.BooleanVar(value=False)
        override_manual_check = ctk.CTkCheckBox(
            manual_frame,
            text="Zamijeni rucno dodane cijene iz ove kalkulacije",
            variable=override_manual_var,
        )
        override_manual_check.pack(side="left", padx=10, pady=6)

        move_frame = ctk.CTkFrame(top)
        move_frame.pack(fill="x", padx=10, pady=(0, 6))

        ctk.CTkLabel(move_frame, text="Pomjeri kolonu:").pack(
            side="left", padx=(10, 6), pady=6
        )
        field_options = [
            ("SKU", "sku"),
            ("Nabavna", "nabavna"),
            ("MP bez PDV", "mpc_wo"),
            ("MP s PDV", "mpc_w"),
            ("PDV %", "pdv_rate"),
            ("Kolicina", "kolicina"),
        ]
        field_labels = [label for label, _ in field_options]
        field_map = {label: key for label, key in field_options}

        active_field = tk.StringVar(value=field_labels[0])
        field_combo = ctk.CTkComboBox(
            move_frame, values=field_labels, width=140, variable=active_field
        )
        field_combo.pack(side="left", padx=6, pady=6)

        ctk.CTkLabel(move_frame, text="Korak:").pack(
            side="left", padx=(10, 6), pady=6
        )
        step_var = tk.StringVar(value="1")
        step_combo = ctk.CTkComboBox(
            move_frame, values=["1", "2"], width=70, variable=step_var
        )
        step_combo.pack(side="left", padx=6, pady=6)

        def reset_offsets_current_sheet():
            offsets_by_sheet[active_key()] = {
                "sku": 0,
                "nabavna": 0,
                "mpc_wo": 0,
                "mpc_w": 0,
                "pdv_rate": 0,
                "kolicina": 0,
            }
            set_vars_from_offsets(offsets_by_sheet[active_key()])
            update_last_offsets_from_vars()
            refresh_preview()
            update_confirm_label()
            ensure_window_fits()

        var_map = {
            "sku": var_sku,
            "nabavna": var_nabavna,
            "mpc_wo": var_mpc_wo,
            "mpc_w": var_mpc_w,
            "pdv_rate": var_pdv,
            "kolicina": var_qty,
        }

        def set_vars_from_offsets(offsets):
            updating_vars["active"] = True
            var_sku.set(offset_to_label(offsets.get("sku", 0)))
            var_nabavna.set(offset_to_label(offsets.get("nabavna", 0)))
            var_mpc_wo.set(offset_to_label(offsets.get("mpc_wo", 0)))
            var_mpc_w.set(offset_to_label(offsets.get("mpc_w", 0)))
            var_pdv.set(offset_to_label(offsets.get("pdv_rate", 0)))
            var_qty.set(offset_to_label(offsets.get("kolicina", 0)))
            updating_vars["active"] = False

        def update_last_offsets_from_vars():
            global KALK_LAST_OFFSETS
            offsets = offsets_by_sheet.setdefault(active_key(), last_offsets.copy())
            offsets.update(
                {
                    "sku": offset_map.get(var_sku.get(), 0),
                    "nabavna": offset_map.get(var_nabavna.get(), 0),
                    "mpc_wo": offset_map.get(var_mpc_wo.get(), 0),
                    "mpc_w": offset_map.get(var_mpc_w.get(), 0),
                    "pdv_rate": offset_map.get(var_pdv.get(), 0),
                    "kolicina": offset_map.get(var_qty.get(), 0),
                }
            )
            last_offsets.update(offsets)
            KALK_LAST_OFFSETS = last_offsets

        def adjust_offset(direction):
            key = field_map.get(active_field.get(), "nabavna")
            var = var_map.get(key, var_nabavna)
            current = offset_map.get(var.get(), 0)
            try:
                step = int(step_var.get())
            except ValueError:
                step = 1
            new_offset = current + (direction * step)
            new_offset = max(-2, min(2, new_offset))
            var.set(offset_to_label(new_offset))

        btn_left = ctk.CTkButton(
            move_frame, text="Pomjeri lijevo", command=lambda: adjust_offset(-1)
        )
        btn_left.pack(side="left", padx=6, pady=6)

        btn_right = ctk.CTkButton(
            move_frame, text="Pomjeri desno", command=lambda: adjust_offset(1)
        )
        btn_right.pack(side="left", padx=6, pady=6)

        btn_reset = ctk.CTkButton(
            move_frame, text="Resetuj offsete", fg_color="#555555", hover_color="#444444", command=reset_offsets_current_sheet
        )
        btn_reset.pack(side="left", padx=6, pady=6)

        def auto_detect_offsets():
            if df_preview is None:
                return
            offsets = [-2, -1, 0, 1, 2]

            def score_rows(rows_list):
                sample = rows_list[: min(len(rows_list), 30)]
                if not sample:
                    return 0.0
                total = 0.0
                for r in sample:
                    nab = parse_float_safe(r.get("nab_unit"))
                    mpc_wo = parse_float_safe(r.get("mpc_wo_vat_unit"))
                    mpc_w = parse_float_safe(r.get("mpc_w_vat_unit"))
                    vat = parse_float_safe(r.get("vat_rate"))
                    qty = parse_float_safe(r.get("qty"))
                    score = 0.0
                    if qty > 0:
                        score += 0.5
                        if abs(qty - round(qty)) < 0.01:
                            score += 0.5
                    if 10 <= vat <= 25:
                        score += 0.5
                        if 16 <= vat <= 18:
                            score += 1.0
                    if 0 < mpc_wo < 1000:
                        score += 0.5
                    if 0 < mpc_w < 1000:
                        score += 0.5
                    if mpc_w > 0 and mpc_wo > 0 and vat > 0:
                        expected = mpc_wo * (1.0 + vat / 100.0)
                        if expected > 0:
                            diff = abs(expected - mpc_w) / expected
                            if diff < 0.03:
                                score += 2.0
                            elif diff < 0.06:
                                score += 1.0
                    if nab > 0 and mpc_wo > 0:
                        if nab <= mpc_wo * 1.2:
                            score += 0.5
                        if nab <= mpc_wo:
                            score += 0.5
                    total += score
                return total / len(sample)

            best_score = -1.0
            best_offsets = None
            for o_nab in offsets:
                for o_wo in offsets:
                    for o_w in offsets:
                        for o_pdv in offsets:
                            for o_qty in offsets:
                                ok, _msg, rows_try = parse_calc_df_to_rows(
                                    df_preview,
                                    path,
                                    col_offsets={
                                        "nabavna": o_nab,
                                        "mpc_wo": o_wo,
                                        "mpc_w": o_w,
                                        "pdv_rate": o_pdv,
                                        "kolicina": o_qty,
                                    },
                                )
                                if not ok:
                                    continue
                                score = score_rows(rows_try)
                                if score > best_score:
                                    best_score = score
                                    best_offsets = (o_nab, o_wo, o_w, o_pdv, o_qty)
            if best_offsets:
                var_nabavna.set(offset_to_label(best_offsets[0]))
                var_mpc_wo.set(offset_to_label(best_offsets[1]))
                var_mpc_w.set(offset_to_label(best_offsets[2]))
                var_pdv.set(offset_to_label(best_offsets[3]))
                var_qty.set(offset_to_label(best_offsets[4]))

        btn_auto = ctk.CTkButton(
            move_frame,
            text="Auto detektuj",
            fg_color="#3b6ea5",
            hover_color="#345f90",
            command=auto_detect_offsets,
        )
        btn_auto.pack(side="left", padx=6, pady=6)

        table_frame = ctk.CTkFrame(top)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = ("SKU", "Nabavna", "MP bez PDV", "MP s PDV", "Kolicina")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=140, anchor="w")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def fmt_num(val):
            try:
                num = float(val)
            except (TypeError, ValueError):
                return ""
            return f"{num:.4f}".rstrip("0").rstrip(".")

        count_label = ctk.CTkLabel(top, text="")
        count_label.pack(anchor="w", padx=12, pady=(0, 6))

        error_label = ctk.CTkLabel(top, text="", text_color="#ff6b6b")
        error_label.pack(anchor="w", padx=12, pady=(0, 6))
        if initial_error:
            error_label.configure(text=initial_error)
        qty_warn_label = ctk.CTkLabel(top, text="", text_color="#ffb84d")
        qty_warn_label.pack(anchor="w", padx=12, pady=(0, 6))

        btn_confirm_ref = {"btn": None}

        def update_confirm_label():
            btn = btn_confirm_ref["btn"]
            if btn is None:
                return
            if not sheet_names:
                btn.configure(text="Potvrdi")
                return
            unvisited = [s for s in sheet_names if s not in visited_sheets]
            btn.configure(text="Potvrdi" if not unvisited else "Dalje (sheet)")

        def on_sheet_change(*_args):
            if not sheet_names:
                return
            key = active_key()
            visited_sheets.add(key)
            set_vars_from_offsets(offsets_by_sheet.get(key, last_offsets))
            refresh_preview()
            update_confirm_label()
            ensure_window_fits()

        if active_sheet is not None:
            active_sheet.trace_add("write", on_sheet_change)

        def update_manual_overlap(rows_list):
            overlap = sorted(
                {
                    str(r.get("sku") or "").strip().upper()
                    for r in rows_list
                    if str(r.get("sku") or "").strip().upper() in PRICE_BY_SKU
                }
            )
            if overlap:
                manual_label.configure(
                    text=f"Rucno dodani SKU u ovoj kalkulaciji: {len(overlap)}"
                )
                override_manual_check.configure(state="normal")
            else:
                manual_label.configure(text="Rucno dodani SKU u ovoj kalkulaciji: 0")
                override_manual_var.set(False)
                override_manual_check.configure(state="disabled")
            return overlap

        def populate(limit):
            tree.delete(*tree.get_children())
            data = current_rows if limit is None else current_rows[:limit]
            for row in data:
                tree.insert(
                    "",
                    "end",
                    values=(
                        str(row.get("sku") or "").strip(),
                        fmt_num(row.get("nab_unit")),
                        fmt_num(row.get("mpc_wo_vat_unit")),
                        fmt_num(row.get("mpc_w_vat_unit")),
                        fmt_num(row.get("qty")),
                    ),
                )
            shown = len(data)
            count_label.configure(
                text=f"Prikazano {shown} od {len(current_rows)} stavki."
            )
            update_manual_overlap(current_rows)

        def refresh_preview(*_args):
            nonlocal current_rows
            if updating_vars["active"]:
                return
            offsets = {
                "sku": offset_map.get(var_sku.get(), 0),
                "nabavna": offset_map.get(var_nabavna.get(), 0),
                "mpc_wo": offset_map.get(var_mpc_wo.get(), 0),
                "mpc_w": offset_map.get(var_mpc_w.get(), 0),
                "pdv_rate": offset_map.get(var_pdv.get(), 0),
                "kolicina": offset_map.get(var_qty.get(), 0),
            }
            update_last_offsets_from_vars()
            if sheet_names:
                sheet_key = active_key()
                try:
                    df_sheet = load_sheet_df(sheet_key)
                except Exception as e:
                    df_sheet = None
                    ok, msg, rows_new = False, f"Greska pri citanju sheeta: {e}", []
                if df_sheet is not None:
                    ok, msg, rows_new = parse_calc_df_to_rows(
                        df_sheet, path, col_offsets=offsets
                    )
            else:
                ok, msg, rows_new = parse_calc_excel_to_rows(
                    path, col_offsets=offsets, sheet_name=0
                )
            if ok:
                error_label.configure(text="")
                current_rows = rows_new
                missing_qty = sum(
                    1
                    for r in current_rows
                    if parse_float_safe(r.get("qty")) <= 0
                )
                if missing_qty:
                    qty_warn_label.configure(
                        text=f"Upozorenje: {missing_qty} stavki bez kolicine (ne ulaze u prosjek)."
                    )
                else:
                    qty_warn_label.configure(text="")
                info.configure(
                    text=f"Fajl: {os.path.basename(path)} | Stavki: {len(current_rows)}"
                )
                total = 0
                for sheet_key in (sheet_names or [0]):
                    offsets_sheet = offsets_by_sheet.get(sheet_key, last_offsets)
                    if sheet_names:
                        try:
                            df_sheet = load_sheet_df(sheet_key)
                        except Exception as e:
                            continue
                        ok_total, msg_total, rows_total = parse_calc_df_to_rows(
                            df_sheet, path, col_offsets=offsets_sheet
                        )
                    else:
                        ok_total, msg_total, rows_total = parse_calc_excel_to_rows(
                            path, col_offsets=offsets_sheet, sheet_name=0
                        )
                    if not ok_total:
                        continue
                    total += len(rows_total)
                total_info.configure(text=f"Ukupno stavki (svi sheetovi): {total}")
                populate(20 if len(rows_new) > 20 else None)
                show_all_btn.configure(
                    state="disabled" if len(rows_new) <= 20 else "normal"
                )
                ensure_window_fits()
            else:
                error_label.configure(text=msg)
                current_rows = []
                info.configure(
                    text=f"Fajl: {os.path.basename(path)} | Stavki: 0"
                )
                total_info.configure(text="Ukupno stavki (svi sheetovi): 0")
                tree.delete(*tree.get_children())
                count_label.configure(text="Prikazano 0 od 0 stavki.")
                show_all_btn.configure(state="disabled")
                qty_warn_label.configure(text="")

        for var in (var_sku, var_nabavna, var_mpc_wo, var_mpc_w, var_pdv, var_qty):
            var.trace_add("write", refresh_preview)

        btn_frame = ctk.CTkFrame(top)
        btn_frame.pack(fill="x", padx=10, pady=(0, 10))

        def on_show_all():
            populate(None)
            show_all_btn.configure(state="disabled")

        show_all_btn = ctk.CTkButton(
            btn_frame,
            text="Prikazi sve",
            command=on_show_all,
        )
        show_all_btn.pack(side="left", padx=10, pady=8)

        def on_back():
            result["back"] = True
            top.destroy()

        btn_back = ctk.CTkButton(
            btn_frame,
            text="Nazad",
            fg_color="#666666",
            hover_color="#555555",
            command=on_back,
        )
        btn_back.pack(side="left", padx=10, pady=8)

        def on_confirm():
            global KALK_LAST_OFFSETS
            if sheet_names:
                unvisited = [s for s in sheet_names if s not in visited_sheets]
                if unvisited:
                    active_sheet.set(unvisited[0])
                    return
            combined_rows = []
            per_sheet_counts = []
            for sheet_key in (sheet_names or [0]):
                offsets = offsets_by_sheet.get(sheet_key, last_offsets)
                if sheet_names:
                    try:
                        df_sheet = load_sheet_df(sheet_key)
                    except Exception as e:
                        df_sheet = None
                        ok, msg, rows_new = False, f"Greska pri citanju sheeta: {e}", []
                    if df_sheet is not None:
                        ok, msg, rows_new = parse_calc_df_to_rows(
                            df_sheet, path, col_offsets=offsets
                        )
                else:
                    ok, msg, rows_new = parse_calc_excel_to_rows(
                        path, col_offsets=offsets, sheet_name=0
                    )
                if not ok:
                    error_label.configure(text=msg)
                    if sheet_names:
                        active_sheet.set(sheet_key)
                        set_vars_from_offsets(offsets)
                    return
                combined_rows.extend(rows_new)
                if sheet_names:
                    per_sheet_counts.append((sheet_key, len(rows_new)))

            total_rows = len(combined_rows)
            if total_rows == 0:
                messagebox.showerror(
                    "Greska",
                    "Nije pronadjena nijedna stavka (SKU) u ovoj kalkulaciji.",
                )
                return
            if per_sheet_counts:
                empty_sheets = [name for name, count in per_sheet_counts if count == 0]
                if empty_sheets:
                    if messagebox.askyesno(
                        "Upozorenje",
                        "Neki sheetovi nemaju nijednu stavku:\n"
                        + ", ".join(empty_sheets)
                        + "\n\nZelis li preskociti ovaj fajl?",
                    ):
                        result["ok"] = False
                        result["reason"] = "empty_sheet"
                        top.destroy()
                        return
                    active_sheet.set(empty_sheets[0])
                    return
            missing_nab = sum(
                1 for r in combined_rows if parse_float_safe(r.get("nab_unit")) <= 0
            )
            missing_mpc_wo = sum(
                1
                for r in combined_rows
                if parse_float_safe(r.get("mpc_wo_vat_unit")) <= 0
            )
            missing_mpc_w = sum(
                1
                for r in combined_rows
                if parse_float_safe(r.get("mpc_w_vat_unit")) <= 0
            )
            missing_qty = sum(
                1 for r in combined_rows if parse_float_safe(r.get("qty")) <= 0
            )
            if missing_nab or missing_mpc_wo or missing_mpc_w or missing_qty:
                if messagebox.askyesno(
                    "Greska",
                    "Nisu ucitane sve potrebne vrijednosti za potvrdu:\n"
                    f"Nabavna bez vrijednosti: {missing_nab}\n"
                    f"MP bez PDV bez vrijednosti: {missing_mpc_wo}\n"
                    f"MP s PDV bez vrijednosti: {missing_mpc_w}\n"
                    f"Kolicina bez vrijednosti: {missing_qty}\n\n"
                    "Zelis li preskociti ovaj fajl?",
                ):
                    result["ok"] = False
                    result["reason"] = "missing_values"
                    top.destroy()
                    return
                return
            bad_nab = 0
            bad_vat = 0
            for r in combined_rows[: min(total_rows, 50)]:
                nab = parse_float_safe(r.get("nab_unit"))
                mpc_wo = parse_float_safe(r.get("mpc_wo_vat_unit"))
                vat = parse_float_safe(r.get("vat_rate"))
                if mpc_wo > 0 and nab > mpc_wo:
                    bad_nab += 1
                if vat > 0 and not (16 <= vat <= 18):
                    bad_vat += 1
            sample = min(total_rows, 50) if total_rows else 0
            if sample:
                nab_pct = bad_nab / sample
                vat_pct = bad_vat / sample
                if nab_pct > 0.3 or vat_pct > 0.3:
                    if not messagebox.askyesno(
                        "Upozorenje",
                        "Uoceni su sumnjivi podaci u previewu "
                        f"(nabavna>prodajna: {nab_pct:.0%}, PDV!=17%: {vat_pct:.0%}).\n"
                        "Zelis li ipak potvrditi?",
                    ):
                        return
            if per_sheet_counts:
                summary = "\n".join(
                    f"{name}: {count} stavki"
                    for name, count in per_sheet_counts
                )
                if not messagebox.askyesno(
                    "Potvrda sheetova",
                    "Pregledani su svi sheetovi. Zabiljezene stavke po sheetu:\n"
                    f"{summary}\n\nZelis li potvrditi ovu kalkulaciju?",
                ):
                    return
            save_settings()
            result["ok"] = True
            result["rows"] = combined_rows
            if override_manual_var.get():
                result["override_skus"] = update_manual_overlap(combined_rows)
            top.destroy()

        def on_skip():
            top.destroy()

        btn_skip = ctk.CTkButton(
            btn_frame,
            text="Preskoci",
            fg_color="#666666",
            hover_color="#555555",
            command=on_skip,
        )
        btn_skip.pack(side="right", padx=10, pady=8)

        btn_confirm = ctk.CTkButton(
            btn_frame,
            text="Potvrdi",
            fg_color=ROSEGOLD,
            hover_color=ROSEGOLD_DARK,
            command=on_confirm,
        )
        btn_confirm.pack(side="right", padx=10, pady=8)
        btn_confirm_ref["btn"] = btn_confirm
        update_confirm_label()

        refresh_preview()
        if len(current_rows) <= 20:
            show_all_btn.configure(state="disabled")
            populate(None)
        else:
            populate(20)

        top.protocol("WM_DELETE_WINDOW", on_skip)
        top.wait_window()
        return result

    def refresh_kalk_prices():
        global KALK_NABAVNE_BY_SKU, KALK_PROCESSED_FILES
        if not kalkulacije_folder or not os.path.isdir(kalkulacije_folder):
            messagebox.showerror("Greska", "Odaberi validan folder kalkulacija.")
            return
        all_files = find_excel_files(kalkulacije_folder)
        candidates = []
        log_entries = []
        for path in all_files:
            try:
                mtime = os.path.getmtime(path)
                size = os.path.getsize(path)
            except Exception:
                mtime = None
                size = None
            prev = KALK_PROCESSED_FILES.get(path)
            if (
                isinstance(prev, dict)
                and prev.get("mtime") == mtime
                and prev.get("size") == size
            ):
                log_entries.append(
                    f"{datetime.now().isoformat(timespec='seconds')} SKIP unchanged {path}"
                )
                continue
            candidates.append(path)

        def _extract_seq_num(p):
            base = os.path.basename(p)
            m = re.search(r"FEMMA_(\\d{2})-160-(\\d{6})", base, re.IGNORECASE)
            if not m:
                return None, None
            try:
                return m.group(1), int(m.group(2))
            except Exception:
                return None, None

        seq_by_year = {}
        for p in all_files:
            year, seq = _extract_seq_num(p)
            if year is None or seq is None:
                continue
            seq_by_year.setdefault(year, []).append(seq)

        missing_chunks = []
        for year, nums in sorted(seq_by_year.items()):
            nums = sorted(set(nums))
            missing = []
            for a, b in zip(nums, nums[1:]):
                if b > a + 1:
                    for n in range(a + 1, b):
                        missing.append(n)
            if missing:
                preview = ", ".join(f"{n:06d}" for n in missing[:5])
                tail = " ..." if len(missing) > 5 else ""
                missing_chunks.append(f"{year}: {preview}{tail}")

        if missing_chunks:
            messagebox.showwarning(
                "Upozorenje",
                "Nedostaju kalkulacije u nizu (po godini):\n"
                + "\\n".join(missing_chunks[:10])
                + ("\\n..." if len(missing_chunks) > 10 else "")
                + "\\nProvjerite da li fajlovi postoje u folderu.",
            )

        if not candidates:
            messagebox.showinfo(
                "Info",
                "Nema novih kalkulacija za obradu.",
            )
            return

        def review_cb(path, rows, initial_error=None):
            result = show_kalk_review(path, rows, initial_error=initial_error)
            status = "ACCEPT" if result.get("ok") else "SKIP"
            reason = result.get("reason", "")
            if reason:
                reason = f" reason={reason}"
            log_entries.append(
                f"{datetime.now().isoformat(timespec='seconds')} {status} {path}{reason}"
            )
            return result

        kalk_cache = load_kalk_file_cache()
        kalk_cache = {p: v for p, v in kalk_cache.items() if os.path.exists(p)}
        if os.path.exists(KALK_NABAVNE_PATH):
            backup_kalk_db()
        accepted_files = 0
        total_rows = 0
        qty_found = 0
        history = []
        index = 0

        def rebuild_from_cache():
            sku_data = _merge_kalk_file_cache(kalk_cache)
            _kalk_finalize_and_write(sku_data, KALK_NABAVNE_PATH)
            ensure_kalk_nabavne_loaded(kalkulacije_folder)

        while index < len(candidates):
            path = candidates[index]
            log_len_before = len(log_entries)
            try:
                ok, msg, rows = parse_calc_excel_to_rows(path)
            except Exception as e:
                ok = False
                msg = f"read_error:{e}"
                rows = []
            if not ok:
                rows = rows or []
            review_result = review_cb(path, rows, initial_error=msg if not ok else None)
            if isinstance(review_result, dict) and review_result.get("back"):
                if history:
                    prev = history.pop()
                    if prev.get("cache_changed"):
                        prev_entry = prev.get("prev_cache_entry")
                        if prev_entry is None:
                            kalk_cache.pop(prev.get("path"), None)
                        else:
                            kalk_cache[prev.get("path")] = prev_entry
                        save_kalk_file_cache(kalk_cache)
                        prev_proc = prev.get("prev_processed_entry")
                        if prev_proc is None:
                            KALK_PROCESSED_FILES.pop(prev.get("path"), None)
                        else:
                            KALK_PROCESSED_FILES[prev.get("path")] = prev_proc
                        save_settings()
                        rebuild_from_cache()
                        if prev.get("accepted"):
                            accepted_files = max(0, accepted_files - 1)
                            total_rows = max(0, total_rows - prev.get("file_total_rows", 0))
                            qty_found = max(0, qty_found - prev.get("file_qty_found", 0))
                    log_entries[:] = log_entries[: prev.get("log_len_before", 0)]
                    index = prev.get("index", max(index - 1, 0))
                else:
                    messagebox.showinfo("Info", "Nema prethodne kalkulacije.")
                continue
            accepted = (
                bool(review_result.get("ok"))
                if isinstance(review_result, dict)
                else bool(review_result)
            )
            reject_reason = ""
            reason_map = {
                "empty_sheet": "Prazan sheet",
                "missing_values": "Nedostaju obavezne vrijednosti",
                "missing_sku": "Nedostaju SKU u bazi",
                "user_skip": "Rucno preskoceno",
            }
            if isinstance(review_result, dict):
                reject_reason = str(review_result.get("reason") or "")
            if reject_reason:
                reject_reason = reason_map.get(reject_reason, reject_reason)
            prev_cache_entry = copy.deepcopy(kalk_cache.get(path))
            prev_processed_entry = copy.deepcopy(KALK_PROCESSED_FILES.get(path))
            file_total_rows = 0
            file_qty_found = 0
            cache_changed = False
            history.append(
                {
                    "path": path,
                    "accepted": accepted,
                    "cache_changed": False,
                    "prev_cache_entry": prev_cache_entry,
                    "prev_processed_entry": prev_processed_entry,
                    "log_len_before": log_len_before,
                    "index": index,
                    "file_total_rows": 0,
                    "file_qty_found": 0,
                }
            )
            if accepted:
                rows_final = rows
                override_skus = []
                if isinstance(review_result, dict):
                    rows_final = list(review_result.get("rows", rows))
                    override_skus = review_result.get("override_skus", [])
                file_items = _build_kalk_file_items(rows_final, os.path.basename(path))
                try:
                    mtime = os.path.getmtime(path)
                    size = os.path.getsize(path)
                except Exception:
                    mtime = None
                    size = None
                kalk_cache[path] = {
                    "mtime": mtime,
                    "size": size,
                    "items": file_items,
                }
                cache_changed = True
                save_kalk_file_cache(kalk_cache)
                rebuild_from_cache()
                data_items = load_kalk_json(KALK_NABAVNE_PATH)
                missing = []
                for r in rows_final:
                    sku = str(r.get("sku") or "").strip().upper()
                    if not sku:
                        continue
                    if parse_float_safe(r.get("qty")) <= 0:
                        continue
                    if sku not in data_items:
                        missing.append(sku)
                if missing:
                    if prev_cache_entry is None:
                        kalk_cache.pop(path, None)
                    else:
                        kalk_cache[path] = prev_cache_entry
                    save_kalk_file_cache(kalk_cache)
                    if prev_processed_entry is None:
                        KALK_PROCESSED_FILES.pop(path, None)
                    else:
                        KALK_PROCESSED_FILES[path] = prev_processed_entry
                    save_settings()
                    rebuild_from_cache()
                    messagebox.showerror(
                        "Greska",
                        "Kalkulacija nije upisana (nedostaju SKU).\n"
                        "Fajl je preskocen i bice ponudjen ponovo.",
                    )
                    reject_reason = reason_map["missing_sku"]
                    accepted = False
                else:
                    if override_skus:
                        removed = 0
                        for sku in override_skus:
                            if sku in PRICE_BY_SKU:
                                del PRICE_BY_SKU[sku]
                                removed += 1
                        if removed:
                            save_settings()
                    try:
                        KALK_PROCESSED_FILES[path] = {
                            "mtime": os.path.getmtime(path),
                            "size": os.path.getsize(path),
                        }
                    except Exception:
                        pass
                    save_settings()
                    accepted_files += 1
                    for row in rows_final:
                        sku = (row.get("sku") or "").strip()
                        if not sku:
                            continue
                        file_total_rows += 1
                        qty_val = parse_float_safe(row.get("qty"))
                        if qty_val > 0:
                            file_qty_found += 1
                    total_rows += file_total_rows
                    qty_found += file_qty_found
                    cache_changed = True
            else:
                log_entries.append(
                    f"{datetime.now().isoformat(timespec='seconds')} SKIP {path} reason=user_skip"
                )
                if path in kalk_cache:
                    kalk_cache.pop(path, None)
                    save_kalk_file_cache(kalk_cache)
                    rebuild_from_cache()
                    cache_changed = True
                KALK_PROCESSED_FILES.pop(path, None)
                save_settings()
                if not reject_reason:
                    reject_reason = reason_map["user_skip"]
                if reject_reason:
                    messagebox.showinfo(
                        "Preskoceno",
                        f"Kalkulacija je preskocena: {os.path.basename(path)}\n"
                        f"Razlog: {reject_reason}",
                    )
            history[-1]["accepted"] = accepted
            history[-1]["cache_changed"] = cache_changed
            history[-1]["file_total_rows"] = file_total_rows
            history[-1]["file_qty_found"] = file_qty_found
            index += 1
        count = len(KALK_NABAVNE_BY_SKU)
        qty_missing = total_rows - qty_found
        pct = (qty_found / total_rows * 100.0) if total_rows else 0.0
        messagebox.showinfo(
            "Uspjeh",
            "Azurirano nabavnih cijena iz kalkulacija: "
            f"{count} SKU.\n"
            f"Fajlovi potvrdjeni: {accepted_files}/{len(candidates)}.\n"
            f"Kolicine pronadjene: {qty_found}/{total_rows} "
            f"({pct:.1f}%). Nedostaje: {qty_missing}.",
        )
        if log_entries:
            try:
                log_path = os.path.join(app_base_dir(), "kalk_import_log.txt")
                with open(log_path, "a", encoding="utf-8") as f:
                    for line in log_entries:
                        f.write(line + "\n")
            except Exception:
                pass

    def reset_kalk_history():
        global KALK_PROCESSED_FILES
        if not KALK_PROCESSED_FILES:
            messagebox.showinfo("Info", "Historija uvoza je vec prazna.")
            return
        if not messagebox.askyesno(
            "Potvrda",
            "Da li zelite obrisati historiju uvoza kalkulacija?\n"
            "Naredno osvjezavanje ce obraditi sve fajlove.",
        ):
            return
        KALK_PROCESSED_FILES = {}
        save_settings()
        try:
            if os.path.exists(KALK_FILE_CACHE_PATH):
                os.remove(KALK_FILE_CACHE_PATH)
        except Exception:
            pass
        messagebox.showinfo("Uspjeh", "Historija uvoza je obrisana.")

    def reset_kalk_all():
        global KALK_PROCESSED_FILES
        if not messagebox.askyesno(
            "Potvrda",
            "Ovo ce obrisati historiju uvoza i sve podatke iz kalkulacija.\n"
            "Nastavak?",
        ):
            return
        KALK_PROCESSED_FILES = {}
        save_settings()
        try:
            if os.path.exists(KALK_NABAVNE_PATH):
                os.remove(KALK_NABAVNE_PATH)
            if os.path.exists(KALK_FILE_CACHE_PATH):
                os.remove(KALK_FILE_CACHE_PATH)
        except Exception:
            messagebox.showerror("Greska", "Ne mogu obrisati kalkulacije JSON.")
            return
        ensure_kalk_nabavne_loaded("")
        messagebox.showinfo(
            "Uspjeh",
            "Obrisana historija uvoza i kalkulacije JSON.",
        )

    btn_kalk_folder = ctk.CTkButton(
        kalk_frame,
        text="Odaberi folder",
        fg_color="#2b8a3e",
        hover_color="#237133",
        command=choose_kalk_folder,
    )
    btn_kalk_folder.grid(row=2, column=0, padx=10, pady=6, sticky="w")

    btn_kalk_refresh = ctk.CTkButton(
        kalk_frame,
        text="Osvjezi cijene iz kalkulacija",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=refresh_kalk_prices,
    )
    btn_kalk_refresh.grid(row=2, column=1, padx=10, pady=6, sticky="w")

    btn_kalk_reset = ctk.CTkButton(
        kalk_frame,
        text="Resetuj historiju uvoza",
        fg_color="#555555",
        hover_color="#444444",
        command=reset_kalk_history,
    )
    btn_kalk_reset.grid(row=2, column=2, padx=10, pady=6, sticky="w")

    btn_kalk_reset_all = ctk.CTkButton(
        kalk_frame,
        text="Resetuj sve kalkulacije",
        fg_color="#7a2f2f",
        hover_color="#662727",
        command=reset_kalk_all,
    )
    btn_kalk_reset_all.grid(row=2, column=3, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(
        kalk_frame,
        text="* Na startu aplikacije se automatski ucitava nabavnecijene_kalkulacije.json, ako postoji.",
        text_color=TEXT_MUTED,
    ).grid(row=3, column=0, columnspan=3, sticky="w", padx=10, pady=(6, 10))

    # Rucni unos prosjecne nabavne cijene
    ctk.CTkLabel(
        kalk_frame,
        text="Rucni unos prosjecne nabavne cijene (KM)",
        font=ctk.CTkFont(size=13, weight="bold"),
    ).grid(row=4, column=0, columnspan=3, sticky="w", padx=10, pady=(10, 5))

    ctk.CTkLabel(kalk_frame, text="SKU:").grid(
        row=5, column=0, padx=10, pady=6, sticky="e"
    )
    kalk_sku_entry = ctk.CTkEntry(kalk_frame, width=160)
    kalk_sku_entry.grid(row=5, column=1, padx=10, pady=6, sticky="w")

    ctk.CTkLabel(kalk_frame, text="Avg nabavna (KM):").grid(
        row=5, column=2, padx=10, pady=6, sticky="e"
    )
    kalk_avg_entry = ctk.CTkEntry(kalk_frame, width=120)
    kalk_avg_entry.grid(row=5, column=3, padx=10, pady=6, sticky="w")

    def save_kalk_manual_avg():
        sku = kalk_sku_entry.get().strip().upper()
        if not sku:
            messagebox.showerror("Greska", "Unesi SKU.")
            return
        try:
            avg_val = float(kalk_avg_entry.get().strip())
        except Exception:
            messagebox.showerror("Greska", "Unesi validnu cijenu.")
            return
        data = load_kalk_json(KALK_NABAVNE_PATH)
        if not isinstance(data, dict):
            data = {}
        item = data.get(sku, {})
        if not isinstance(item, dict):
            item = {}
        item["avg_nabavna"] = float(avg_val)
        data[sku] = item
        out = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source": "FEMMA12_MANUAL",
            "items": data,
        }
        with open(KALK_NABAVNE_PATH, "w", encoding="utf-8") as f_json:
            json.dump(out, f_json, ensure_ascii=False, indent=2)
        ensure_kalk_nabavne_loaded(kalkulacije_folder)
        messagebox.showinfo("Uspjeh", f"Sacuvano: {sku} = {avg_val:.4f} KM")

    btn_kalk_save_manual = ctk.CTkButton(
        kalk_frame,
        text="Sacuvaj",
        fg_color=ROSEGOLD,
        hover_color=ROSEGOLD_DARK,
        command=save_kalk_manual_avg,
    )
    btn_kalk_save_manual.grid(row=6, column=3, padx=10, pady=(4, 10), sticky="e")
    ctk.CTkLabel(
        tab_excel_out,
        text="Excel izlaz podesavanja ce biti dodana kasnije.",
        text_color=TEXT_MUTED,
    ).pack(pady=20)
    ctk.CTkLabel(
        tab_excel_in,
        text="Excel ulaz podesavanja ce biti dodana kasnije.",
        text_color=TEXT_MUTED,
    ).pack(pady=20)

    refresh_category_options()


# Bind into class
ctk.CTk._build_proc_page = build_proc_page
ctk.CTk._build_settings_page = build_settings_page
# -------------------------------------------------------------------------
# GLAVNA KLASA APLIKACIJE — FEMMA 12.0
# -------------------------------------------------------------------------


class FemmaApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # prozor
        self.title("FEMMA 14.0 – BiH Edition (Modern UI)")
        self.geometry("1200x900")
        self.minsize(1100, 800)

        def _on_close():
            if not messagebox.askyesno(
                "Potvrda",
                "Da li zelite zavrsiti s radom?",
            ):
                return
            try:
                if os.path.exists(META_TMP_PATH):
                    os.remove(META_TMP_PATH)
            except Exception:
                pass
            self.destroy()

        self.protocol("WM_DELETE_WINDOW", _on_close)

        # Tri glavne stranice
        self.page_pantheon = ctk.CTkFrame(self)
        self.page_sales = ctk.CTkFrame(self)
        self.page_promet = ctk.CTkFrame(self)
        self.page_proc = ctk.CTkFrame(self)
        self.page_settings = ctk.CTkFrame(self)

        # Sidebar
        self.build_sidebar()

        # Kreiranje stranica
        self._build_pantheon_page(self.page_pantheon)
        self._build_sales_page(self.page_sales)
        self._build_promet_page(self.page_promet)
        self._build_proc_page(self.page_proc)
        self._build_settings_page(self.page_settings)

        # Default prikaz
        self.show_pantheon()


# -------------------------------------------------------------------------
# MAIN – POKRETANJE APLIKACIJE
# -------------------------------------------------------------------------

if __name__ == "__main__":
    app = FemmaApp()
    app.mainloop()
